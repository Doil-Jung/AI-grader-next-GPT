import io
import json
import zipfile

from openpyxl import load_workbook

import app as app_module
import models.project as project_model
import services.file_manager as file_manager
import services.grading as grading
import services.overview as overview
import services.pdf_splitter as splitter
import services.review as review
import services.submissions as submissions
from models.project import Category, Criterion, StudentRecord
from services.feedback_export import (
    BUNDLE_FILENAMES,
    create_feedback_bundle,
    preview_feedback_export,
)


def configure_temp_projects(tmp_path, monkeypatch):
    projects = tmp_path / "projects"
    for module in (
        app_module,
        project_model,
        file_manager,
        grading,
        overview,
        splitter,
        submissions,
    ):
        monkeypatch.setattr(module, "PROJECTS_DIR", projects)
    projects.mkdir()
    return projects


def make_sensitive_project(projects):
    config = project_model.create_project("민감 프로젝트", workflow_type="report")
    config.description = "홍길동 교사의 내부 평가"
    config.setup.target = "충북과학고등학교 2학년 3반"
    config.setup.assessment_name = "비공개 수행평가"
    config.categories = [Category(
        name="내용",
        criteria=[Criterion(
            id="c1",
            name="정확성",
            description="핵심 내용의 정확성",
            scale=[10, 8, 6, 4, 2],
            scale_labels=["매우 우수", "우수", "보통", "미흡", "매우 미흡"],
        )],
    )]
    config.submissions.students = [
        StudentRecord(
            number=17,
            name="홍길동",
            grade="2",
            class_name="3",
            student_id="20317",
        ),
        StudentRecord(
            number=24,
            name="김영희",
            grade="2",
            class_name="3",
            student_id="20324",
        ),
    ]
    project_model.save_project(config)
    materials = projects / config.id / "materials"
    materials.mkdir(parents=True, exist_ok=True)
    (materials / "17. 홍길동_비밀답안.pdf").write_bytes(b"%PDF-sensitive")
    (materials / "24. 김영희_비밀답안.pdf").write_bytes(b"%PDF-sensitive")
    return config


def save_sensitive_round(config, round_id, number, name, score):
    grading.save_result(
        config.id,
        round_id,
        {
            "team_number": number,
            "team_name": name,
            "c1": score,
            "c1_answer_summary": (
                f"{name} 학생의 2학년 3반 답안. "
                "teacher@example.com, 010-1234-5678, "
                "충청북도 청주시 흥덕구 과학로 1, 123-456-789012"
            ),
            "c1_reason": (
                r"충북과학고등학교 C:\Users\doilm\Documents\답안.pdf "
                "AIza123456789012345678901234567890"
            ),
            "c1_confidence": 0.8,
            "c1_review_required": False,
            "total_score": score,
        },
        number,
    )
    grading.save_round_metadata(
        config.id,
        round_id,
        {
            "status": "completed_with_errors",
            "execution_context": {
                "provider": "openai_api",
                "model": "test-model",
                "criteria_version": 3,
                "criteria_status": "approved",
            },
            "target_team_numbers": [17, 24],
            "started_at": "2026-07-17T10:00:00+09:00",
            "finished_at": "2026-07-17T10:01:00+09:00",
            "attempts": [{
                "status": "completed",
                "requested_team_numbers": [17, 24],
                "success_count": 1,
                "failure_count": 1,
            }],
            "failures": [{
                "team_number": 24,
                "category": "quota",
                "retryable": False,
                "error": "김영희 teacher@example.com 쿼터 오류",
                "action": r"C:\Users\doilm\secret 경로 확인",
            }],
            "request_plan": {
                "expected_requests": 2,
                "estimated_minutes_range": [1, 4],
                "estimated_cost_range_krw": [40, 200],
            },
        },
    )


def bundle_text(path):
    fragments = []
    with zipfile.ZipFile(path) as bundle:
        for name in bundle.namelist():
            payload = bundle.read(name)
            if name.endswith(".xlsx"):
                with zipfile.ZipFile(io.BytesIO(payload)) as workbook:
                    for member in workbook.namelist():
                        if member.endswith((".xml", ".rels")):
                            fragments.append(
                                workbook.read(member).decode("utf-8", errors="replace")
                            )
            else:
                fragments.append(payload.decode("utf-8-sig", errors="replace"))
    return "\n".join(fragments)


def test_preview_writes_nothing_and_reports_missing_manual(tmp_path, monkeypatch):
    projects = configure_temp_projects(tmp_path, monkeypatch)
    config = make_sensitive_project(projects)
    save_sensitive_round(config, 1, 17, "홍길동", 8)

    preview = preview_feedback_export(config)

    assert preview["student_count"] == 2
    assert preview["round_count"] == 1
    assert preview["manual_measurement"] == "측정 불가"
    assert preview["final_measurement"] == "측정 불가"
    assert preview["originals_included"] is False
    assert not (projects / config.id / "exports").exists()
    assert not (projects / config.id / "private").exists()


def test_feedback_bundle_is_allowlisted_anonymous_and_consistent(
    tmp_path, monkeypatch
):
    projects = configure_temp_projects(tmp_path, monkeypatch)
    config = make_sensitive_project(projects)
    save_sensitive_round(config, 1, 17, "홍길동", 8)
    save_sensitive_round(config, 2, 17, "홍길동", 9)
    review.set_manual_score(
        config, 17, total_score=8, item_scores={"c1": 8}
    )
    review.approve_final_score(
        config,
        17,
        final_total_score=8,
        item_scores={"c1": 8},
        teacher_note="홍길동 확인",
        decision_source="manual",
        basis_rounds=[1, 2],
        participant_numbers=[17, 24],
    )

    path, metadata = create_feedback_bundle(config)

    assert path.exists()
    with zipfile.ZipFile(path) as bundle:
        assert set(bundle.namelist()) == BUNDLE_FILENAMES
        assert not any(name.lower().endswith(".pdf") for name in bundle.namelist())
        records = [
            json.loads(line)
            for line in bundle.read("채점데이터.jsonl").decode("utf-8").splitlines()
            if line
        ]
        student_records = [
            record for record in records if record["round_id"] in {1, 2}
        ]
        assert len({record["virtual_student_id"] for record in student_records}) == 1
        assert {record["round_id"] for record in student_records} == {1, 2}
        assert all("team_number" not in record for record in records)
        assert all("team_name" not in record for record in records)
        workbook = load_workbook(
            io.BytesIO(bundle.read("성능요약.xlsx")), data_only=True
        )
        assert workbook.sheetnames == ["성능 요약", "학생별", "항목별", "회차별"]

    combined = bundle_text(path)
    for secret in (
        "홍길동",
        "김영희",
        "20317",
        "20324",
        "충북과학고등학교",
        "teacher@example.com",
        "010-1234-5678",
        "충청북도 청주시 흥덕구 과학로 1",
        "123-456-789012",
        "2학년",
        "3반",
        r"C:\Users\doilm",
        "비밀답안",
        "AIza123456789012345678901234567890",
    ):
        assert secret not in combined

    mapping_path = metadata["mapping_path"]
    assert mapping_path.exists()
    assert mapping_path.parent.name == "feedback_mappings"
    mapping = json.loads(mapping_path.read_text(encoding="utf-8"))
    assert any(student["actual_name"] == "홍길동" for student in mapping["students"])
    assert mapping_path.name not in zipfile.ZipFile(path).namelist()


def test_feedback_export_api_downloads_zip_and_keeps_mapping_local(
    tmp_path, monkeypatch
):
    projects = configure_temp_projects(tmp_path, monkeypatch)
    config = make_sensitive_project(projects)
    save_sensitive_round(config, 1, 17, "홍길동", 8)
    client = app_module.app.test_client()

    preview = client.get(
        f"/api/projects/{config.id}/feedback-export/preview"
    )
    assert preview.status_code == 200
    assert preview.get_json()["originals_included"] is False

    response = client.post(
        f"/api/projects/{config.id}/feedback-export"
    )
    assert response.status_code == 200
    assert response.mimetype == "application/zip"
    assert response.headers["X-AI-Grader-Mapping-Saved"] == "true"
    with zipfile.ZipFile(io.BytesIO(response.data)) as bundle:
        assert set(bundle.namelist()) == BUNDLE_FILENAMES
    mappings = list(
        (projects / config.id / "private" / "feedback_mappings").glob("*.json")
    )
    assert len(mappings) == 1
