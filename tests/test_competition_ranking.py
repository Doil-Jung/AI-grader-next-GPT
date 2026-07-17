import io

from openpyxl import load_workbook

import app as app_module
import models.project as project_model
import services.file_manager as file_manager
import services.grading as grading
import services.overview as overview
import services.pdf_splitter as splitter
import services.submissions as submissions
from models.project import Category, Criterion
from services.competition import (
    approve_competition_plan,
    build_competition_plan,
    build_competition_workbook,
    competition_state_view,
    load_competition_state,
)


def configure_temp_projects(tmp_path, monkeypatch):
    projects = tmp_path / "projects"
    for module in (
        app_module,
        project_model,
        file_manager,
        grading,
        overview,
        splitter,
        submissions,
    ):
        monkeypatch.setattr(module, "PROJECTS_DIR", projects)
    projects.mkdir()
    return projects


def make_competition():
    config = project_model.create_project(
        "탐구대회", workflow_type="competition"
    )
    config.categories = [
        Category("내용", [
            Criterion(
                "c1", "정확성", "정확성",
                [10, 8, 6], ["우수", "보통", "미흡"],
            ),
            Criterion(
                "c2", "논리성", "논리성",
                [10, 8, 6], ["우수", "보통", "미흡"],
            ),
        ])
    ]
    project_model.save_project(config)
    for number, name, c1, c2 in (
        (1, "하늘팀", 10, 10),
        (2, "바다팀", 10, 8),
        (3, "별빛팀", 8, 8),
    ):
        grading.save_result(
            config.id,
            1,
            {
                "team_number": number,
                "team_name": name,
                "c1": c1,
                "c2": c2,
                "total_score": c1 + c2,
            },
            number,
        )
    return config


def ordered_teams(*, override=None, reason=""):
    return [
        {
            "team_number": 2,
            "override_score": override,
            "exception_reason": reason,
        },
        {
            "team_number": 1,
            "tie_with_previous": True,
            "override_score": override,
            "exception_reason": reason,
        },
        {"team_number": 3},
    ]


def test_competition_plan_preserves_evaluation_and_supports_ties(
    tmp_path, monkeypatch
):
    configure_temp_projects(tmp_path, monkeypatch)
    config = make_competition()

    plan = build_competition_plan(
        config,
        source_round=1,
        teams=ordered_teams(),
    )

    assert [entry["rank"] for entry in plan["entries"]] == [1, 1, 3]
    assert [entry["evaluation_score"] for entry in plan["entries"]] == [
        18, 20, 16
    ]
    assert plan["entries"][0]["final_score"] == plan["entries"][1]["final_score"]
    assert plan["summary"]["tie_team_count"] == 2
    assert grading.load_completed(config.id, 1)[1]["total_score"] == 20
    assert "final_score" not in grading.load_completed(config.id, 1)[1]


def test_manual_exception_requires_reason_and_keeps_rank_order(
    tmp_path, monkeypatch
):
    configure_temp_projects(tmp_path, monkeypatch)
    config = make_competition()

    try:
        build_competition_plan(
            config,
            source_round=1,
            teams=ordered_teams(override=19),
        )
        assert False, "예외 사유 없는 수동 점수는 거부해야 한다."
    except ValueError as exc:
        assert "예외 사유" in str(exc)

    plan = build_competition_plan(
        config,
        source_round=1,
        teams=ordered_teams(override=19, reason="심사위원 합의"),
    )
    assert [entry["final_score"] for entry in plan["entries"][:2]] == [19, 19]
    assert all(entry["manual_exception"] for entry in plan["entries"][:2])
    assert plan["entries"][0]["exception_reason"] == "심사위원 합의"


def test_approval_history_stale_detection_and_excel(tmp_path, monkeypatch):
    configure_temp_projects(tmp_path, monkeypatch)
    config = make_competition()
    first = approve_competition_plan(
        config,
        source_round=1,
        teams=ordered_teams(),
        allowed_scores=[20, 18, 16, 14, 12],
        approval_note="1차 회의",
    )
    second = approve_competition_plan(
        config,
        source_round=1,
        teams=ordered_teams(override=19, reason="최종 합의"),
        allowed_scores=[20, 18, 16, 14, 12],
        approval_note="2차 회의",
    )

    assert first["version"] == 1
    assert second["version"] == 2
    state = load_competition_state(config.id)
    assert len(state["history"]) == 2
    workbook = build_competition_workbook(
        config, state["current"], state["history"]
    )
    assert workbook.sheetnames == ["최종 순위표", "원 평가점수", "변경 이력"]
    final_headers = [cell.value for cell in workbook["최종 순위표"][1]]
    assert "원 평가점수" in final_headers
    assert "최종 배정점수" in final_headers
    assert competition_state_view(config)["current_stale"] is False

    changed = grading.load_completed(config.id, 1)[1]
    changed["total_score"] = 18
    grading.save_result(config.id, 1, changed, 1)
    assert competition_state_view(config)["current_stale"] is True


def test_competition_api_preview_approve_and_excel(tmp_path, monkeypatch):
    configure_temp_projects(tmp_path, monkeypatch)
    config = make_competition()
    client = app_module.app.test_client()
    payload = {
        "source_round": 1,
        "teams": ordered_teams(),
        "allowed_scores": [20, 18, 16, 14, 12],
    }

    preview = client.post(
        f"/api/projects/{config.id}/competition-ranking/preview",
        json=payload,
    )
    assert preview.status_code == 200
    assert preview.get_json()["entries"][1]["rank"] == 1

    approved = client.post(
        f"/api/projects/{config.id}/competition-ranking/approve",
        json={**payload, "approval_note": "확정"},
    )
    assert approved.status_code == 200
    state = client.get(
        f"/api/projects/{config.id}/competition-ranking"
    ).get_json()
    assert state["current"]["version"] == 1

    exported = client.get(
        f"/api/projects/{config.id}/competition-ranking/excel"
    )
    assert exported.status_code == 200
    workbook = load_workbook(io.BytesIO(exported.data), data_only=True)
    assert workbook.sheetnames == ["최종 순위표", "원 평가점수", "변경 이력"]
