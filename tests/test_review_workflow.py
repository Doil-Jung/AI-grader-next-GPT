import io
import json

from openpyxl import load_workbook
import app as app_module
import models.project as project_model
import services.file_manager as file_manager
import services.grading as grading
import services.overview as overview
import services.pdf_splitter as splitter
import services.review as review
import services.submissions as submissions
from models.project import Category, Criterion, ExamQuestion, StudentRecord


def configure_temp_projects(tmp_path, monkeypatch):
    projects = tmp_path / "projects"
    for module in (
        app_module,
        project_model,
        file_manager,
        grading,
        overview,
        splitter,
        submissions,
    ):
        monkeypatch.setattr(module, "PROJECTS_DIR", projects)
    projects.mkdir()
    return projects


def make_report_project(projects, name="검토 흐름"):
    config = project_model.create_project(name, workflow_type="report")
    config.categories = [Category(
        name="내용",
        criteria=[Criterion(
            id="c1",
            name="정확성",
            description="핵심 내용의 정확성",
            scale=[10, 8, 6, 4, 2],
            scale_labels=["매우 우수", "우수", "보통", "미흡", "매우 미흡"],
        )],
    )]
    project_model.save_project(config)
    materials = projects / config.id / "materials"
    (materials / "1. 가.pdf").write_bytes(b"%PDF-test")
    (materials / "2. 나.pdf").write_bytes(b"%PDF-test")
    return config


def save_round(config, round_id, team_number, name, score, **extra):
    grading.save_result(
        config.id,
        round_id,
        {
            "team_number": team_number,
            "team_name": name,
            "c1": score,
            "c1_reason": "근거",
            "total_score": score,
            **extra,
        },
        team_number,
    )


def test_ai_manual_and_final_scores_remain_separate(tmp_path, monkeypatch):
    projects = configure_temp_projects(tmp_path, monkeypatch)
    config = make_report_project(projects)
    save_round(config, 1, 1, "가", 4)
    save_round(config, 2, 1, "가", 8)
    review.set_manual_score(config, 1, total_score=5, item_scores={"c1": 5})

    dashboard = review.build_review_dashboard(
        config,
        round_ids=[1, 2],
        participant_numbers=[1],
        std_threshold=1,
        range_threshold=2,
        manual_diff_threshold=1,
    )
    student = dashboard["students"][0]
    assert student["scores_by_round"] == {1: 4.0, 2: 8.0}
    assert student["ai_average"] == 6
    assert student["ai_median"] == 6
    assert student["manual_score"] == 5
    assert student["ai_manual_difference"] == 1
    assert {reason["code"] for reason in student["review_reasons"]} >= {
        "high_std_dev", "high_range", "manual_difference",
    }

    decision = review.approve_final_score(
        config,
        1,
        final_total_score=5,
        item_scores={"c1": 5},
        teacher_note="수동 채점 확인",
        decision_source="manual",
        basis_rounds=[1, 2],
        participant_numbers=[1],
    )
    assert decision["status"] == "approved"
    assert decision["total_score"] == 5

    # 확정 뒤 수동 점수가 바뀌어도 확정값은 자동으로 덮어쓰지 않고 재검토로 표시한다.
    review.set_manual_score(config, 1, total_score=6, item_scores={"c1": 6})
    changed = review.build_review_dashboard(
        config,
        round_ids=[1, 2],
        participant_numbers=[1],
    )["students"][0]
    assert changed["decision"]["total_score"] == 5
    assert changed["decision_stale"] is True
    assert "approved_data_changed" in {
        reason["code"] for reason in changed["review_reasons"]
    }
    filtered = review.build_review_dashboard(
        config,
        round_ids=[1],
        participant_numbers=[1],
    )["students"][0]
    assert filtered["decision_stale"] is False

    reopened = review.reopen_final_score(config, 1, reason="수동 점수 변경")
    assert reopened["status"] == "pending"
    state = review.load_review_state(config.id)
    actions = [
        entry["action"]
        for entry in state["students"]["1"]["audit_log"]
    ]
    assert actions == [
        "manual_score_saved",
        "final_score_approved",
        "manual_score_saved",
        "final_score_reopened",
    ]


def test_review_api_and_analysis_do_not_mix_manual_into_ai_average(tmp_path, monkeypatch):
    projects = configure_temp_projects(tmp_path, monkeypatch)
    config = make_report_project(projects)
    save_round(config, 1, 1, "가", 4)
    save_round(config, 2, 1, "가", 8)
    client = app_module.app.test_client()

    manual = client.put(
        f"/api/projects/{config.id}/review/1/manual",
        json={"total_score": 9, "item_scores": {"c1": 9}},
    )
    assert manual.status_code == 200

    analysis = client.get(
        f"/api/projects/{config.id}/analysis?rounds=1,2&include_manual=true"
    )
    assert analysis.status_code == 200
    row = analysis.get_json()[0]
    assert row["average"] == 6
    assert row["manual_score"] == 9
    assert row["ai_manual_difference"] == -3

    invalid = client.post(
        f"/api/projects/{config.id}/review/1/approve",
        json={
            "final_total_score": 7,
            "item_scores": {"c1": 6},
            "decision_source": "custom",
            "basis_rounds": [1, 2],
        },
    )
    assert invalid.status_code == 400
    assert "합계" in invalid.get_json()["error"]

    approved = client.post(
        f"/api/projects/{config.id}/review/1/approve",
        json={
            "final_total_score": 9,
            "item_scores": {"c1": 9},
            "teacher_note": "확인",
            "decision_source": "manual",
            "basis_rounds": [1, 2],
        },
    )
    assert approved.status_code == 200
    summary = client.get(
        f"/api/projects/{config.id}/analysis/summary"
        "?rounds=1,2&include_manual=true"
    )
    assert summary.status_code == 200
    summary_data = summary.get_json()
    assert summary_data["manual_mae"] == 3
    assert summary_data["final_mae"] == 3
    assert summary_data["approved_count"] == 1
    assert summary_data["items"][0]["manual_mae"] == 3
    dashboard = client.get(
        f"/api/projects/{config.id}/review?rounds=1,2"
    ).get_json()
    student = next(
        value for value in dashboard["students"] if value["team_number"] == 1
    )
    assert student["decision"]["status"] == "approved"
    assert student["decision"]["total_score"] == 9

    exported = client.get(
        f"/api/projects/{config.id}/analysis/excel"
        "?rounds=1,2&include_manual=true"
    )
    assert exported.status_code == 200
    workbook = load_workbook(io.BytesIO(exported.data), data_only=True)
    assert workbook.sheetnames == [
        "확정 점수표", "종합 분석", "항목별 비교", "문항별 분석"
    ]
    final_headers = [
        cell.value for cell in workbook["확정 점수표"][1]
    ]
    assert "최종 확정 총점" in final_headers
    assert "확정 정확성" in final_headers
    headers = [
        cell.value for cell in workbook["종합 분석"][1]
    ]
    assert "교사 확정점수" in headers
    assert "차이(AI-수동)" in headers
    detail_headers = [
        cell.value for cell in workbook["항목별 비교"][1]
    ]
    assert "AI 중앙값" in detail_headers
    assert "최종 확정" in detail_headers
    item_headers = [
        cell.value for cell in workbook["문항별 분석"][1]
    ]
    assert "AI-수동 평균 절대차" in item_headers


def test_exam_final_sheet_contains_subquestions_and_parent_total(
    tmp_path, monkeypatch
):
    projects = configure_temp_projects(tmp_path, monkeypatch)
    config = project_model.create_project("서술형 확정표", workflow_type="exam")
    config.exam.questions = [
        ExamQuestion("q1", "1", "공통 지문", 10),
        ExamQuestion(
            "q1a", "1-(1)", "값", 4, parent_id="q1", sub_index=1
        ),
        ExamQuestion(
            "q1b", "1-(2)", "유도", 6, parent_id="q1", sub_index=2
        ),
    ]
    config.submissions.students = [StudentRecord(1, "학생")]
    project_model.save_project(config)
    grading.save_result(
        config.id,
        1,
        {
            "team_number": 1,
            "team_name": "학생",
            "q1a": 3,
            "q1b": 5,
            "total_score": 8,
        },
        1,
    )
    review.approve_final_score(
        config,
        1,
        final_total_score=8,
        item_scores={"q1a": 3, "q1b": 5},
        teacher_note="확인",
        decision_source="ai_suggested",
        basis_rounds=[1],
        participant_numbers=[1],
    )
    client = app_module.app.test_client()

    exported = client.get(
        f"/api/projects/{config.id}/analysis/excel?rounds=1"
    )

    assert exported.status_code == 200
    workbook = load_workbook(io.BytesIO(exported.data), data_only=True)
    sheet = workbook["확정 점수표"]
    headers = [cell.value for cell in sheet[1]]
    assert "확정 1번 합계" in headers
    assert "확정 1-(1)번" in headers
    assert "확정 1-(2)번" in headers
    values = {
        header: sheet.cell(row=2, column=index + 1).value
        for index, header in enumerate(headers)
    }
    assert values["확정 1번 합계"] == 8
    assert values["확정 1-(1)번"] == 3
    assert values["확정 1-(2)번"] == 5


def test_round_adjustment_preserves_original_ai_result_for_report(tmp_path, monkeypatch):
    projects = configure_temp_projects(tmp_path, monkeypatch)
    config = make_report_project(projects)
    save_round(config, 1, 1, "가", 4)
    client = app_module.app.test_client()

    updated = client.put(
        f"/api/projects/{config.id}/result/1?round=1",
        json={"c1": 8, "c1_reason": "교사 판정 보정"},
    )
    assert updated.status_code == 200
    data = updated.get_json()["data"]
    assert data["total_score"] == 8
    assert data["ai_original"]["total_score"] == 4
    assert data["ai_original"]["c1"] == 4
    assert data["audit_log"][-1]["action"] == "round_result_adjusted"

    dashboard = review.build_review_dashboard(
        config,
        round_ids=[1],
        participant_numbers=[1],
    )
    student = dashboard["students"][0]
    assert student["scores_by_round"][1] == 4
    assert student["adjusted_scores_by_round"][1] == 8


def test_review_screen_exposes_score_layers_and_teacher_confirmation():
    client = app_module.app.test_client()
    html = client.get("/").get_data(as_text=True)
    assert 'id="reviewRoundChecks"' in html
    assert 'id="reviewBody"' in html
    assert 'id="reviewDecisionModal"' in html
    assert 'id="reviewManualTotal"' in html
    assert 'id="reviewFinalTotal"' in html
    assert "AI 원점수 · 수동 점수 · 최종 확정 분리" in html
    assert "평균에 포함" not in html


def test_legacy_manual_score_file_is_read_without_rewrite(tmp_path, monkeypatch):
    projects = configure_temp_projects(tmp_path, monkeypatch)
    config = make_report_project(projects)
    path = projects / config.id / "manual_scores.json"
    path.write_text(json.dumps({"1": 7.5}), encoding="utf-8")

    loaded = review.load_manual_scores(config.id)
    assert loaded[1]["total_score"] == 7.5
    assert loaded[1]["source"] == "legacy"
    assert json.loads(path.read_text(encoding="utf-8")) == {"1": 7.5}
