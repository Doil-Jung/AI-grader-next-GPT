"""
범용 AI 채점 시스템 - 메인 애플리케이션
"""
import json
import re
import sys
import os
import threading
import time
import webbrowser
import uuid
from datetime import datetime
from pathlib import Path
from queue import Queue

from flask import Flask, render_template, request, jsonify, Response, send_file

# 경로 설정 (PyInstaller 호환)
if getattr(sys, 'frozen', False):
    # exe 실행 시: 데이터 파일은 _MEIPASS/_internal 안에 번들됨
    BUNDLE_DIR = Path(getattr(sys, '_MEIPASS', Path(sys.executable).parent / '_internal'))
    BASE_DIR = Path(sys.executable).parent
else:
    BUNDLE_DIR = Path(__file__).parent
    BASE_DIR = BUNDLE_DIR

sys.path.insert(0, str(BUNDLE_DIR))

from config import PROJECTS_DIR, AI_MODELS, OPENAI_AI_MODELS, SCALE_PRESETS, API_KEY_FILE
from models.project import (
    ProjectConfig, Criterion, Category, MaterialsConfig,
    ExamConfig, ExamQuestion, ScoringElement, StudentRecord, ScanSplitConfig,
    ProjectSetup, SubmissionLink, WORKFLOW_TYPES, ANSWER_TYPES,
    QUESTION_GRADING_MODES, normalize_exam_questions,
    create_project, save_project, load_project, list_projects, delete_project,
    generate_default_prompt, portable_project_path, resolve_project_path,
)
from models.evaluation import compute_scores, generate_fake_subscores_dynamic
from models.criteria_versions import (
    load_versions, get_version, snapshot as snapshot_criteria,
    approve as approve_criteria, summary as criteria_version_summary,
)
from services.file_manager import find_materials, get_participant_files, save_uploaded_file
from services.grading import (
    grading_state, event_queues, broadcast_event,
    load_completed, save_result, get_latest_round_id,
    grading_worker, get_round_dir,
)
from services.export import generate_excel
from services.pdf_splitter import build_split_plan, split_integrated_pdf, SplitValidationError
from services.providers import get_provider
from services.providers.base import ProviderNeedsUserAction
from services.overview import build_project_overview
from services.submissions import build_submission_status


app = Flask(__name__, template_folder=str(BUNDLE_DIR / "templates"), static_folder=str(BUNDLE_DIR / "static"))
app.config["MAX_CONTENT_LENGTH"] = 500 * 1024 * 1024  # 500MB
app.json.sort_keys = False  # 모델 목록은 권장 순서(최신 기본 모델 우선)를 유지한다.

PROJECTS_DIR.mkdir(parents=True, exist_ok=True)


def _apply_project_setup(config: ProjectConfig, setup_data: dict | None) -> None:
    """프로젝트 생성 마법사의 공통 정보를 안전하게 반영한다."""
    data = setup_data or {}
    participant_mode = data.get("participant_mode", config.setup.participant_mode)
    if participant_mode not in {"individual", "group"}:
        participant_mode = "individual"
    ai_setup_mode = data.get("ai_setup_mode", config.setup.ai_setup_mode)
    if ai_setup_mode not in {"recommended", "advanced"}:
        ai_setup_mode = "recommended"
    config.setup = ProjectSetup(
        target=str(data.get("target", config.setup.target)),
        assessment_name=str(data.get("assessment_name", config.setup.assessment_name)),
        participant_mode=participant_mode,
        expected_count=max(0, int(data.get("expected_count", config.setup.expected_count) or 0)),
        materials_status=str(
            data.get("materials_status", config.setup.materials_status) or "later"
        ),
        ai_setup_mode=ai_setup_mode,
    )


def _parse_students(raw_students: list[dict], *, require_students: bool = False) -> list[StudentRecord]:
    students = []
    for index, item in enumerate(raw_students or [], 1):
        try:
            number = int(item.get("number", index))
        except (TypeError, ValueError) as exc:
            raise ValueError(f"{index}번째 학생 번호가 올바르지 않습니다.") from exc
        if number < 1:
            raise ValueError(f"{index}번째 학생 번호는 1 이상이어야 합니다.")
        students.append(StudentRecord(
            number=number,
            name=str(item.get("name", "")).strip(),
            grade=str(item.get("grade", "")).strip(),
            class_name=str(item.get("class_name", "")).strip(),
            student_id=str(item.get("student_id", "")).strip(),
        ))
    if require_students and not students:
        raise ValueError("학생 명렬을 한 명 이상 입력하세요.")
    if len({student.number for student in students}) != len(students):
        raise ValueError("학생 번호가 중복되었습니다.")
    return students


def _parse_categories(raw_categories: list[dict]) -> list[Category]:
    categories = []
    used_ids = set()
    for category_index, category_data in enumerate(raw_categories or [], 1):
        criteria = []
        for criterion_index, criterion_data in enumerate(
            category_data.get("criteria", []), 1
        ):
            criterion_id = str(
                criterion_data.get("id") or f"c{category_index}_{criterion_index}"
            )
            while criterion_id in used_ids:
                criterion_id = f"{criterion_id}_{criterion_index}"
            used_ids.add(criterion_id)
            scale = [
                int(value)
                for value in criterion_data.get("scale", [5, 4, 3, 2, 1])
            ]
            if not scale:
                scale = [5, 4, 3, 2, 1]
            labels = [
                str(value)
                for value in criterion_data.get("scale_labels", [])
            ]
            if len(labels) != len(scale):
                labels = [f"등급 {index}" for index in range(1, len(scale) + 1)]
            criteria.append(Criterion(
                id=criterion_id,
                name=str(criterion_data.get("name", criterion_id)),
                description=str(
                    criterion_data.get(
                        "description", criterion_data.get("name", criterion_id)
                    )
                ),
                scale=scale,
                scale_labels=labels,
                required_elements=[
                    str(value).strip()
                    for value in criterion_data.get("required_elements", [])
                    if str(value).strip()
                ],
                deduction_rules=[
                    str(value).strip()
                    for value in criterion_data.get("deduction_rules", [])
                    if str(value).strip()
                ],
                exceptions=[
                    str(value).strip()
                    for value in criterion_data.get("exceptions", [])
                    if str(value).strip()
                ],
                feedback_focus=str(criterion_data.get("feedback_focus", "")),
                core_criteria=[
                    str(value).strip()
                    for value in criterion_data.get("core_criteria", [])
                    if str(value).strip()
                ],
            ))
        categories.append(Category(
            name=str(category_data.get("name", f"영역 {category_index}")),
            criteria=criteria,
        ))
    return categories


def _set_roster(config: ProjectConfig, students: list[StudentRecord]) -> None:
    config.submissions.students = list(students)
    if config.project_type == "exam":
        config.exam.students = list(students)


def _apply_submissions_data(config: ProjectConfig, submissions_data: dict | None) -> None:
    data = submissions_data or {}
    if "students" in data:
        _set_roster(config, _parse_students(data.get("students", [])))
    if "manual_links" in data:
        links = []
        for item in data.get("manual_links", []):
            file_path = str(item.get("file_path", "")).strip()
            try:
                student_number = int(item.get("student_number", 0))
            except (TypeError, ValueError):
                continue
            if file_path and student_number > 0:
                links.append(SubmissionLink(
                    file_path=file_path,
                    student_number=student_number,
                ))
        config.submissions.manual_links = links
    if "split_output_dir" in data:
        config.submissions.split_output_dir = str(data.get("split_output_dir", ""))


# ═══════════════════════════════════════════════════════════
# API Key 관리
# ═══════════════════════════════════════════════════════════

def load_api_keys() -> dict:
    if API_KEY_FILE.exists():
        with open(API_KEY_FILE, "r") as f:
            return json.load(f)
    return {}


def save_api_keys(keys: dict):
    with open(API_KEY_FILE, "w") as f:
        json.dump(keys, f, indent=2)


def _parse_exam_questions(question_data: list[dict]) -> list[ExamQuestion]:
    """API/UI에서 받은 문항 JSON을 검증 가능한 데이터 모델로 변환."""
    questions = []
    used_ids = set()
    id_map = {}
    for index, item in enumerate(question_data or [], 1):
        raw_id = str(item.get("id") or f"q{index}")
        question_id = "".join(ch if ch.isalnum() or ch == "_" else "_" for ch in raw_id)
        if not question_id or question_id[0].isdigit():
            question_id = f"q{index}"
        while question_id in used_ids:
            question_id = f"{question_id}_{index}"
        used_ids.add(question_id)
        id_map[raw_id] = question_id
        max_score = max(1, int(item.get("max_score", 1)))
        elements = []
        for raw_element in item.get("scoring_elements", []):
            points = max(0, int(raw_element.get("points", 0)))
            if points:
                elements.append(ScoringElement(
                    description=str(raw_element.get("description", "")).strip(),
                    points=points,
                    required=bool(raw_element.get("required", True)),
                ))
        questions.append(ExamQuestion(
            id=question_id,
            number=str(item.get("number", index)),
            question_text=str(item.get("question_text", "")).strip(),
            max_score=max_score,
            model_answer=str(item.get("model_answer", "")).strip(),
            scoring_elements=elements,
            accepted_answers=[str(v).strip() for v in item.get("accepted_answers", []) if str(v).strip()],
            common_errors=[str(v).strip() for v in item.get("common_errors", []) if str(v).strip()],
            core_criteria=[str(v).strip() for v in item.get("core_criteria", []) if str(v).strip()],
            parent_id=str(item.get("parent_id", "") or ""),
            sub_index=max(0, int(item.get("sub_index", 0) or 0)),
            answer_type=(
                str(item.get("answer_type", "text") or "text")
                if str(item.get("answer_type", "text") or "text") in ANSWER_TYPES
                else "text"
            ),
            grading_mode=(
                str(item.get("grading_mode", "inherit") or "inherit")
                if str(item.get("grading_mode", "inherit") or "inherit")
                in QUESTION_GRADING_MODES
                else "inherit"
            ),
            teacher_notes=str(item.get("teacher_notes", "")).strip(),
        ))
    for question in questions:
        if question.parent_id in id_map:
            question.parent_id = id_map[question.parent_id]
    normalize_exam_questions(questions)
    return questions


def _mark_criteria_changed(config: ProjectConfig, source: str = "manual") -> None:
    """현재 기준이 마지막 스냅샷과 달라졌음을 명시한다."""
    has_criteria = bool(config.categories or config.exam.questions)
    config.criteria_state.active_version = 0
    config.criteria_state.status = (
        "generated" if source not in {"manual", "restore"} else
        ("modified" if has_criteria else "empty")
    )
    config.criteria_state.source = source
    config.criteria_state.updated_at = datetime.now().isoformat(timespec="seconds")


def _criteria_signature(config: ProjectConfig) -> str:
    data = config.to_dict()
    return json.dumps({
        "categories": data.get("categories", []),
        "exam_questions": (data.get("exam") or {}).get("questions", []),
        "exam_grading_mode": (data.get("exam") or {}).get("grading_mode", ""),
        "additional_instructions": (
            data.get("exam") or {}
        ).get("additional_instructions", ""),
        "report_delivery_mode": config.criteria_state.delivery_mode,
        "report_prompt_template": (
            config.prompt_template if config.project_type != "exam" else ""
        ),
    }, ensure_ascii=False, sort_keys=True)


def _apply_exam_data(config: ProjectConfig, exam_data: dict) -> None:
    """부분 업데이트를 허용하면서 시험 전용 설정을 적용."""
    if "questions" in exam_data:
        config.exam.questions = _parse_exam_questions(exam_data.get("questions", []))
    if "students" in exam_data:
        _set_roster(config, _parse_students(exam_data.get("students", [])))
    if "question_source_path" in exam_data:
        config.exam.question_source_path = str(exam_data["question_source_path"])
    if "rubric_source_path" in exam_data:
        config.exam.rubric_source_path = str(exam_data["rubric_source_path"])
    if "source_mode" in exam_data:
        source_mode = str(exam_data["source_mode"] or "auto")
        config.exam.source_mode = source_mode if source_mode in {
            "auto", "question_only", "combined_answers", "answers_only"
        } else "auto"
    if "expected_question_count" in exam_data:
        config.exam.expected_question_count = max(
            0, int(exam_data.get("expected_question_count", 0) or 0)
        )
    if "additional_instructions" in exam_data:
        config.exam.additional_instructions = str(exam_data["additional_instructions"] or "").strip()
    if "grading_mode" in exam_data:
        mode = str(exam_data["grading_mode"] or "").strip()
        if mode in {"autonomous", "core", "strict"}:
            config.exam.grading_mode = mode
    if "scan_split" in exam_data:
        value = exam_data.get("scan_split") or {}
        config.exam.scan_split = ScanSplitConfig(
            source_path=str(value.get("source_path", config.exam.scan_split.source_path)),
            start_page=max(1, int(value.get("start_page", config.exam.scan_split.start_page))),
            pages_per_student=max(0, int(value.get("pages_per_student", config.exam.scan_split.pages_per_student))),
            boundaries=[int(v) for v in value.get("boundaries", config.exam.scan_split.boundaries)],
            completed_at=str(value.get("completed_at", config.exam.scan_split.completed_at)),
        )


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/browse-folder", methods=["POST"])
def api_browse_folder():
    """네이티브 폴더 선택 다이얼로그"""
    import threading
    result = {"path": ""}

    def pick():
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)
            path = filedialog.askdirectory(title="심사자료 폴더 선택")
            root.destroy()
            if path:
                result["path"] = path
        except Exception:
            pass

    t = threading.Thread(target=pick)
    t.start()
    t.join(timeout=120)
    return jsonify(result)


@app.route("/api/browse-files", methods=["POST"])
def api_browse_files():
    """네이티브 파일 선택 다이얼로그 (여러 파일 선택 가능, 추가 모드)"""
    import threading
    result = {"files": [], "project_id": ""}
    
    data = request.json or {}
    project_id = data.get("project_id", "")
    file_types_str = data.get("file_types", "")
    overwrite = data.get("overwrite", False)
    pre_selected = data.get("selected_files", [])
    
    filetypes = [("모든 파일", "*.*")]
    if file_types_str:
        exts = [f"*.{ft.strip().strip('.')}" for ft in file_types_str.split(",") if ft.strip()]
        if exts:
            filetypes.insert(0, ("지원 파일", " ".join(exts)))

    if pre_selected:
        result["files"] = pre_selected
    else:
        def pick():
            try:
                import tkinter as tk
                from tkinter import filedialog
                root = tk.Tk()
                root.withdraw()
                root.attributes("-topmost", True)
                paths = filedialog.askopenfilenames(
                    title="심사자료 파일 추가 (여러 파일 선택 가능)",
                    filetypes=filetypes,
                )
                root.destroy()
                if paths:
                    result["files"] = list(paths)
            except Exception:
                pass

        t = threading.Thread(target=pick)
        t.start()
        t.join(timeout=120)
    
    if result["files"] and project_id:
        import shutil
        materials_dir = PROJECTS_DIR / project_id / "materials"
        materials_dir.mkdir(parents=True, exist_ok=True)
        
        duplicates = []
        if not overwrite:
            for fpath in result["files"]:
                src = Path(fpath)
                dst = materials_dir / src.name
                if dst.exists():
                    duplicates.append(src.name)
        
        if duplicates and not overwrite:
            result["duplicates"] = duplicates
            result["materials_dir"] = str(materials_dir)
            return jsonify(result)
        
        copied = []
        skipped = []
        for fpath in result["files"]:
            src = Path(fpath)
            dst = materials_dir / src.name
            try:
                shutil.copy2(str(src), str(dst))
                copied.append(src.name)
            except Exception as e:
                skipped.append(src.name)
                print(f"파일 복사 실패: {src.name} - {e}")
        result["copied"] = copied
        result["skipped"] = skipped
        result["materials_dir"] = str(materials_dir)
    
    return jsonify(result)


@app.route("/api/open-file", methods=["POST"])
def api_open_file():
    """OS 기본 앱으로 파일 열기"""
    data = request.json or {}
    file_path = data.get("path", "")
    if not file_path:
        return jsonify({"error": "파일 경로가 필요합니다."}), 400
    
    p = Path(file_path)
    if not p.exists():
        return jsonify({"error": f"파일이 존재하지 않습니다: {p.name}"}), 404
    
    try:
        import os, subprocess, sys
        if sys.platform == 'win32':
            os.startfile(str(p))
        elif sys.platform == 'darwin':
            subprocess.Popen(['open', str(p)])
        else:
            subprocess.Popen(['xdg-open', str(p)])
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": f"파일 열기 실패: {str(e)}"}), 500


@app.route("/api/projects/<project_id>/materials/remove", methods=["POST"])
def api_remove_material(project_id):
    """심사자료 파일 제거 (업로드: 삭제, 폴더: 제외 목록에 추가)"""
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    
    data = request.json or {}
    file_path = data.get("path", "")
    if not file_path:
        return jsonify({"error": "파일 경로가 필요합니다."}), 400
    
    p = Path(file_path)
    materials_dir = PROJECTS_DIR / project_id / "materials"
    
    try:
        is_uploaded = materials_dir.exists() and p.resolve().is_relative_to(materials_dir.resolve())
        
        if is_uploaded:
            # 업로드된 파일은 실제 삭제
            if p.exists():
                p.unlink()
            return jsonify({"success": True, "deleted": p.name})
        else:
            # 외부 폴더 파일은 제외 목록에 추가
            if not config.materials.excluded_files:
                config.materials.excluded_files = []
            if file_path not in config.materials.excluded_files:
                config.materials.excluded_files.append(file_path)
            save_project(config)
            return jsonify({"success": True, "excluded": p.name})
    except Exception as e:
        return jsonify({"error": f"파일 제거 실패: {str(e)}"}), 500


@app.route("/api/models")
def api_models():
    provider = request.args.get("provider", "gemini_api")
    if provider == "openai_api":
        return jsonify(OPENAI_AI_MODELS)
    return jsonify(AI_MODELS)


@app.route("/api/scale-presets")
def api_scale_presets():
    return jsonify(SCALE_PRESETS)


@app.route("/api/keys", methods=["GET", "POST"])
def api_keys():
    if request.method == "GET":
        keys = load_api_keys()
        # 마스킹
        masked = {}
        for provider, key in keys.items():
            if key and len(key) > 8:
                masked[provider] = key[:4] + "..." + key[-4:]
            else:
                masked[provider] = key
        return jsonify({"keys": masked, "has_keys": {p: bool(k) for p, k in keys.items()}})
    else:
        data = request.json
        keys = load_api_keys()
        for provider, key in data.items():
            if key and "..." not in key:
                keys[provider] = key
        save_api_keys(keys)
        return jsonify({"success": True})


@app.route("/api/validate-key", methods=["POST"])
def api_validate_key():
    data = request.json
    key = data.get("key", "")
    
    # 저장된 키 사용
    if not key or "..." in key:
        keys = load_api_keys()
        key = keys.get("google", "")
    
    if not key:
        return jsonify({"valid": False, "error": "API 키가 없습니다."})
    
    try:
        from google import genai
        client = genai.Client(api_key=key)
        client.models.list()
        return jsonify({"valid": True})
    except Exception as e:
        return jsonify({"valid": False, "error": str(e)[:200]})


@app.route("/api/providers")
def api_providers():
    """현재 구현된 제공자와 향후 확장 슬롯."""
    return jsonify([
        {"id": "openai_api", "name": "OpenAI API (GPT)", "stable": True},
        {"id": "gemini_api", "name": "Gemini API", "stable": True},
    ])


# ═══════════════════════════════════════════════════════════

@app.route("/api/projects", methods=["GET"])
def api_list_projects():
    return jsonify(list_projects())


@app.route("/api/projects", methods=["POST"])
def api_create_project():
    data = request.get_json(silent=True) or {}
    workflow_type = data.get("workflow_type")
    if workflow_type not in WORKFLOW_TYPES:
        workflow_type = "exam" if data.get("project_type") == "exam" else "report"
    config = create_project(
        name=data.get("name", "새 프로젝트"),
        description=data.get("description", ""),
        project_type="exam" if workflow_type == "exam" else "report",
        workflow_type=workflow_type,
    )
    _apply_project_setup(config, data.get("setup"))
    
    # 기본 설정 적용
    if data.get("ai_model"):
        config.ai_model = data["ai_model"]
    if data.get("ai_provider"):
        config.ai_provider = data["ai_provider"]
    if data.get("exam"):
        _apply_exam_data(config, data["exam"])
    
    if data.get("materials"):
        m = data["materials"]
        config.materials = MaterialsConfig(
            source_type=m.get("source_type", "folder"),
            folder_path=m.get("folder_path", ""),
            file_types=m.get("file_types", ["pdf", "mp4"]),
            naming_pattern=m.get("naming_pattern", r"(\d+)\.\s*(.+)"),
        )
    
    if data.get("categories"):
        config.categories = _parse_categories(data["categories"])
    
    if data.get("prompt_template"):
        config.prompt_template = data["prompt_template"]
    else:
        config.prompt_template = generate_default_prompt(config)
    
    if config.project_type == "exam" and config.exam.questions:
        config.total_max_score = config.exam.scored_max_score
    else:
        config.total_max_score = data.get(
            "total_max_score",
            sum(cat.max_score for cat in config.categories) or 100,
        )
    
    save_project(config)
    return jsonify({
        "id": config.id,
        "name": config.name,
        "workflow_type": config.workflow_type,
        "project_type": config.project_type,
    })


@app.route("/api/projects/<project_id>", methods=["GET"])
def api_get_project(project_id):
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    return jsonify(config.to_dict())


@app.route("/api/projects/<project_id>/overview")
def api_project_overview(project_id):
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    return jsonify(build_project_overview(config))


def _criteria_validation(config: ProjectConfig) -> dict:
    errors = []
    warnings = []
    if config.project_type == "exam":
        scored = config.exam.scored_questions()
        if not scored:
            errors.append("실제로 채점할 문항 또는 소문항이 없습니다.")
        for issue in config.exam.score_validation():
            errors.append(issue["message"])
        for question in scored:
            element_total = sum(
                element.points for element in question.scoring_elements
            )
            if element_total > question.max_score:
                errors.append(
                    f"{question.number}번 부분점 합 {element_total}점이 "
                    f"배점 {question.max_score}점을 초과합니다."
                )
            elif question.scoring_elements and element_total < question.max_score:
                warnings.append(
                    f"{question.number}번 부분점 합이 배점보다 "
                    f"{question.max_score - element_total}점 적습니다."
                )
            if not question.model_answer.strip():
                warnings.append(f"{question.number}번 모범답안이 비어 있습니다.")
            effective_mode = (
                question.grading_mode
                if question.grading_mode != "inherit"
                else config.exam.grading_mode
            )
            if effective_mode == "core" and not question.core_criteria:
                warnings.append(
                    f"{question.number}번은 핵심 기준 채점이지만 압축 기준이 없습니다."
                )
        score = config.exam.scored_max_score
        unit_count = len(scored)
    else:
        criteria = config.all_criteria
        if not criteria:
            errors.append("평가 항목이 없습니다.")
        for criterion in criteria:
            if not criterion.scale:
                errors.append(f"{criterion.name} 항목의 배점 척도가 비어 있습니다.")
            if (
                config.criteria_state.delivery_mode == "core"
                and not criterion.core_criteria
            ):
                warnings.append(
                    f"{criterion.name} 항목은 AI용 압축 기준이 없어 상세 설명을 사용합니다."
                )
        score = sum(category.max_score for category in config.categories) if criteria else 0
        unit_count = len(criteria)
    return {
        "valid": not errors,
        "errors": errors,
        "warnings": warnings,
        "scored_unit_count": unit_count,
        "total_max_score": score,
    }


@app.route("/api/projects/<project_id>/criteria-versions", methods=["GET"])
def api_criteria_versions(project_id):
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    return jsonify({
        "state": vars(config.criteria_state),
        "validation": _criteria_validation(config),
        "versions": [
            criteria_version_summary(entry)
            for entry in load_versions(project_id)
        ],
    })


@app.route("/api/projects/<project_id>/criteria-versions", methods=["POST"])
def api_create_criteria_version(project_id):
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    validation = _criteria_validation(config)
    if validation["scored_unit_count"] == 0:
        return jsonify({"error": "저장할 평가기준이 없습니다."}), 400
    data = request.get_json(silent=True) or {}
    source = str(data.get("source") or config.criteria_state.source or "manual")
    entry = snapshot_criteria(
        config,
        source=source,
        note=str(data.get("note", "")),
    )
    config.criteria_state.active_version = entry["version"]
    config.criteria_state.status = "draft"
    config.criteria_state.source = entry["source"]
    config.criteria_state.updated_at = entry["created_at"]
    save_project(config)
    return jsonify({
        "success": True,
        "state": vars(config.criteria_state),
        "validation": validation,
        "version": criteria_version_summary(entry),
    })


@app.route(
    "/api/projects/<project_id>/criteria-versions/<int:version>",
    methods=["GET"],
)
def api_criteria_version_detail(project_id, version):
    entry = get_version(project_id, version)
    if not entry:
        return jsonify({"error": f"기준 v{version}을 찾을 수 없습니다."}), 404
    return jsonify(entry)


@app.route(
    "/api/projects/<project_id>/criteria-versions/<int:version>/approve",
    methods=["POST"],
)
def api_approve_criteria_version(project_id, version):
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    if config.criteria_state.active_version != version:
        return jsonify({
            "error": "현재 편집 내용과 연결된 버전만 승인할 수 있습니다. 먼저 현재 기준을 새 버전으로 저장하세요."
        }), 409
    validation = _criteria_validation(config)
    if validation["errors"]:
        return jsonify({
            "error": "배점 오류를 먼저 수정하세요.",
            "validation": validation,
        }), 409
    entry = approve_criteria(project_id, version)
    if not entry:
        return jsonify({"error": f"기준 v{version}을 찾을 수 없습니다."}), 404
    config.criteria_state.approved_version = version
    config.criteria_state.status = "approved"
    config.criteria_state.updated_at = entry["approved_at"]
    save_project(config)
    return jsonify({
        "success": True,
        "state": vars(config.criteria_state),
        "version": criteria_version_summary(entry),
    })


@app.route(
    "/api/projects/<project_id>/criteria-versions/<int:version>/restore",
    methods=["POST"],
)
def api_restore_criteria_version(project_id, version):
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    entry = get_version(project_id, version)
    if not entry:
        return jsonify({"error": f"기준 v{version}을 찾을 수 없습니다."}), 404
    payload = entry.get("payload", {})
    if payload.get("project_type", config.project_type) != config.project_type:
        return jsonify({
            "error": "현재 프로젝트 유형과 다른 평가기준 버전은 복원할 수 없습니다."
        }), 409

    config.categories = _parse_categories(payload.get("categories", []))
    exam_payload = payload.get("exam", {}) or {}
    config.exam.questions = _parse_exam_questions(exam_payload.get("questions", []))
    mode = str(exam_payload.get("grading_mode", config.exam.grading_mode))
    if mode in {"autonomous", "core", "strict"}:
        config.exam.grading_mode = mode
    config.exam.additional_instructions = str(
        exam_payload.get(
            "additional_instructions", config.exam.additional_instructions
        )
    )
    delivery_mode = str(
        payload.get("report_delivery_mode", config.criteria_state.delivery_mode)
    )
    if delivery_mode in {"core", "strict"}:
        config.criteria_state.delivery_mode = delivery_mode
    config.total_max_score = (
        config.exam.scored_max_score
        if config.project_type == "exam"
        else sum(category.max_score for category in config.categories) or 100
    )
    config.criteria_state.active_version = version
    config.criteria_state.status = "approved" if entry.get("approved") else "draft"
    config.criteria_state.source = "restored"
    config.criteria_state.updated_at = datetime.now().isoformat(timespec="seconds")
    if entry.get("approved"):
        config.criteria_state.approved_version = version
    if config.project_type == "exam":
        config.prompt_template = generate_default_prompt(config)
    else:
        config.prompt_template = str(
            payload.get("prompt_template", "")
        ) or generate_default_prompt(config)
    save_project(config)
    return jsonify({
        "success": True,
        "project": config.to_dict(),
        "validation": _criteria_validation(config),
    })


@app.route("/api/projects/<project_id>", methods=["PUT"])
def api_update_project(project_id):
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    
    data = request.json
    criteria_signature_before = _criteria_signature(config)
    
    if "name" in data:
        config.name = data["name"]
    if "description" in data:
        config.description = data["description"]
    if "ai_model" in data:
        config.ai_model = data["ai_model"]
    if "ai_provider" in data:
        config.ai_provider = data["ai_provider"]
    if "workflow_type" in data and data["workflow_type"] in WORKFLOW_TYPES:
        config.workflow_type = data["workflow_type"]
        config.project_type = "exam" if config.workflow_type == "exam" else "report"
    elif "project_type" in data and data["project_type"] in ("report", "exam"):
        config.project_type = data["project_type"]
        if config.project_type == "exam":
            config.workflow_type = "exam"
        elif config.workflow_type == "exam":
            config.workflow_type = "report"
    if "setup" in data:
        _apply_project_setup(config, data["setup"])
    if "submissions" in data:
        _apply_submissions_data(config, data["submissions"])
    if "temperature" in data:
        config.temperature = max(0.0, min(2.0, float(data["temperature"])))
    
    if "materials" in data:
        m = data["materials"]
        config.materials = MaterialsConfig(
            source_type=m.get("source_type", config.materials.source_type),
            folder_path=m.get("folder_path", config.materials.folder_path),
            file_types=m.get("file_types", config.materials.file_types),
            naming_pattern=m.get("naming_pattern", config.materials.naming_pattern),
            excluded_files=m.get("excluded_files", config.materials.excluded_files),
        )
    
    if "categories" in data:
        config.categories = _parse_categories(data["categories"])
        config.total_max_score = sum(cat.max_score for cat in config.categories) or 100

    if "criteria_state" in data:
        delivery_mode = str(
            (data.get("criteria_state") or {}).get(
                "delivery_mode", config.criteria_state.delivery_mode
            )
        )
        if delivery_mode in {"core", "strict"} and delivery_mode != config.criteria_state.delivery_mode:
            config.criteria_state.delivery_mode = delivery_mode
    
    if "prompt_template" in data:
        config.prompt_template = data["prompt_template"]
    if "exam" in data:
        _apply_exam_data(config, data["exam"] or {})
        if config.project_type == "exam":
            config.total_max_score = config.exam.scored_max_score
            # 채점 방식·문항이 바뀌면 실제 채점 프롬프트도 함께 갱신한다.
            # (서술형 프로젝트의 프롬프트는 UI에서 직접 편집하지 않는다.)
            config.prompt_template = generate_default_prompt(config)
    elif "total_max_score" in data:
        config.total_max_score = max(1, int(data["total_max_score"] or 1))

    if _criteria_signature(config) != criteria_signature_before:
        source = str(data.get("criteria_change_source", "manual"))
        _mark_criteria_changed(
            config,
            source if source in {
                "manual", "official", "generated", "synthesized", "compressed"
            } else "manual",
        )
    
    save_project(config)
    return jsonify(config.to_dict())


@app.route("/api/projects/<project_id>", methods=["DELETE"])
def api_delete_project(project_id):
    if delete_project(project_id):
        return jsonify({"success": True})
    return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404


@app.route("/api/projects/<project_id>/generate-prompt", methods=["POST"])
def api_generate_prompt(project_id):
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    prompt = generate_default_prompt(config)
    return jsonify({"prompt": prompt})


# ═══════════════════════════════════════════════════════════
# 심사자료 관리
# ═══════════════════════════════════════════════════════════

@app.route("/api/projects/<project_id>/materials")
def api_materials(project_id):
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    
    materials = find_materials(config)
    return jsonify({
        "participants": materials,
        "total_count": len(materials),
        "source_type": config.materials.source_type,
        "folder_path": config.materials.folder_path,
    })


def _submission_path_key(config: ProjectConfig, value: str) -> str:
    return str(
        resolve_project_path(config.id, value, expected_subdir="materials").resolve(strict=False)
    ).casefold()


@app.route("/api/projects/<project_id>/submissions")
def api_submissions(project_id):
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    return jsonify(build_submission_status(config))


@app.route("/api/projects/<project_id>/roster", methods=["PUT"])
def api_roster(project_id):
    """보고서·대회·시험이 함께 쓰는 명렬을 저장한다."""
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    try:
        students = _parse_students((request.json or {}).get("students", []))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    _set_roster(config, students)
    save_project(config)
    return jsonify({
        "success": True,
        "students": [vars(student) for student in students],
        "status": build_submission_status(config),
    })


@app.route("/api/projects/<project_id>/submissions/link", methods=["PUT", "DELETE"])
def api_submission_link(project_id):
    """자동 파일명 연결 한 건을 다른 학생에게 옮기거나 수동 연결을 해제한다."""
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    data = request.json or {}
    file_path = str(data.get("file_path", "")).strip()
    if not file_path:
        return jsonify({"error": "연결할 파일 경로가 필요합니다."}), 400

    available_files = {
        _submission_path_key(config, file_info["path"]): file_info["path"]
        for participant in find_materials(config, apply_links=False)
        for file_info in participant.get("files", [])
    }
    requested_key = _submission_path_key(config, file_path)
    if requested_key not in available_files:
        return jsonify({"error": "현재 프로젝트에서 찾을 수 없는 파일입니다."}), 404

    config.submissions.manual_links = [
        link
        for link in config.submissions.manual_links
        if _submission_path_key(config, link.file_path) != requested_key
    ]
    if request.method == "PUT":
        try:
            student_number = int(data.get("student_number", 0))
        except (TypeError, ValueError):
            student_number = 0
        if student_number < 1:
            return jsonify({"error": "연결할 학생 번호가 올바르지 않습니다."}), 400
        roster_numbers = {student.number for student in config.roster_students}
        if roster_numbers and student_number not in roster_numbers:
            return jsonify({"error": "명렬에 없는 학생 번호입니다."}), 400
        config.submissions.manual_links.append(SubmissionLink(
            file_path=portable_project_path(config.id, available_files[requested_key]),
            student_number=student_number,
        ))

    save_project(config)
    return jsonify({"success": True, "status": build_submission_status(config)})


@app.route("/api/projects/<project_id>/submissions/links/reset", methods=["POST"])
def api_reset_submission_links(project_id):
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    config.submissions.manual_links = []
    save_project(config)
    return jsonify({"success": True, "status": build_submission_status(config)})


@app.route("/api/projects/<project_id>/upload", methods=["POST"])
def api_upload_file(project_id):
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    
    if "file" not in request.files:
        return jsonify({"error": "파일이 없습니다."}), 400
    
    file = request.files["file"]
    if file.filename:
        saved = save_uploaded_file(config, file, file.filename)
        return jsonify({"success": True, "path": saved})
    
    return jsonify({"error": "파일명이 없습니다."}), 400


def _save_exam_source(project_id: str, storage, stem: str, allowed_extensions: set[str]) -> Path:
    original_name = Path(storage.filename or "")
    suffix = original_name.suffix.lower()
    if suffix not in allowed_extensions:
        raise ValueError(f"지원하지 않는 파일 형식입니다: {suffix or '확장자 없음'}")
    source_dir = PROJECTS_DIR / project_id / "exam_sources"
    source_dir.mkdir(parents=True, exist_ok=True)
    target = source_dir / f"{stem}{suffix}"
    storage.save(str(target))
    return target.resolve()


@app.route("/api/projects/<project_id>/exam/sources", methods=["POST"])
def api_exam_sources(project_id):
    """문제지·채점기준표·한 반 통합 스캔을 프로젝트에 보관."""
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    if config.project_type != "exam":
        return jsonify({"error": "정기고사 프로젝트에서만 사용할 수 있습니다."}), 400

    saved = {}
    try:
        if request.files.get("question_file") and request.files["question_file"].filename:
            path = _save_exam_source(
                project_id, request.files["question_file"], "questions",
                {".pdf", ".hwp", ".hwpx", ".doc", ".docx", ".txt"},
            )
            stored_path = portable_project_path(project_id, path)
            config.exam.question_source_path = stored_path
            saved["question_source_path"] = stored_path
        if request.files.get("rubric_file") and request.files["rubric_file"].filename:
            path = _save_exam_source(
                project_id, request.files["rubric_file"], "rubric",
                {".pdf", ".hwp", ".hwpx", ".doc", ".docx", ".txt"},
            )
            stored_path = portable_project_path(project_id, path)
            config.exam.rubric_source_path = stored_path
            saved["rubric_source_path"] = stored_path
        if request.files.get("scan_file") and request.files["scan_file"].filename:
            path = _save_exam_source(project_id, request.files["scan_file"], "integrated_scan", {".pdf"})
            stored_path = portable_project_path(project_id, path)
            config.exam.scan_split.source_path = stored_path
            config.exam.scan_split.completed_at = ""
            saved["scan_source_path"] = stored_path
        if not saved:
            return jsonify({"error": "업로드할 파일을 선택하세요."}), 400
        save_project(config)
        return jsonify({"success": True, **saved})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"파일 저장 실패: {str(e)[:300]}"}), 500


@app.route("/api/projects/<project_id>/exam/students", methods=["PUT"])
def api_exam_students(project_id):
    """구형 API 호환: 공통 명렬과 시험 분할 명렬을 함께 저장."""
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    try:
        students = _parse_students(
            (request.json or {}).get("students", []),
            require_students=True,
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    _set_roster(config, students)
    save_project(config)
    return jsonify({"success": True, "students": [vars(student) for student in students]})


@app.route("/api/projects/<project_id>/exam/split/preview", methods=["POST"])
def api_exam_split_preview(project_id):
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    data = request.json or {}
    try:
        plan = build_split_plan(
            resolve_project_path(
                project_id, config.exam.scan_split.source_path, expected_subdir="exam_sources"
            ),
            config.roster_students,
            start_page=data.get("start_page", 1),
            pages_per_student=data.get("pages_per_student", 0),
            boundaries=data.get("boundaries", []),
        )
        return jsonify(plan)
    except SplitValidationError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"분할 계획 생성 실패: {str(e)[:300]}"}), 500


@app.route("/api/projects/<project_id>/exam/split", methods=["POST"])
def api_exam_split(project_id):
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    data = request.json or {}
    try:
        plan = split_integrated_pdf(
            config,
            start_page=data.get("start_page", 1),
            pages_per_student=data.get("pages_per_student", 0),
            boundaries=data.get("boundaries", []),
        )
        return jsonify({"success": True, **plan})
    except SplitValidationError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"PDF 분할 실패: {str(e)[:300]}"}), 500


EXAM_QUESTION_SCHEMA = {
    "type": "object",
    "properties": {
        "id": {"type": "string"},
        "number": {"type": "string"},
        "question_text": {"type": "string"},
        "max_score": {"type": "integer"},
        "model_answer": {"type": "string"},
        "scoring_elements": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "description": {"type": "string"},
                    "points": {"type": "integer"},
                    "required": {"type": "boolean"},
                },
                "required": ["description", "points", "required"],
            },
        },
        "accepted_answers": {"type": "array", "items": {"type": "string"}},
        "common_errors": {"type": "array", "items": {"type": "string"}},
        "core_criteria": {"type": "array", "items": {"type": "string"}},
        "parent_id": {"type": "string"},
        "sub_index": {"type": "integer"},
        "answer_type": {
            "type": "string",
            "enum": list(ANSWER_TYPES),
        },
        "grading_mode": {
            "type": "string",
            "enum": list(QUESTION_GRADING_MODES),
        },
        "teacher_notes": {"type": "string"},
        "source_kind": {"type": "string"},
        "source_notes": {"type": "string"},
    },
    "required": [
        "number", "question_text", "max_score", "model_answer",
        "scoring_elements", "accepted_answers", "common_errors",
    ],
}


EXAM_RUBRIC_SCHEMA = {
    "type": "object",
    "properties": {
        "questions": {
            "type": "array",
            "items": EXAM_QUESTION_SCHEMA,
        },
    },
    "required": ["questions"],
}


EXAM_DISCOVERY_SCHEMA = {
    "type": "object",
    "properties": {
        "document_kind": {"type": "string"},
        "detected_main_question_numbers": {
            "type": "array", "items": {"type": "string"},
        },
        "coverage_notes": {"type": "string"},
        "questions": {"type": "array", "items": EXAM_QUESTION_SCHEMA},
    },
    "required": [
        "document_kind", "detected_main_question_numbers", "coverage_notes", "questions",
    ],
}


EXAM_SOURCE_MODES = {
    "auto": "문제지만인지, 문제와 여러 학생 답안이 함께 있는지, 답안만 있는지 먼저 판별",
    "question_only": "문제 원문 중심의 문제지",
    "combined_answers": "같은 인쇄 문항과 여러 학생의 필기 답안이 반복되는 통합 스캔",
    "answers_only": "문제 원문이 없거나 일부만 있고 여러 학생 답안으로 문제를 추정해야 하는 자료",
}

EXAM_STRUCTURE_GUIDANCE = (
    "\n문항 구조 규칙:\n"
    "- 소문항이 없는 대문항은 하나의 채점 문항으로 출력하세요.\n"
    "- 소문항이 있는 대문항은 공통 지문용 부모 객체와 소문항 객체들로 나누세요. "
    "부모에는 고유한 id(예: q1), 대문항 번호, 공통 지문, 소문항 합계와 같은 max_score를 넣고 "
    "model_answer와 scoring_elements는 비워 두세요.\n"
    "- 각 소문항은 parent_id에 부모 id를 넣고 sub_index를 1부터 순서대로 지정하세요. "
    "소문항 번호는 1-(1), 1-(2)처럼 대문항과 구별하고, 실제 모범답안·부분점 기준·배점은 "
    "소문항 객체에 넣으세요. 모든 소문항 max_score의 합은 부모 max_score와 같아야 합니다.\n"
    "- answer_type은 short, text, formula, diagram, mixed 중 하나로 판단하고, "
    "grading_mode는 별도 지시가 없으면 inherit로 두세요. 모든 id는 응답 안에서 고유해야 합니다.\n"
)


def _document_page_count(path: Path) -> int | None:
    if path.suffix.lower() != ".pdf":
        return None
    try:
        from pypdf import PdfReader
        return len(PdfReader(str(path)).pages)
    except Exception:
        return None


def _split_answer_candidates(project_id: str, limit: int = 5) -> list[Path]:
    """문항 복원에 유리하도록 쪽 수가 많은 학생별 답지부터 반환한다."""
    project_dir = (PROJECTS_DIR / project_id).resolve()
    answer_dir = (project_dir / "materials" / "student_answers").resolve()
    try:
        answer_dir.relative_to(project_dir)
    except ValueError:
        return []
    if not answer_dir.is_dir():
        return []
    candidates = [path for path in answer_dir.glob("*.pdf") if path.is_file()]
    candidates.sort(key=lambda path: (-(_document_page_count(path) or 0), path.name))
    return candidates[:max(1, limit)]


def _main_question_number(value: str) -> int | None:
    match = re.match(r"^\s*(\d+)", str(value or ""))
    return int(match.group(1)) if match else None


def _question_numbers(questions: list[ExamQuestion]) -> list[int]:
    return sorted({
        number for number in (_main_question_number(question.number) for question in questions)
        if number is not None
    })


def _detected_numbers(result: dict) -> list[int]:
    return sorted({
        number for number in (
            _main_question_number(value)
            for value in result.get("detected_main_question_numbers", [])
        ) if number is not None
    })


def _missing_question_numbers(
    questions: list[ExamQuestion], detected: list[int], expected_count: int = 0
) -> list[int]:
    present = set(_question_numbers(questions))
    expected = set(detected)
    if present:
        highest = max(max(present), max(expected) if expected else 0)
        if min(present) == 1 and highest <= 100:
            expected.update(range(1, highest + 1))
    if expected_count > 0:
        expected.update(range(1, expected_count + 1))
    return sorted(expected - present)


def _merge_question_candidates(
    first: list[ExamQuestion], second: list[ExamQuestion]
) -> list[ExamQuestion]:
    """두 차례 추출 결과를 합치면서 소문항의 부모 연결을 보존한다."""
    merged: dict[tuple[str, str], tuple[ExamQuestion, str]] = {}
    for batch in (first, second):
        by_id = {question.id: question for question in batch}
        for question in batch:
            number = str(question.number).strip()
            if not number:
                continue
            parent = by_id.get(question.parent_id) if question.parent_id else None
            parent_number = str(parent.number).strip() if parent else ""
            key = (parent_number.casefold(), number.casefold())
            merged[key] = (question, parent_number)

    entries = list(merged.values())
    top_entries = [entry for entry in entries if not entry[1]]
    top_numbers = {
        str(question.number).strip().casefold() for question, _ in top_entries
    }
    # 부모 객체가 한 추출 응답에 빠진 고아 소문항은 독립 문항으로 보존한다.
    orphan_entries = [
        (question, "")
        for question, parent_number in entries
        if parent_number and parent_number.casefold() not in top_numbers
    ]
    child_entries = [
        entry
        for entry in entries
        if entry[1] and entry[1].casefold() in top_numbers
    ]
    top_entries.extend(orphan_entries)
    top_entries.sort(key=lambda entry: (
        _main_question_number(entry[0].number) is None,
        _main_question_number(entry[0].number) or 0,
        str(entry[0].number),
    ))

    result = []
    root_ids = {}
    for index, (question, _) in enumerate(top_entries, 1):
        question.id = f"q{index}"
        question.parent_id = ""
        question.sub_index = 0
        root_ids[str(question.number).strip().casefold()] = question.id
        result.append(question)

    children_by_parent: dict[str, list[ExamQuestion]] = {}
    for question, parent_number in child_entries:
        children_by_parent.setdefault(parent_number.casefold(), []).append(question)
    for parent_number, children in children_by_parent.items():
        children.sort(key=lambda question: (
            question.sub_index <= 0,
            question.sub_index if question.sub_index > 0 else 0,
            str(question.number),
        ))
        parent_id = root_ids[parent_number]
        parent_position = next(
            index for index, item in enumerate(result) if item.id == parent_id
        )
        normalized_children = []
        for sub_index, question in enumerate(children, 1):
            question.id = f"{parent_id}_s{sub_index}"
            question.parent_id = parent_id
            question.sub_index = sub_index
            normalized_children.append(question)
        result[parent_position + 1:parent_position + 1] = normalized_children

    normalize_exam_questions(result)
    return result


def _exam_extraction_warnings(
    questions: list[ExamQuestion],
    detected: list[int],
    expected_count: int,
) -> list[str]:
    warnings = []
    missing = _missing_question_numbers(questions, detected, expected_count)
    if missing:
        warnings.append(
            "누락 가능성이 있는 대문항 번호: " + ", ".join(map(str, missing))
        )
    numbers = [str(question.number).strip() for question in questions]
    duplicates = sorted({number for number in numbers if number and numbers.count(number) > 1})
    if duplicates:
        warnings.append("중복된 문항 번호: " + ", ".join(duplicates))
    main_count = len(_question_numbers(questions))
    if expected_count and main_count != expected_count:
        warnings.append(
            f"예상 대문항 수는 {expected_count}개이지만 {main_count}개를 추출했습니다."
        )
    for question in questions:
        points = sum(element.points for element in question.scoring_elements)
        if points > question.max_score:
            warnings.append(
                f"{question.number}번 부분점 합 {points}점이 총점 {question.max_score}점을 초과합니다."
            )
    return warnings


def _build_exam_discovery_prompt(
    *,
    has_rubric: bool,
    source_mode: str,
    page_count: int | None,
    expected_count: int,
    audit_numbers: list[int] | None = None,
    missing_numbers: list[int] | None = None,
    split_answer: bool = False,
) -> str:
    page_text = f"이 파일은 총 {page_count}쪽입니다. " if page_count else ""
    count_text = (
        f"교사가 예상한 대문항 수는 {expected_count}개입니다. 반드시 그 수와 대조하세요. "
        if expected_count else ""
    )
    rubric_text = (
        "두 번째 첨부 파일은 공식 채점기준표이므로 배점과 채점 요소는 이를 최우선으로 따르세요."
        if has_rubric else
        "별도 공식 채점기준표가 없으므로 문제를 해결해 모범답안과 객관적인 부분점 기준을 만드세요."
    )
    audit_text = ""
    if audit_numbers is not None:
        audit_text = (
            "이 요청은 1차 결과의 누락 검증입니다. 1차에서 찾은 번호는 "
            + (", ".join(map(str, audit_numbers)) or "없음") + "입니다. "
        )
        if missing_numbers:
            audit_text += (
                "특히 다음 번호가 빠졌을 가능성이 큽니다: "
                + ", ".join(map(str, missing_numbers)) + ". "
            )
    source_text = (
        "첫 번째 첨부 파일은 통합 스캔에서 분리한 학생 한 명의 답안지입니다. "
        "이 학생이 작성한 문항만 있을 수 있으므로, 인쇄된 문제와 필기 답안을 구분하고 현재 파일에 실제로 보이는 대문항만 추출하세요. "
        if split_answer else
        "첫 번째 첨부 파일에서 서술형 시험의 완전한 대문항 목록과 채점 초안을 만드세요. "
    )

    return (
        source_text + page_text + count_text + rubric_text + "\n"
        + f"자료 유형 설정: {EXAM_SOURCE_MODES.get(source_mode, EXAM_SOURCE_MODES['auto'])}.\n"
        + audit_text
        + "반드시 문서의 첫 쪽부터 마지막 쪽까지 확인한 뒤 출력하세요. "
          "여러 학생의 답안지가 이어진 통합 스캔이면 같은 인쇄 문항이 반복됩니다. 반복본 중 가장 선명한 인쇄 영역을 결합해 "
          "문항별로 한 번만 작성하고, 학생별 필기 답안을 서로 다른 문항으로 세지 마세요. "
          "'1.', '2.' 같은 대문항 번호와 '1)', '2)' 같은 소문항 번호를 구분하세요. "
          "중간 번호가 빠지면 다른 학생의 반복 페이지와 문서 후반부를 다시 검색하세요. "
          "문제 원문이 전혀 없거나 일부만 있으면 여러 학생 답안의 공통 개념을 비교해 문제를 보수적으로 추정할 수 있으며, "
          "그 경우 source_kind를 inferred 또는 mixed로 표시하세요. 임의로 서로 다른 문제를 합치지 마세요.\n"
          "detected_main_question_numbers에는 문서 전체에서 발견하거나 확실히 추정한 모든 대문항 번호를 먼저 나열하세요. "
          "questions에는 그 번호 각각이 정확히 한 번씩 있어야 합니다. document_kind는 question_only, combined_answers, "
          "answers_only 중 하나로 쓰고, coverage_notes에는 페이지 반복·누락·추정 여부를 간단히 기록하세요. "
          "각 문항에는 원문 또는 보수적으로 복원한 문제, 총 배점, 모범답안, 부분점 요소, 허용 답안과 주요 오류를 작성하세요. "
          "부분점 합은 총 배점을 초과할 수 없습니다."
        + EXAM_STRUCTURE_GUIDANCE
    )


def _merge_official_exam_questions(
    existing_questions: list[ExamQuestion], extracted_questions: list[ExamQuestion]
) -> list[ExamQuestion]:
    """공식 기준의 점수 요소를 우선하고, 문제지에서 얻은 원문은 보존."""
    existing_by_number = {str(question.number).strip(): question for question in existing_questions}
    for question in extracted_questions:
        existing = existing_by_number.get(str(question.number).strip())
        if existing is None:
            continue
        if existing.question_text.strip():
            question.question_text = existing.question_text
        elif not question.question_text.strip():
            question.question_text = f"{question.number}번 문항"
        if not question.model_answer.strip() and existing.model_answer.strip():
            question.model_answer = existing.model_answer
        if not question.accepted_answers and existing.accepted_answers:
            question.accepted_answers = list(existing.accepted_answers)
        if not question.common_errors and existing.common_errors:
            question.common_errors = list(existing.common_errors)
    return extracted_questions


@app.route("/api/projects/<project_id>/exam/rubric/extract", methods=["POST"])
def api_exam_extract_official_rubric(project_id):
    """보고서형 루브릭과 같은 흐름으로 공식 시험 채점기준표를 즉시 구조화."""
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    if config.project_type != "exam":
        return jsonify({"error": "서술형 시험 프로젝트에서만 사용할 수 있습니다."}), 400
    storage = request.files.get("file")
    if storage is None or not storage.filename:
        return jsonify({"error": "공식 채점기준표 파일을 선택하세요."}), 400

    try:
        rubric_path = _save_exam_source(
            project_id, storage, "rubric",
            {".pdf", ".hwp", ".hwpx", ".doc", ".docx", ".txt"},
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"error": f"채점기준표 저장 실패: {str(exc)[:300]}"}), 500

    existing_summary = [
        {
            "id": question.id,
            "number": question.number,
            "question_text": question.question_text,
            "max_score": question.max_score,
            "parent_id": question.parent_id,
            "sub_index": question.sub_index,
            "answer_type": question.answer_type,
        }
        for question in config.exam.questions
    ]
    prompt = (
        "첨부 파일은 교사가 확정한 공식 서술형 채점기준표입니다. 문서 내용을 임의로 보완하거나 "
        "새 기준을 만들지 말고, 문항 번호·배점·모범답안·부분점 요소·허용 답안·감점 사례를 "
        "편집 가능한 구조로 정확히 옮기세요. 부분점 합계는 문항 배점을 초과할 수 없습니다. "
        "문제 원문이 기준표에 생략되어 있으면 아래 기존 문항 정보의 같은 번호를 사용하세요.\n"
        + EXAM_STRUCTURE_GUIDANCE
        + "\n"
        "기존 문항 정보:\n"
        + json.dumps(existing_summary, ensure_ascii=False)
    )
    try:
        provider = get_provider(config, load_api_keys())
        result = provider.generate_json(
            prompt,
            schema=EXAM_RUBRIC_SCHEMA,
            files=[str(rubric_path)],
            model_name=config.ai_model,
            temperature=0.1,
        )
        extracted = _parse_exam_questions(result.get("questions", []))
        if not extracted:
            return jsonify({"error": "공식 채점기준표에서 문항을 추출하지 못했습니다."}), 500
        questions = _merge_official_exam_questions(config.exam.questions, extracted)
        config.exam.rubric_source_path = portable_project_path(project_id, rubric_path)
        config.exam.questions = questions
        # 공식 기준표를 올렸다는 것은 그 기준의 준수가 중요하다는 뜻이므로
        # 엄격 적용을 기본으로 두고, 교사가 채점 방식에서 바꿀 수 있게 한다.
        config.exam.grading_mode = "strict"
        config.total_max_score = config.exam.scored_max_score
        _mark_criteria_changed(config, "official")
        config.prompt_template = generate_default_prompt(config)
        save_project(config)
        return jsonify({
            "success": True,
            "rubric_source_path": config.to_dict()["exam"]["rubric_source_path"],
            "questions": config.to_dict()["exam"]["questions"],
            "total_max_score": config.total_max_score,
            "prompt_template": config.prompt_template,
        })
    except ProviderNeedsUserAction as exc:
        return jsonify({"error": str(exc), "needs_user_action": True}), 409
    except Exception as exc:
        return jsonify({"error": f"공식 채점기준표 추출 실패: {str(exc)[:400]}"}), 500


EXAM_SYNTHESIS_SCHEMA = {
    "type": "object",
    "properties": {
        "questions": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "number": {"type": "string"},
                    "model_answer": {"type": "string"},
                    "scoring_elements": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "description": {"type": "string"},
                                "points": {"type": "integer"},
                                "required": {"type": "boolean"},
                            },
                            "required": ["description", "points"],
                        },
                    },
                    "accepted_answers": {"type": "array", "items": {"type": "string"}},
                    "common_errors": {"type": "array", "items": {"type": "string"}},
                    "rationale": {"type": "string"},
                },
                "required": ["number", "model_answer"],
            },
        },
    },
    "required": ["questions"],
}


def _build_rubric_synthesis_prompt(config, completed: dict) -> str:
    """개별 채점 결과를 종합해 기준안을 만들기 위한 프롬프트."""
    lines = [
        f"'{config.name}' 서술형 시험의 개별 채점이 완료되었습니다.",
        f"아래는 학생 {len(completed)}명의 문항별 점수, 답안 요약, 채점 근거입니다.",
        "이 실제 채점 결과를 종합해 교사가 검토할 문항별 채점 기준안을 만드세요.",
        "",
        "규칙:",
        "- model_answer(예시답안)는 만점 답의 핵심 내용을 새로 정리해 작성하세요.",
        "- scoring_elements(부분점 기준안)는 실제 학생들 사이에서 점수 차이를 만든 요소를"
        " 기준으로 나누고, 점수 합이 배점을 넘지 않게 하세요.",
        "- accepted_answers(정답 인정 답안)에는 표현은 다르지만 정답으로 인정된, 혹은"
        " 인정해야 할 실제 답안 표현을 넣으세요.",
        "- common_errors(감점·오답 사례)에는 실제로 관찰된 대표적인 오답·감점 유형을 넣으세요.",
        "- rationale에는 이 기준안이 실제 채점 결과와 어떻게 부합하는지 한두 문장으로 쓰세요.",
        "- 채점 결과에서 관찰되지 않은 내용을 지어내지 마세요.",
    ]
    for question in config.exam.scored_questions():
        lines.extend([
            "",
            f"## {question.number}번 (배점 {question.max_score}점)",
            f"문제: {question.question_text}",
        ])
        if question.model_answer.strip():
            lines.append(f"교사가 채점 전 입력한 참고 답: {question.model_answer.strip()}")
        lines.append("학생별 채점 결과:")
        for team in sorted(completed):
            result = completed[team]
            score = result.get(question.id)
            if score is None:
                continue
            summary = str(result.get(f"{question.id}_answer_summary", "")).strip()[:250]
            reason = str(result.get(f"{question.id}_reason", "")).strip()[:250]
            lines.append(f"- 학생{team}: {score}점 | 답안: {summary} | 근거: {reason}")
    return "\n".join(lines)


@app.route("/api/projects/<project_id>/exam/synthesize-rubric", methods=["POST"])
def api_exam_synthesize_rubric(project_id):
    """전체 채점 결과를 종합해 채점 기준안·예시답안·인정 답안 초안을 만든다."""
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    if config.project_type != "exam" or not config.exam.scored_questions():
        return jsonify({"error": "서술형 시험 프로젝트에서 문항을 먼저 준비하세요."}), 400

    data = request.get_json(silent=True) or {}
    try:
        round_id = int(data.get("round_id") or get_latest_round_id(project_id))
    except (TypeError, ValueError):
        return jsonify({"error": "회차 번호가 올바르지 않습니다."}), 400
    completed = load_completed(project_id, round_id)
    if not completed:
        return jsonify({"error": f"{round_id}회차에 종합할 채점 결과가 없습니다. 먼저 채점을 완료하세요."}), 400

    prompt = _build_rubric_synthesis_prompt(config, completed)
    try:
        provider = get_provider(config, load_api_keys())
        result = provider.generate_json(
            prompt,
            schema=EXAM_SYNTHESIS_SCHEMA,
            model_name=config.ai_model,
            temperature=0.2,
        )
    except ProviderNeedsUserAction as e:
        return jsonify({"error": str(e), "needs_user_action": True}), 409
    except Exception as e:
        return jsonify({"error": f"채점 결과 종합 실패: {str(e)[:300]}"}), 500

    drafts = {}
    for item in result.get("questions", []):
        number = str(item.get("number", "")).strip()
        if number:
            drafts[number] = item

    merged = []
    for question in config.exam.scored_questions():
        draft = drafts.get(str(question.number).strip())
        if not draft:
            continue
        elements = []
        for raw_element in draft.get("scoring_elements", []) or []:
            try:
                points = max(0, int(raw_element.get("points", 0)))
            except (TypeError, ValueError):
                continue
            description = str(raw_element.get("description", "")).strip()
            if points and description:
                elements.append({
                    "description": description,
                    "points": points,
                    "required": bool(raw_element.get("required", True)),
                })
        merged.append({
            "number": question.number,
            "max_score": question.max_score,
            "model_answer": str(draft.get("model_answer", "")).strip(),
            "scoring_elements": elements,
            "accepted_answers": [
                str(v).strip() for v in (draft.get("accepted_answers") or []) if str(v).strip()
            ],
            "common_errors": [
                str(v).strip() for v in (draft.get("common_errors") or []) if str(v).strip()
            ],
            "rationale": str(draft.get("rationale", "")).strip(),
        })
    if not merged:
        return jsonify({"error": "종합 결과에서 문항 기준안을 만들지 못했습니다. 다시 시도하세요."}), 500
    return jsonify({
        "round_id": round_id,
        "student_count": len(completed),
        "questions": merged,
    })


@app.route("/api/projects/<project_id>/exam/compress-criteria", methods=["POST"])
def api_exam_compress_criteria(project_id):
    """교사용 상세 기준을 AI 채점용 핵심 확인 요소(3~5개)로 문항별 압축한다."""
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    if config.project_type != "exam" or not config.exam.scored_questions():
        return jsonify({"error": "서술형 문항을 먼저 준비하세요."}), 400

    scored_ids = {question.id for question in config.exam.scored_questions()}
    detailed = [
        question
        for question in config.to_dict()["exam"]["questions"]
        if question.get("id") in scored_ids
    ]
    prompt = (
        "당신은 서술형 채점 기준을 정리하는 교사 보조자입니다.\n"
        "아래 문항별 상세 채점 기준을, AI가 채점할 때 반드시 확인해야 할 "
        "핵심 확인 요소로 문항마다 3~5개씩 압축하세요.\n"
        "규칙:\n"
        "1. 각 요소는 한 문장으로, 실제로 점수 차이를 만드는 판단 기준만 남기세요.\n"
        "2. 세부 배점 수치는 점수 배분에 꼭 필요한 경우에만 포함하세요.\n"
        "3. 상세 기준에 없는 내용을 새로 만들지 마세요.\n"
        "4. 상세 기준이 거의 없는 문항은 모범답안의 핵심을 요소로 삼으세요.\n\n"
        "## 문항별 상세 기준\n"
        + json.dumps(detailed, ensure_ascii=False)
    )
    schema = {
        "type": "object",
        "properties": {
            "questions": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "number": {"type": "string"},
                        "core_criteria": {"type": "array", "items": {"type": "string"}},
                    },
                    "required": ["number", "core_criteria"],
                },
            },
        },
        "required": ["questions"],
    }
    try:
        provider = get_provider(config, load_api_keys())
        result = provider.generate_json(
            prompt, schema=schema, model_name=config.ai_model, temperature=0.1,
        )
    except ProviderNeedsUserAction as e:
        return jsonify({"error": str(e), "needs_user_action": True}), 409
    except Exception as e:
        return jsonify({"error": f"핵심 기준 압축 실패: {str(e)[:300]}"}), 500

    by_number = {}
    for item in result.get("questions", []):
        number = str(item.get("number", "")).strip()
        if number:
            by_number[number] = [
                str(v).strip() for v in (item.get("core_criteria") or []) if str(v).strip()
            ][:6]
    updated = 0
    for question in config.exam.scored_questions():
        criteria = by_number.get(str(question.number).strip())
        if criteria:
            question.core_criteria = criteria
            updated += 1
    if not updated:
        return jsonify({"error": "핵심 기준을 만들지 못했습니다. 상세 기준을 확인하세요."}), 500

    _mark_criteria_changed(config, "compressed")
    config.prompt_template = generate_default_prompt(config)
    save_project(config)
    return jsonify({
        "success": True,
        "updated": updated,
        "questions": config.to_dict()["exam"]["questions"],
        "prompt_template": config.prompt_template,
        "grading_mode": config.exam.grading_mode,
    })


@app.route("/api/projects/<project_id>/exam/rubric/from-text", methods=["POST"])
def api_exam_rubric_from_text(project_id):
    """텍스트 지시만으로 문항 초안을 만들거나(문항 없음) 기존 문항을 일괄 수정한다.

    보고서 평가의 '방법 1. 텍스트로 루브릭 생성'을 서술형에 이식한 기능.
    """
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    if config.project_type != "exam":
        return jsonify({"error": "서술형 시험 프로젝트에서만 사용할 수 있습니다."}), 400

    data = request.get_json(silent=True) or {}
    text = str(data.get("text", "")).strip()
    if not text:
        return jsonify({"error": "지시 텍스트를 입력하세요."}), 400

    existing = config.to_dict()["exam"]["questions"]
    if existing:
        mode = "revise"
        prompt = (
            "당신은 서술형 시험의 문항별 채점 기준을 관리하는 교사 보조자입니다.\n"
            "아래는 현재 저장된 문항 목록입니다. 교사의 지시를 반영해 전체 문항 목록을 다시 출력하세요.\n"
            "규칙:\n"
            "1. 지시와 무관한 문항과 필드는 원래 값을 그대로 유지하세요 (number, question_text, max_score 포함).\n"
            "2. 지시가 특정 문항만 언급하면 그 문항만 바꾸고 나머지는 건드리지 마세요.\n"
            "3. 각 문항의 부분점 합이 배점(max_score)을 넘지 않게 하세요.\n"
            "4. 문항의 추가·삭제·번호 변경은 교사가 명시적으로 지시했을 때만 하세요.\n\n"
            + EXAM_STRUCTURE_GUIDANCE
            + "\n"
            "## 현재 문항 목록\n"
            + json.dumps(existing, ensure_ascii=False)
            + "\n\n## 교사 지시\n"
            + text
        )
    else:
        mode = "create"
        prompt = (
            "당신은 서술형 시험의 문항과 채점 기준 초안을 만드는 교사 보조자입니다.\n"
            "아래 교사의 설명만으로 문항 목록을 만드세요. 문제 원문이 포함되어 있으면 "
            "question_text에 그대로 옮기고, 요약 설명이면 취지에 맞게 간결히 작성하세요.\n"
            "규칙:\n"
            "1. 각 문항에 번호, 문제, 배점, 모범답안, 부분점 요소(합이 배점 이하), "
            "정답 인정 답안, 주요 감점 사례를 채우세요.\n"
            "2. 교사가 문항 수나 배점을 명시했으면 정확히 따르세요.\n"
            "3. 교사가 주지 않은 내용은 과목과 문제 맥락에 맞는 합리적인 초안으로 채우되, "
            "교사가 검토·수정할 것을 전제로 간결하게 작성하세요.\n\n"
            + EXAM_STRUCTURE_GUIDANCE
            + "\n"
            "## 교사 설명\n"
            + text
        )

    schema = {
        "type": "object",
        "properties": {"questions": {"type": "array", "items": EXAM_QUESTION_SCHEMA}},
        "required": ["questions"],
    }
    try:
        provider = get_provider(config, load_api_keys())
        result = provider.generate_json(
            prompt, schema=schema, model_name=config.ai_model, temperature=0.2,
        )
    except ProviderNeedsUserAction as e:
        return jsonify({"error": str(e), "needs_user_action": True}), 409
    except Exception as e:
        return jsonify({"error": f"문항 생성·수정 실패: {str(e)[:300]}"}), 500

    questions = _parse_exam_questions(result.get("questions", []))
    if not questions:
        return jsonify({"error": "지시에서 문항을 만들지 못했습니다. 설명을 조금 더 구체적으로 써 주세요."}), 500

    config.exam.questions = questions
    config.total_max_score = config.exam.scored_max_score
    _mark_criteria_changed(config, "generated")
    config.prompt_template = generate_default_prompt(config)
    save_project(config)
    return jsonify({
        "success": True,
        "mode": mode,
        "questions": config.to_dict()["exam"]["questions"],
        "total_max_score": config.total_max_score,
        "prompt_template": config.prompt_template,
    })


@app.route("/api/projects/<project_id>/exam/generate-rubric", methods=["POST"])
def api_exam_generate_rubric(project_id):
    """기준표가 있으면 추출하고, 없으면 문제를 풀어 모범답안·부분점 기준을 생성."""
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404

    request_data = request.get_json(silent=True) or {}
    source_mode = str(request_data.get("source_mode", config.exam.source_mode) or "auto")
    if source_mode not in EXAM_SOURCE_MODES:
        return jsonify({"error": "지원하지 않는 문제 자료 유형입니다."}), 400
    try:
        expected_count = max(
            0,
            int(request_data.get(
                "expected_question_count", config.exam.expected_question_count
            ) or 0),
        )
    except (TypeError, ValueError):
        return jsonify({"error": "예상 대문항 수는 0 이상의 정수로 입력하세요."}), 400

    split_sources = _split_answer_candidates(project_id)
    prefer_split_answers = bool(request_data.get("prefer_split_answers", False))
    question_source = config.exam.question_source_path.strip()
    question_path = (
        resolve_project_path(project_id, question_source, expected_subdir="exam_sources")
        if question_source else None
    )
    if question_path is not None and not question_path.is_file():
        question_path = None
    use_split_answers = bool(split_sources) and (
        prefer_split_answers
        or source_mode in {"combined_answers", "answers_only"}
        or question_path is None
    )
    if question_path is None and not split_sources:
        return jsonify({"error": "문제지 또는 분할된 학생 답지 PDF가 필요합니다."}), 400
    source_paths = split_sources if use_split_answers else [question_path]

    rubric_source = config.exam.rubric_source_path.strip()
    rubric_path = (
        resolve_project_path(project_id, rubric_source, expected_subdir="exam_sources")
        if rubric_source else None
    )
    if rubric_path is not None and not rubric_path.is_file():
        return jsonify({"error": "저장된 채점기준표 파일을 찾을 수 없습니다. 다시 업로드하세요."}), 400
    has_rubric = rubric_path is not None

    config.exam.source_mode = source_mode
    config.exam.expected_question_count = expected_count
    first_source = source_paths[0]
    page_count = _document_page_count(first_source)
    prompt = _build_exam_discovery_prompt(
        has_rubric=has_rubric,
        source_mode=source_mode,
        page_count=page_count,
        expected_count=expected_count,
        split_answer=use_split_answers,
    )
    try:
        provider = get_provider(config, load_api_keys())
        first_files = [str(first_source)]
        if has_rubric:
            first_files.append(str(rubric_path))
        first_result = provider.generate_json(
            prompt,
            schema=EXAM_DISCOVERY_SCHEMA,
            files=first_files,
            model_name=config.ai_model,
            temperature=0.1,
        )
        first_questions = _parse_exam_questions(first_result.get("questions", []))
        if not first_questions:
            return jsonify({"error": "문항을 추출하지 못했습니다."}), 500

        first_detected = sorted(set(
            _detected_numbers(first_result) + _question_numbers(first_questions)
        ))
        first_missing = _missing_question_numbers(
            first_questions, first_detected, expected_count
        )
        first_kind = str(first_result.get("document_kind", "")).strip()
        audit_required = bool(
            first_missing
            or source_mode in {"combined_answers", "answers_only"}
            or first_kind in {"combined_answers", "answers_only"}
            or (
                page_count is not None
                and page_count >= 12
                and len(_question_numbers(first_questions)) <= 5
            )
            or (
                expected_count
                and len(_question_numbers(first_questions)) != expected_count
            )
        )

        questions = first_questions
        detected = first_detected
        document_kind = first_kind or source_mode
        coverage_notes = str(first_result.get("coverage_notes", "")).strip()
        extraction_changed = False
        source_files_checked = [first_source.name]
        audit_performed = False
        if audit_required:
            audit_sources = source_paths[1:] if use_split_answers else [first_source]
            if not audit_sources:
                audit_sources = [first_source]
        else:
            audit_sources = []

        for audit_source in audit_sources:
            current_missing = _missing_question_numbers(
                questions, detected, expected_count
            )
            if audit_performed and not current_missing:
                break
            audit_performed = True
            source_files_checked.append(audit_source.name)
            audit_prompt = _build_exam_discovery_prompt(
                has_rubric=has_rubric,
                source_mode=source_mode,
                page_count=_document_page_count(audit_source),
                expected_count=expected_count,
                audit_numbers=detected,
                missing_numbers=current_missing,
                split_answer=use_split_answers,
            )
            audit_files = [str(audit_source)]
            if has_rubric:
                audit_files.append(str(rubric_path))
            second_result = provider.generate_json(
                audit_prompt,
                schema=EXAM_DISCOVERY_SCHEMA,
                files=audit_files,
                model_name=config.ai_model,
                temperature=0.1,
            )
            second_questions = _parse_exam_questions(second_result.get("questions", []))
            second_detected = sorted(set(
                _detected_numbers(second_result) + _question_numbers(second_questions)
            ))
            if second_questions:
                previous_numbers = _question_numbers(questions)
                second_numbers = _question_numbers(second_questions)
                extraction_changed = extraction_changed or previous_numbers != second_numbers
                questions = _merge_question_candidates(questions, second_questions)
                detected = sorted(set(detected + second_detected))
                second_kind = str(second_result.get("document_kind", "")).strip()
                if second_kind:
                    document_kind = second_kind
                second_notes = str(second_result.get("coverage_notes", "")).strip()
                if second_notes and second_notes != coverage_notes:
                    coverage_notes = " / ".join(filter(None, [coverage_notes, second_notes]))

        warnings = _exam_extraction_warnings(questions, detected, expected_count)
        if extraction_changed:
            warnings.insert(
                0,
                "1차 추출과 누락 검증의 문항 번호가 달라 두 결과를 통합했습니다.",
            )
        remaining_missing = _missing_question_numbers(questions, detected, expected_count)
        if expected_count and remaining_missing:
            return jsonify({
                "error": (
                    "분할 답지를 여러 장 확인했지만 다음 문항을 찾지 못했습니다: "
                    + ", ".join(map(str, remaining_missing))
                ),
                "questions": [question.number for question in questions],
                "missing_question_numbers": remaining_missing,
                "source_files_checked": source_files_checked,
                "needs_review": True,
            }), 422
        config.exam.questions = questions
        config.total_max_score = config.exam.scored_max_score
        _mark_criteria_changed(
            config, "official" if has_rubric else "generated"
        )
        config.prompt_template = generate_default_prompt(config)
        save_project(config)
        return jsonify({
            "success": True,
            "used_rubric": has_rubric,
            "questions": config.to_dict()["exam"]["questions"],
            "total_max_score": config.total_max_score,
            "prompt_template": config.prompt_template,
            "document_kind": document_kind,
            "detected_main_question_numbers": detected,
            "coverage_notes": coverage_notes,
            "page_count": page_count,
            "audit_performed": audit_performed,
            "used_split_answers": use_split_answers,
            "source_files_checked": source_files_checked,
            "warnings": warnings,
            "needs_review": bool(warnings),
        })
    except ProviderNeedsUserAction as e:
        return jsonify({"error": str(e), "needs_user_action": True}), 409
    except Exception as e:
        return jsonify({"error": f"모범답안·채점기준 생성 실패: {str(e)[:400]}"}), 500


@app.route("/api/projects/<project_id>/convert-hwp", methods=["POST"])
def api_convert_hwp(project_id):
    """심사자료 폴더 내 HWP/HWPX 파일을 PDF로 일괄 변환 (한컴오피스 한글 COM)"""
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    
    # 심사자료 폴더 확인
    folder_path = Path(config.materials.folder_path) if config.materials.folder_path else None
    if not folder_path or not folder_path.exists():
        return jsonify({"error": "심사자료 폴더가 설정되지 않았거나 존재하지 않습니다."}), 400
    
    # 폴더 직접 스캔: 변환 대상 문서 수집 (PDF가 아직 없는 것만)
    convert_exts = (".hwp", ".hwpx", ".docx", ".doc")
    doc_files = []
    for f in folder_path.rglob("*"):
        if f.suffix.lower() in convert_exts and f.is_file():
            pdf_path = f.with_suffix(".pdf")
            if not pdf_path.exists():
                doc_files.append({"path": str(f), "name": f.name})
    
    if not doc_files:
        return jsonify({"message": "변환할 문서 파일이 없습니다. (이미 모두 변환되었거나 HWP/DOCX 파일이 없습니다)", "converted": 0})
    
    # 한컴오피스 한글 COM으로 변환
    try:
        import win32com.client
    except ImportError:
        return jsonify({"error": "pywin32가 설치되지 않았습니다.\npip install pywin32 로 설치하세요."}), 400
    
    converted = []
    failed = []
    hwp = None
    
    try:
        hwp = win32com.client.Dispatch("HWPFrame.HwpObject")
        hwp.XHwpWindows.Item(0).Visible = False
        hwp.RegisterModule("FilePathCheckDLL", "SecurityModule")
        try:
            hwp.SetAltMode(0)
        except Exception:
            pass
        
        for f in doc_files:
            src = Path(f["path"])
            pdf_path = src.with_suffix(".pdf")
            try:
                hwp.Open(str(src), "HWP", "forceopen:true;versionwarning:false")
                hwp.HAction.GetDefault("FileSaveAs_S", hwp.HParameterSet.HFileOpenSave.HSet)
                hwp.HParameterSet.HFileOpenSave.filename = str(pdf_path)
                hwp.HParameterSet.HFileOpenSave.Format = "PDF"
                hwp.HAction.Execute("FileSaveAs_S", hwp.HParameterSet.HFileOpenSave.HSet)
                
                # PDF 저장 완료 대기 (비동기 저장 문제 방지)
                import time
                for _ in range(30):  # 최대 15초 대기
                    time.sleep(0.5)
                    if pdf_path.exists() and pdf_path.stat().st_size > 500:
                        break
                
                hwp.Clear(False)
                
                if pdf_path.exists() and pdf_path.stat().st_size > 500:
                    converted.append(f["name"])
                else:
                    if pdf_path.exists():
                        pdf_path.unlink()  # 백지 PDF 삭제
                    failed.append({"name": f["name"], "error": "PDF 변환이 완료되지 않았습니다 (백지)"})
            except Exception as e:
                failed.append({"name": f["name"], "error": str(e)[:200]})
                try:
                    hwp.Clear(False)
                except Exception:
                    pass
    except Exception as e:
        return jsonify({"error": f"한컴오피스 한글을 실행할 수 없습니다.\n한글(2010 이상)이 설치되어 있는지 확인하세요.\n\n오류: {str(e)[:200]}"}), 400
    finally:
        if hwp:
            try:
                hwp.Quit()
            except Exception:
                pass
    
    return jsonify({
        "converted": len(converted),
        "converted_files": converted,
        "failed": len(failed),
        "failed_files": failed,
    })

# ═══════════════════════════════════════════════════════════
# 채점 실행
# ═══════════════════════════════════════════════════════════

@app.route("/api/projects/<project_id>/rounds")
def api_rounds(project_id):
    results_dir = PROJECTS_DIR / project_id / "results"
    results_dir.mkdir(parents=True, exist_ok=True)
    rounds = []
    for d in sorted(results_dir.iterdir()):
        if d.is_dir() and d.name.startswith("round_"):
            try:
                round_id = int(d.name.split("_")[1])
                count = len(list(d.glob("team_*.json")))
                rounds.append({"id": round_id, "count": count})
            except ValueError:
                pass
    return jsonify(rounds)


@app.route("/api/projects/<project_id>/start", methods=["POST"])
def api_start_grading(project_id):
    if grading_state["running"]:
        return jsonify({"error": "이미 채점이 진행 중입니다."}), 409
    
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    
    # 채점 기준 확인
    if config.project_type == "exam":
        if not config.exam.scored_questions():
            return jsonify({"error": "서술형 문항과 채점 기준을 먼저 준비하세요."}), 400
        validation = _criteria_validation(config)
        if validation["errors"]:
            return jsonify({
                "error": "문항 배점 오류를 먼저 수정하세요.",
                "validation": validation,
            }), 409
    elif not config.all_criteria:
        return jsonify({"error": "채점 기준(Rubric)이 설정되지 않았습니다. 설정 탭에서 채점 기준을 추가하세요."}), 400
    
    data = request.json or {}
    start_from = data.get("start_from", 1)
    delay = data.get("delay", 5)
    repeat_count = data.get("repeat_count", 1)
    new_round = data.get("new_round", False)
    team_numbers = data.get("team_numbers") # 특정 팀 번호 리스트 (개별/선택 채점용)
    
    # 제공자별 사전 조건 확인
    keys = load_api_keys()
    if config.ai_provider == "gemini_api" and not keys.get("google", ""):
        return jsonify({"error": "Google API 키가 설정되지 않았습니다. 우측 상단 'API 키' 버튼에서 설정하세요."}), 400
    if config.ai_provider == "openai_api" and not keys.get("openai", ""):
        return jsonify({"error": "OpenAI API 키가 설정되지 않았습니다. 우측 상단 'API 키' 버튼에서 설정하세요."}), 400
    
    # 라운드 설정
    latest_round = get_latest_round_id(project_id)
    latest_has_results = any(get_round_dir(project_id, latest_round).glob("team_*.json"))
    current_round = latest_round + 1 if new_round and latest_has_results else latest_round
    get_round_dir(project_id, current_round).mkdir(parents=True, exist_ok=True)
    
    grading_state["running"] = True
    grading_state["should_stop"] = False
    grading_state["project_id"] = project_id
    grading_state["current_round"] = current_round
    grading_state["completed_count"] = 0
    grading_state["total_count"] = 0
    grading_state["success_count"] = 0
    grading_state["fail_count"] = 0
    grading_state["current_team"] = None
    grading_state["current_step"] = "채점 준비 중..."
    grading_state["started_at"] = time.time()
    grading_state["team_started_at"] = None
    grading_state["stop_requested_at"] = None
    
    thread = threading.Thread(
        target=grading_worker,
        args=(project_id, keys, start_from, delay, repeat_count, team_numbers),
        daemon=True,
    )
    thread.start()
    
    return jsonify({"message": "채점을 시작합니다.", "round": current_round})


@app.route("/api/stop", methods=["POST"])
def api_stop():
    if not grading_state["running"]:
        return jsonify({"success": True, "pending": False})
    grading_state["should_stop"] = True
    grading_state["stop_requested_at"] = time.time()
    grading_state["current_step"] = "중단 요청됨 - 현재 AI 응답이 끝나면 중단합니다."
    return jsonify({"success": True, "pending": True})


@app.route("/api/status")
def api_status():
    return jsonify(grading_state)


@app.route("/api/events")
def api_events():
    q = Queue()
    event_queues.append(q)
    
    def stream():
        try:
            while True:
                try:
                    msg = q.get(timeout=60)
                    yield f"data: {msg}\n\n"
                except Exception:
                    # 하트비트 전송 (연결 유지)
                    yield ": heartbeat\n\n"
        except GeneratorExit:
            pass
        finally:
            if q in event_queues:
                event_queues.remove(q)
    
    return Response(stream(), mimetype="text/event-stream",
                   headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ═══════════════════════════════════════════════════════════
# 채점 결과 조회
# ═══════════════════════════════════════════════════════════

@app.route("/api/projects/<project_id>/results")
def api_results(project_id):
    round_id = request.args.get("round", type=int)
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    
    completed = load_completed(project_id, round_id)
    results = sorted(completed.values(), key=lambda x: x.get("total_score", 0), reverse=True)
    
    for i, r in enumerate(results, 1):
        r["rank"] = i
    
    return jsonify(results)


@app.route("/api/projects/<project_id>/result/<int:team_num>")
def api_result_detail(project_id, team_num):
    round_id = request.args.get("round", type=int)
    completed = load_completed(project_id, round_id)
    if team_num not in completed:
        return jsonify({"error": "결과가 없습니다."}), 404
    return jsonify(completed[team_num])


@app.route("/api/projects/<project_id>/result/<int:team_num>", methods=["PUT"])
def api_update_result(project_id, team_num):
    """채점 결과 수동 수정"""
    round_id = request.args.get("round", type=int) or get_latest_round_id(project_id)
    new_data = request.json
    if not new_data:
        return jsonify({"error": "데이터가 없습니다."}), 400
    
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    
    completed = load_completed(project_id, round_id)
    if team_num not in completed:
        return jsonify({"error": "기존 결과를 찾을 수 없습니다."}), 404
    
    # 데이터 업데이트 (팀 번호는 고정)
    old_data = completed[team_num]
    before = dict(old_data)
    for k, v in new_data.items():
        if k not in ("team_number", "teacher_status", "teacher_approved_at", "audit_log"):
            old_data[k] = v
            
    # 소계/합계 재계산
    updated_data = compute_scores(config, old_data)
    if config.project_type == "exam":
        changed = {
            key: {"before": before.get(key), "after": updated_data.get(key)}
            for key in new_data
            if before.get(key) != updated_data.get(key)
            and key not in ("audit_log", "teacher_status", "teacher_approved_at")
        }
        audit_log = list(before.get("audit_log", []))
        audit_log.append({
            "timestamp": datetime.now().isoformat(),
            "action": "manual_edit",
            "changes": changed,
        })
        updated_data["audit_log"] = audit_log
        updated_data["teacher_status"] = "pending"
        updated_data.pop("teacher_approved_at", None)
    
    # 저장
    save_result(project_id, round_id, updated_data, team_num)
    return jsonify({"success": True, "data": updated_data})


@app.route("/api/projects/<project_id>/result/<int:team_num>/approve", methods=["POST"])
def api_approve_result(project_id, team_num):
    """교사가 확인한 시험 채점만 확정 상태로 전환."""
    round_id = request.args.get("round", type=int) or get_latest_round_id(project_id)
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    if config.project_type != "exam":
        return jsonify({"error": "정기고사 결과만 승인할 수 있습니다."}), 400
    completed = load_completed(project_id, round_id)
    if team_num not in completed:
        return jsonify({"error": "결과가 없습니다."}), 404

    result = completed[team_num]
    payload = request.json or {}
    approved_at = datetime.now().isoformat()
    result["teacher_status"] = "approved"
    result["teacher_approved_at"] = approved_at
    result["teacher_note"] = str(payload.get("teacher_note", result.get("teacher_note", "")))
    audit_log = list(result.get("audit_log", []))
    audit_log.append({
        "timestamp": approved_at,
        "action": "approved",
        "note": result["teacher_note"],
    })
    result["audit_log"] = audit_log
    save_result(project_id, round_id, result, team_num)
    return jsonify({"success": True, "data": result})


@app.route("/api/projects/<project_id>/delete-result/<int:team_num>", methods=["DELETE"])
def api_delete_result(project_id, team_num):
    round_id = request.args.get("round", type=int) or get_latest_round_id(project_id)
    round_dir = get_round_dir(project_id, round_id)
    filepath = round_dir / f"team_{team_num:03d}.json"
    if filepath.exists():
        filepath.unlink()
        return jsonify({"success": True})
    return jsonify({"error": "파일을 찾을 수 없습니다."}), 404


# ═══════════════════════════════════════════════════════════
# 종합 분석 (다회차 평균 / 수동 채점 비교)
# ═══════════════════════════════════════════════════════════

@app.route("/api/projects/<project_id>/manual-scores/upload", methods=["POST"])
def api_upload_manual(project_id):
    """수동 채점 Excel 업로드"""
    if 'file' not in request.files:
        return jsonify({"error": "파일이 없습니다."}), 400
    file = request.files['file']
    save_path = PROJECTS_DIR / project_id / "manual_scores.xlsx"
    file.save(str(save_path))
    try:
        from openpyxl import load_workbook
        wb = load_workbook(save_path, data_only=True)
        ws = wb.active
        header_map = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                header_map[str(val).strip()] = col
        num_col = score_col = None
        for name, col in header_map.items():
            if any(k in name for k in ["번호", "팀번호", "참가번호", "심사번호", "No"]):
                num_col = col
            if any(k in name for k in ["점수", "총점", "수동", "채점", "score"]):
                score_col = col
        if not num_col or not score_col:
            return jsonify({"error": f"번호/점수 컬럼을 찾을 수 없습니다. 헤더: {list(header_map.keys())}"}), 400
        manual_data = {}
        for row in range(2, ws.max_row + 1):
            t_num = ws.cell(row=row, column=num_col).value
            t_score = ws.cell(row=row, column=score_col).value
            if t_num is not None and t_score is not None:
                try:
                    manual_data[int(t_num)] = float(t_score)
                except (ValueError, TypeError):
                    pass
        json_path = PROJECTS_DIR / project_id / "manual_scores.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(manual_data, f, ensure_ascii=False)
        return jsonify({"count": len(manual_data)})
    except Exception as e:
        return jsonify({"error": f"파싱 실패: {str(e)[:200]}"}), 500


@app.route("/api/projects/<project_id>/manual-scores")
def api_get_manual(project_id):
    json_path = PROJECTS_DIR / project_id / "manual_scores.json"
    if json_path.exists():
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return jsonify({"count": len(data)})
    return jsonify({"count": 0})


@app.route("/api/projects/<project_id>/analysis")
def api_analysis(project_id):
    """다회차 종합 분석"""
    import statistics
    rounds_param = request.args.get("rounds", "")
    include_manual = request.args.get("include_manual", "false").lower() == "true"
    round_ids = [int(r) for r in rounds_param.split(",") if r.strip().isdigit()]
    manual_scores = {}
    json_path = PROJECTS_DIR / project_id / "manual_scores.json"
    if json_path.exists():
        with open(json_path, "r", encoding="utf-8") as f:
            manual_scores = {int(k): v for k, v in json.load(f).items()}
    team_data = {}
    for r_id in round_ids:
        round_dir = get_round_dir(project_id, r_id)
        if not round_dir.exists():
            continue
        for json_file in round_dir.glob("team_*.json"):
            try:
                with open(json_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                t_num = data.get("team_number")
                if t_num is None:
                    continue
                if t_num not in team_data:
                    team_data[t_num] = {"team_number": t_num, "team_name": data.get("team_name", f"참가자 {t_num}"), "scores": {}, "comments": []}
                team_data[t_num]["scores"][f"{r_id}회"] = data.get("total_score", 0)
                if data.get("overall_comment"):
                    team_data[t_num]["comments"].append(data["overall_comment"])
            except Exception:
                continue
    for t_num, score in manual_scores.items():
        if t_num not in team_data:
            team_data[t_num] = {"team_number": t_num, "team_name": f"참가자 {t_num}", "scores": {}, "comments": []}
        team_data[t_num]["manual_score"] = score
    results = []
    for t_num, t_info in team_data.items():
        calc_scores = list(t_info["scores"].values())
        if include_manual and "manual_score" in t_info:
            calc_scores.append(t_info["manual_score"])
        if not calc_scores:
            continue
        avg = sum(calc_scores) / len(calc_scores)
        std_dev = statistics.pstdev(calc_scores) if len(calc_scores) > 1 else 0.0
        trimmed_avg = avg
        is_trimmed = False
        if len(calc_scores) >= 5:
            trimmed_list = sorted(calc_scores)[1:-1]
            trimmed_avg = sum(trimmed_list) / len(trimmed_list)
            is_trimmed = True
        results.append({
            "team_number": t_num, "team_name": t_info["team_name"],
            "manual_score": t_info.get("manual_score"),
            "scores_by_round": t_info["scores"], "score_count": len(calc_scores),
            "average": round(avg, 2), "trimmed_average": round(trimmed_avg, 2),
            "is_trimmed": is_trimmed, "std_dev": round(std_dev, 2),
            "sample_comment": t_info["comments"][-1] if t_info["comments"] else ""
        })
    results.sort(key=lambda x: x["trimmed_average"] if x["is_trimmed"] else x["average"], reverse=True)
    return jsonify(results)


@app.route("/api/projects/<project_id>/analysis/excel")
def api_analysis_excel(project_id):
    """분석 결과 Excel 다운로드"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    import statistics
    
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    
    rounds_param = request.args.get("rounds", "")
    include_manual = request.args.get("include_manual", "false").lower() == "true"
    round_ids = [int(r) for r in rounds_param.split(",") if r.strip().isdigit()]
    manual_scores = {}
    json_path = PROJECTS_DIR / project_id / "manual_scores.json"
    if json_path.exists():
        with open(json_path, "r", encoding="utf-8") as f:
            manual_scores = {int(k): v for k, v in json.load(f).items()}
    
    team_data = {}
    for r_id in round_ids:
        round_dir = get_round_dir(project_id, r_id)
        if not round_dir.exists():
            continue
        for jf in round_dir.glob("team_*.json"):
            try:
                with open(jf, "r", encoding="utf-8") as f:
                    data = json.load(f)
                t_num = data.get("team_number")
                if t_num is None: continue
                if t_num not in team_data:
                    team_data[t_num] = {"team_number": t_num, "team_name": data.get("team_name", ""), "scores": {}, "comments": [], "seteuk": ""}
                team_data[t_num]["scores"][f"{r_id}회"] = data.get("total_score", 0)
                if data.get("overall_comment"):
                    team_data[t_num]["comments"].append(data["overall_comment"])
                if data.get("seteuk"):
                    team_data[t_num]["seteuk"] = data["seteuk"]
            except Exception: continue
    for t_num, score in manual_scores.items():
        if t_num not in team_data:
            team_data[t_num] = {"team_number": t_num, "team_name": f"참가자 {t_num}", "scores": {}, "comments": [], "seteuk": ""}
        team_data[t_num]["manual_score"] = score
    
    all_round_keys = sorted(set(k for td in team_data.values() for k in td["scores"].keys()))
    results = []
    for t_num, t_info in team_data.items():
        calc_scores = list(t_info["scores"].values())
        if include_manual and "manual_score" in t_info:
            calc_scores.append(t_info["manual_score"])
        if not calc_scores: continue
        avg = sum(calc_scores) / len(calc_scores)
        std_dev = statistics.pstdev(calc_scores) if len(calc_scores) > 1 else 0.0
        trimmed_avg = avg
        if len(calc_scores) >= 5:
            trimmed_list = sorted(calc_scores)[1:-1]
            trimmed_avg = sum(trimmed_list) / len(trimmed_list)
        results.append({**t_info, "average": round(avg, 2), "trimmed_average": round(trimmed_avg, 2), "std_dev": round(std_dev, 2)})
    results.sort(key=lambda x: x["average"], reverse=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "종합 분석"
    hfont = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    dfont = Font(name="맑은 고딕", size=10)
    ca = Alignment(horizontal="center", vertical="center")
    bd = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    headers = ["순위", "번호", "이름"] + all_round_keys + ["수동채점", "평균", "절사평균", "편차", "차이(AI-수동)", "종합의견", "세특"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont; cell.fill = hfill; cell.alignment = ca; cell.border = bd
    for rank, r in enumerate(results, 1):
        manual = r.get("manual_score")
        diff = round(r["average"] - manual, 2) if manual is not None else ""
        vals = [rank, r["team_number"], r["team_name"]]
        for rk in all_round_keys:
            vals.append(r["scores"].get(rk, ""))
        vals += [manual if manual is not None else "", r["average"], r["trimmed_average"], r["std_dev"], diff, r["comments"][-1] if r.get("comments") else "", r.get("seteuk", "")]
        row = rank + 1
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = dfont; cell.alignment = ca; cell.border = bd
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = PROJECTS_DIR / project_id / "results" / f"종합분석_{timestamp}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return send_file(output_path, as_attachment=True, download_name=f"종합분석_{timestamp}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")




REPORT_RUBRIC_SCHEMA = {
    "type": "object",
    "properties": {
        "categories": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                    "criteria": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "name": {"type": "string"},
                                "description": {"type": "string"},
                                "scale": {"type": "array", "items": {"type": "integer"}},
                                "scale_labels": {"type": "array", "items": {"type": "string"}},
                                "required_elements": {"type": "array", "items": {"type": "string"}},
                                "deduction_rules": {"type": "array", "items": {"type": "string"}},
                                "exceptions": {"type": "array", "items": {"type": "string"}},
                                "feedback_focus": {"type": "string"},
                                "core_criteria": {"type": "array", "items": {"type": "string"}},
                            },
                            "required": ["name", "description", "scale", "scale_labels"],
                        },
                    },
                },
                "required": ["name", "criteria"],
            },
        },
        "prompt_template": {"type": "string"},
    },
    "required": ["categories"],
}


def _normalize_report_rubric(result: dict) -> dict:
    categories = result.get("categories", [])
    cid = 1
    for category in categories:
        for criterion in category.get("criteria", []):
            criterion["id"] = f"c{cid}"
            cid += 1
            scale = [int(value) for value in criterion.get("scale", [5, 4, 3, 2, 1])]
            criterion["scale"] = sorted(set(scale), reverse=True)
            labels = criterion.get("scale_labels", [])
            if len(labels) != len(criterion["scale"]):
                labels = [f"등급{i + 1}" for i in range(len(criterion["scale"]))]
            criterion["scale_labels"] = labels
            for field_name in (
                "required_elements", "deduction_rules", "exceptions", "core_criteria"
            ):
                criterion[field_name] = [
                    str(value).strip()
                    for value in criterion.get(field_name, [])
                    if str(value).strip()
                ]
            criterion["feedback_focus"] = str(
                criterion.get("feedback_focus", "")
            )
    return {"categories": categories, "prompt_template": result.get("prompt_template", "")}


@app.route("/api/projects/<project_id>/rubric/extract", methods=["POST"])
def api_extract_rubric(project_id):
    """업로드된 문서에서 AI로 채점 기준(루브릭) 추출"""
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    
    if 'file' not in request.files:
        return jsonify({"error": "파일이 없습니다."}), 400
    
    file = request.files['file']
    if not file.filename:
        return jsonify({"error": "파일이 선택되지 않았습니다."}), 400

    if config.ai_provider != "gemini_api":
        temp_dir = PROJECTS_DIR / project_id / "temp"
        temp_dir.mkdir(exist_ok=True)
        suffix = Path(file.filename).suffix.lower()
        temp_path = temp_dir / f"rubric_web_{uuid.uuid4().hex}{suffix}"
        file.save(str(temp_path))
        prompt = (
            "첨부 문서는 보고서·대회 평가용 채점 기준 또는 운영 요강입니다. "
            "평가 영역, 세부 항목, 배점 척도와 등급별 구체적 기준을 구조화하세요. "
            "문서에 배점이 있으면 그대로 사용하고, 없으면 합리적인 5단계 척도를 만드세요. "
            "prompt_template에는 이 문서의 맥락에 맞는 한국어 채점 지침을 작성하세요."
        )
        try:
            provider = get_provider(config, load_api_keys())
            result = provider.generate_json(
                prompt,
                schema=REPORT_RUBRIC_SCHEMA,
                files=[str(temp_path)],
                model_name=config.ai_model,
                temperature=0.1,
            )
            return jsonify(_normalize_report_rubric(result))
        except ProviderNeedsUserAction as e:
            return jsonify({"error": str(e), "needs_user_action": True}), 409
        except Exception as e:
            return jsonify({"error": f"AI 추출 실패: {str(e)[:300]}"}), 500
        finally:
            if temp_path.exists():
                temp_path.unlink()
    
    # API 키 확인
    keys = load_api_keys()
    api_key = keys.get("google", "")
    if not api_key:
        return jsonify({"error": "Google API 키가 설정되지 않았습니다."}), 400
    
    # 임시 파일 저장 (고유 파일명으로 잠금 충돌 방지)
    import tempfile, shutil, time
    temp_dir = PROJECTS_DIR / project_id / "temp"
    temp_dir.mkdir(exist_ok=True)
    suffix = Path(file.filename).suffix.lower()
    uid = str(int(time.time() * 1000))
    temp_path = temp_dir / f"rubric_{uid}{suffix}"
    file.save(str(temp_path))
    
    # HWP/HWPX → PDF 자동 변환
    if suffix in (".hwp", ".hwpx"):
        pdf_path = temp_dir / f"rubric_{uid}.pdf"
        hwp = None
        try:
            import win32com.client
            hwp = win32com.client.Dispatch("HWPFrame.HwpObject")
            hwp.XHwpWindows.Item(0).Visible = False
            hwp.RegisterModule("FilePathCheckDLL", "SecurityModule")
            try:
                hwp.SetAltMode(0)
            except Exception:
                pass
            hwp.Open(str(temp_path), "HWP", "forceopen:true;versionwarning:false")
            hwp.HAction.GetDefault("FileSaveAs_S", hwp.HParameterSet.HFileOpenSave.HSet)
            hwp.HParameterSet.HFileOpenSave.filename = str(pdf_path)
            hwp.HParameterSet.HFileOpenSave.Format = "PDF"
            hwp.HAction.Execute("FileSaveAs_S", hwp.HParameterSet.HFileOpenSave.HSet)
            hwp.Clear(False)
        except Exception as e:
            return jsonify({"error": f"HWP → PDF 변환 실패: {str(e)[:200]}\n한컴오피스 한글이 설치되어 있는지 확인하세요."}), 400
        finally:
            if hwp:
                try:
                    hwp.Quit()
                except Exception:
                    pass
        
        if not pdf_path.exists():
            return jsonify({"error": "HWP → PDF 변환에 실패했습니다."}), 400
        temp_path = pdf_path
    
    try:
        from google import genai
        from google.genai import types
        
        client = genai.Client(api_key=api_key)
        
        # 파일 업로드
        print(f"[루브릭 추출] 파일 업로드 시작: {file.filename}")
        uploaded = client.files.upload(file=str(temp_path))
        
        # 처리 대기 (영상 등 대용량 파일)
        import time
        while uploaded.state and uploaded.state.name == "PROCESSING":
            print(f"[루브릭 추출] 파일 처리 중...")
            time.sleep(2)
            uploaded = client.files.get(name=uploaded.name)
        
        print(f"[루브릭 추출] AI 분석 요청 중...")
        
        # 프로젝트 설정에서 고른 API 모델을 루브릭 추출에도 동일하게 사용한다.
        extract_model = config.ai_model
        
        # AI에게 루브릭 추출 요청
        extraction_prompt = """이 문서는 채점 기준 또는 대회 요강입니다.
문서에서 채점에 사용할 평가 영역, 세부 항목, 배점을 추출하세요.

반드시 아래 JSON 형식으로만 응답하세요. 다른 텍스트 없이 JSON만 출력하세요:
{
  "categories": [
    {
      "name": "영역 이름",
      "criteria": [
        {
          "name": "항목 이름",
          "description": "이 항목의 평가 기준 상세 설명",
          "scale": [5, 4, 3, 2, 1],
          "scale_labels": ["매우우수", "우수", "보통", "미흡", "매우미흡"]
        }
      ]
    }
  ],
  "prompt_template": "이 문서의 채점 맥락에 맞는 평가 프롬프트 (한국어)"
}

규칙:
1. 문서에 명시된 배점 체계가 있으면 그대로 사용하세요.
2. 배점이 없으면 적절히 5점 척도로 만드세요.
3. scale은 반드시 내림차순 정수 배열이어야 합니다.
4. scale_labels는 scale과 같은 길이여야 합니다.
5. description에는 각 등급별 구체적 기준을 포함하세요.
6. prompt_template에는 채점 시 AI가 참고할 전체 지침을 작성하세요."""

        response = client.models.generate_content(
            model=extract_model,
            contents=[uploaded, extraction_prompt],
            config=types.GenerateContentConfig(temperature=0.1),
        )
        print(f"[루브릭 추출] AI 응답 수신 완료")
        
        # JSON 파싱
        import re
        text = response.text.strip()
        # ```json ... ``` 블록 추출
        json_match = re.search(r'```(?:json)?\s*\n?(.*?)\n?```', text, re.DOTALL)
        if json_match:
            text = json_match.group(1).strip()
        
        data = json.loads(text)
        categories = data.get("categories", [])
        
        # ID 자동 부여
        cid = 1
        for cat in categories:
            for criterion in cat.get("criteria", []):
                criterion["id"] = f"c{cid}"
                cid += 1
                # scale_labels 누락 시 기본값
                if "scale_labels" not in criterion:
                    scale = criterion.get("scale", [5,4,3,2,1])
                    criterion["scale_labels"] = [f"등급{i+1}" for i in range(len(scale))]
        
        return jsonify({
            "categories": categories,
            "prompt_template": data.get("prompt_template", ""),
        })
        
    except json.JSONDecodeError as e:
        return jsonify({"error": f"AI 응답을 JSON으로 파싱할 수 없습니다: {str(e)[:200]}"}), 500
    except Exception as e:
        return jsonify({"error": f"AI 추출 실패: {str(e)[:300]}"}), 500
    finally:
        if temp_path.exists():
            temp_path.unlink()


@app.route("/api/projects/<project_id>/rubric/generate", methods=["POST"])
def api_generate_rubric(project_id):
    """텍스트 설명에서 AI로 채점 기준(루브릭) 생성"""
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    
    data = request.json or {}
    text_description = data.get("text", "").strip()
    if not text_description:
        return jsonify({"error": "루브릭 설명 텍스트를 입력하세요."}), 400

    if config.ai_provider != "gemini_api":
        prompt = (
            "사용자가 보고서 평가용 루브릭을 간단히 설명했습니다. 이를 평가 영역, 세부 항목, "
            "내림차순 정수 배점 척도, 척도별 구체적 기준으로 구조화하세요.\n\n사용자 설명:\n"
            + text_description
        )
        try:
            provider = get_provider(config, load_api_keys())
            result = provider.generate_json(
                prompt,
                schema=REPORT_RUBRIC_SCHEMA,
                model_name=config.ai_model,
                temperature=0.1,
            )
            return jsonify(_normalize_report_rubric(result))
        except ProviderNeedsUserAction as e:
            return jsonify({"error": str(e), "needs_user_action": True}), 409
        except Exception as e:
            return jsonify({"error": f"AI 생성 실패: {str(e)[:300]}"}), 500
    
    # API 키 확인
    keys = load_api_keys()
    api_key = keys.get("google", "")
    if not api_key:
        return jsonify({"error": "Google API 키가 설정되지 않았습니다."}), 400
    
    try:
        from google import genai
        from google.genai import types
        
        client = genai.Client(api_key=api_key)
        
        # 프로젝트 설정에서 고른 API 모델을 루브릭 생성에도 동일하게 사용한다.
        generate_model = config.ai_model
        
        generation_prompt = f"""사용자가 채점 기준(루브릭)을 텍스트로 설명했습니다. 이를 구조화된 JSON으로 변환하세요.

사용자 설명:
{text_description}

반드시 아래 JSON 형식으로만 응답하세요. 다른 텍스트 없이 JSON만 출력하세요:
{{
  "categories": [
    {{
      "name": "영역 이름",
      "criteria": [
        {{
          "name": "항목 이름",
          "description": "이 항목의 평가 기준 상세 설명",
          "scale": [10, 8, 6],
          "scale_labels": ["우수", "보통", "미흡"]
        }}
      ]
    }}
  ]
}}

규칙:
1. 사용자가 명시한 영역명, 항목명, 배점, 급간을 정확히 반영하세요.
2. scale은 반드시 내림차순 정수 배열이어야 합니다.
3. scale_labels는 scale과 같은 길이여야 합니다.
4. 사용자가 급간을 명시하지 않으면 적절한 등간격으로 만드세요.
5. description에는 각 등급별 구체적 기준을 포함하세요.
6. 사용자가 언급하지 않은 세부사항은 합리적으로 추론하세요."""
        
        print(f"[루브릭 생성] AI 생성 요청 중...")
        
        response = client.models.generate_content(
            model=generate_model,
            contents=[generation_prompt],
            config=types.GenerateContentConfig(temperature=0.1),
        )
        print(f"[루브릭 생성] AI 응답 수신 완료")
        
        import re
        text = response.text.strip()
        json_match = re.search(r'```(?:json)?\s*\n?(.*?)\n?```', text, re.DOTALL)
        if json_match:
            text = json_match.group(1).strip()
        
        result = json.loads(text)
        categories = result.get("categories", [])
        
        # ID 자동 부여
        cid = 1
        for cat in categories:
            for criterion in cat.get("criteria", []):
                criterion["id"] = f"c{cid}"
                cid += 1
                if "scale_labels" not in criterion:
                    scale = criterion.get("scale", [5, 4, 3, 2, 1])
                    criterion["scale_labels"] = [f"등급{i+1}" for i in range(len(scale))]
        
        return jsonify({"categories": categories})
        
    except json.JSONDecodeError as e:
        return jsonify({"error": f"AI 응답을 JSON으로 파싱할 수 없습니다: {str(e)[:200]}"}), 500
    except Exception as e:
        return jsonify({"error": f"AI 생성 실패: {str(e)[:300]}"}), 500


# ═══════════════════════════════════════════════════════════
# 제출용 채점표 (수동 순위 조정)
# ═══════════════════════════════════════════════════════════

def _get_valid_totals(config):
    """루브릭 세부 항목 조합으로 실제 나올 수 있는 총점 집합 (DP)"""
    possible = {0}
    for criterion in config.all_criteria:
        new_possible = set()
        for s in possible:
            for v in criterion.scale:
                new_possible.add(s + v)
        possible = new_possible
    return sorted(possible, reverse=True)  # 내림차순


def _pyramid_distribution(num_teams, valid_scores):
    """
    삼각분포형 점수 배분 (항아리형):
    valid_scores: 유효한 총점 목록 (내림차순 정렬)
    - 1~3위: 고유 점수
    - 하위 40% 점수 위치에서 최대 인원 (최빈값)
    - 그 아래로 다시 줄어듦
    """
    if num_teams <= 0: return []
    if not valid_scores: return [0] * num_teams
    if num_teams == 1: return [valid_scores[0]]
    
    num_levels = len(valid_scores)
    if num_levels == 1: return [valid_scores[0]] * num_teams
    
    if num_teams <= num_levels:
        # 팀 수 ≤ 유효 점수 수 → 균등 간격으로 선택
        indices = [round(i * (num_levels - 1) / (num_teams - 1)) for i in range(num_teams)]
        return [valid_scores[idx] for idx in indices]
    
    # 상위 3위: 고유 점수 (유효 점수 중 상위 3개)
    unique_top = min(3, num_levels)
    remaining_teams = num_teams - unique_top
    remaining_levels = num_levels - unique_top
    
    if remaining_levels <= 0:
        per = num_teams // num_levels
        extra = num_teams % num_levels
        scores = []
        for i in range(num_levels):
            scores.extend([valid_scores[i]] * (per + (1 if i < extra else 0)))
        return scores[:num_teams]
    
    scores = [valid_scores[i] for i in range(unique_top)]
    
    # 삼각분포 가중치: 최빈값을 60% 지점 (= 하위 40%)에 배치
    peak_t = 0.6
    weights = []
    for i in range(remaining_levels):
        t = i / max(1, remaining_levels - 1) if remaining_levels > 1 else 0.5
        if t <= peak_t:
            w = t / peak_t
        else:
            w = (1 - t) / (1 - peak_t)
        weights.append(max(0.05, w))
    
    total_w = sum(weights)
    allocated = 0
    for i in range(remaining_levels):
        sv = valid_scores[unique_top + i]
        if i == remaining_levels - 1:
            size = remaining_teams - allocated
        else:
            size = max(1, round(remaining_teams * weights[i] / total_w))
            size = min(size, remaining_teams - allocated)
        if size > 0:
            scores.extend([sv] * size)
            allocated += size
    
    return scores[:num_teams]


def _snap_to_valid(value, valid_scores):
    """값을 가장 가까운 유효 점수로 스냅"""
    return min(valid_scores, key=lambda v: abs(v - value))

@app.route("/api/projects/<project_id>/submission/preview", methods=["POST"])
def api_submission_preview(project_id):
    """순위와 고정 점수를 기반으로 세부 점수를 생성하고 JSON으로 반환 (미리보기)"""
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    
    data = request.json
    ordered_teams = data.get("teams", [])  # [{team_number, team_name, total_score(or null)}]
    
    if not ordered_teams:
        return jsonify({"error": "팀 목록이 없습니다."}), 400
    
    num_teams = len(ordered_teams)
    
    # 유효 총점 계산 (루브릭 기반)
    valid_scores = _get_valid_totals(config)
    max_valid = valid_scores[0]
    min_valid = valid_scores[-1]
    
    # ── 1단계: 고정 점수 수집 (유효 점수로 스냅) ──
    anchors = {}
    for i, entry in enumerate(ordered_teams):
        val = entry.get("total_score")
        if val is not None:
            anchors[i] = _snap_to_valid(int(round(float(val))), valid_scores)
    
    # ── 2단계: 점수 생성 ──
    if not anchors:
        # 고정 점수 없음 → 삼각분포 배분 (유효 점수만 사용)
        target_scores = _pyramid_distribution(num_teams, valid_scores)
    else:
        # 고정 점수 있음 → 구간별 선형 보간 후 유효 점수로 스냅
        target_scores = [None] * num_teams
        for pos, score in anchors.items():
            target_scores[pos] = score
        
        all_anchors = {-1: max_valid, num_teams: min_valid}
        all_anchors.update(anchors)
        sorted_positions = sorted(all_anchors.keys())
        
        for k in range(len(sorted_positions) - 1):
            sp, ep = sorted_positions[k], sorted_positions[k + 1]
            ss, es = all_anchors[sp], all_anchors[ep]
            gaps = [p for p in range(max(0, sp + 1), min(num_teams, ep))
                    if target_scores[p] is None]
            if not gaps:
                continue
            intervals = ep - sp
            for pos in gaps:
                t = (pos - sp) / intervals
                raw = ss + t * (es - ss)
                target_scores[pos] = _snap_to_valid(int(round(raw)), valid_scores)
        
        # 내림차순 보정
        for i in range(1, num_teams):
            if target_scores[i] is not None and target_scores[i - 1] is not None:
                if target_scores[i] > target_scores[i - 1]:
                    # 현재 값 이하인 유효 점수 중 가장 큰 것
                    cands = [v for v in valid_scores if v <= target_scores[i - 1]]
                    target_scores[i] = cands[0] if cands else min_valid
    
    # ── 4단계: 세부 항목 점수 역생성 ──
    results = []
    for i, entry in enumerate(ordered_teams):
        target = target_scores[i]
        
        # 세부 점수 생성 (최대 3회 시도하여 총점 일치 보장)
        for _attempt in range(3):
            fake_scores = generate_fake_subscores_dynamic(config, target)
            # 세부 점수 합계 검증
            actual_sum = sum(fake_scores.get(c.id, 0) for c in config.all_criteria)
            if actual_sum == target:
                break
        
        result_entry = {
            "team_number": entry["team_number"],
            "team_name": entry.get("team_name", f"참가자 {entry['team_number']}"),
            **fake_scores,
            "overall_comment": "",
        }
        # 소계 재계산 (일관성 보장)
        for cat in config.categories:
            cat_total = sum(result_entry.get(c.id, 0) for c in cat.criteria)
            cat_key = cat.name.replace(" ", "_") + "_total"
            result_entry[cat_key] = cat_total
        # 총점은 반드시 target 사용 (순위 보장)
        result_entry["total_score"] = target
        result_entry["rank"] = i + 1
        results.append(result_entry)
    
    # 카테고리 정보도 함께 반환 (미리보기 테이블 렌더링용)
    cat_info = []
    for cat in config.categories:
        cat_info.append({
            "name": cat.name,
            "max_score": cat.max_score,
            "criteria": [{"id": c.id, "name": c.name, "max_score": c.max_score} for c in cat.criteria]
        })
    
    return jsonify({"results": results, "categories": cat_info})


@app.route("/api/projects/<project_id>/submission/excel", methods=["POST"])
def api_submission_excel(project_id):
    """미리보기에서 확인된 결과 데이터를 Excel로 내보내기"""
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    
    results = request.json.get("results", [])
    if not results:
        return jsonify({"error": "데이터가 없습니다."}), 400
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = PROJECTS_DIR / project_id / "results" / f"제출용_채점표_{timestamp}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    generate_excel(config, results, output_path)
    
    return send_file(str(output_path), as_attachment=True, download_name=output_path.name)


# ═══════════════════════════════════════════════════════════
# Excel 다운로드
# ═══════════════════════════════════════════════════════════

@app.route("/api/projects/<project_id>/download")
def api_download(project_id):
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    
    round_id = request.args.get("round", type=int)
    completed = load_completed(project_id, round_id)
    if not completed:
        return jsonify({"error": "채점 결과가 없습니다."}), 404
    
    results = sorted(completed.values(), key=lambda x: x.get("total_score", 0), reverse=True)
    for i, r in enumerate(results, 1):
        r["rank"] = i
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    round_label = f"{round_id}회차" if round_id else "최신"
    output_path = PROJECTS_DIR / project_id / "results" / f"{config.name}_{round_label}_채점결과_{timestamp}.xlsx"
    generate_excel(config, results, output_path)
    
    return send_file(str(output_path), as_attachment=True, download_name=output_path.name)


# ═══════════════════════════════════════════════════════════
# PDF 원문 보기
# ═══════════════════════════════════════════════════════════

@app.route("/api/projects/<project_id>/file/<int:team_num>/<file_type>")
def api_view_file(project_id, team_num, file_type):
    config = load_project(project_id)
    if not config:
        return jsonify({"error": "프로젝트를 찾을 수 없습니다."}), 404
    
    participant = get_participant_files(config, team_num)
    for f in participant.get("files", []):
        if f["type"] == file_type or f["ext"].strip(".") == file_type:
            return send_file(f["path"])
    
    return jsonify({"error": "파일을 찾을 수 없습니다."}), 404


# ═══════════════════════════════════════════════════════════
# 실행
# ═══════════════════════════════════════════════════════════

if __name__ == "__main__":
    port = 5000
    
    # 포트 충돌 시 자동 증가
    import socket
    for p in range(port, port + 10):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            if s.connect_ex(("127.0.0.1", p)) != 0:
                port = p
                break
    
    import io, sys
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    print()
    print("=" * 50)
    print("  AI Grading System")
    print(f"  http://localhost:{port}")
    print("=" * 50)
    print()
    
    if os.environ.get("AI_GRADER_NO_BROWSER") != "1":
        webbrowser.open(f"http://localhost:{port}")
    app.run(host="127.0.0.1", port=port, debug=False)
