"""추가 입력 없이 만드는 익명 개발 피드백 묶음.

원본 파일과 실제 학생 대응표는 ZIP에 넣지 않는다. ZIP은 허용된 통계·진단
필드만 새로 구성하고, 생성 뒤 텍스트와 XLSX 내부 XML까지 개인정보 흔적을
검사한다.
"""

from __future__ import annotations

import csv
import io
import json
import re
import secrets
import unicodedata
import uuid
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models.project import ProjectConfig
from services import grading as grading_service
from services.analysis import build_analysis_summary
from services.file_manager import find_materials
from services.grading import list_round_summaries, load_completed, load_round_metadata
from services.review import build_review_dashboard, scoring_items


FEEDBACK_SCHEMA_VERSION = 1
BUNDLE_FILENAMES = {
    "성능요약.xlsx",
    "채점데이터.jsonl",
    "실행진단.json",
    "자료목록.csv",
    "안내.txt",
}
TEXT_ENTRY_NAMES = BUNDLE_FILENAMES - {"성능요약.xlsx"}
EMAIL_PATTERN = re.compile(r"[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}")
PHONE_PATTERN = re.compile(
    r"(?<!\d)(?:01[016789][-\s]?\d{3,4}[-\s]?\d{4}|"
    r"0\d{1,2}[-\s]?\d{3,4}[-\s]?\d{4})(?!\d)"
)
ADDRESS_PATTERN = re.compile(
    r"(?:서울특별시|부산광역시|대구광역시|인천광역시|광주광역시|"
    r"대전광역시|울산광역시|세종특별자치시|경기도|강원특별자치도|"
    r"충청북도|충청남도|전북특별자치도|전라남도|경상북도|경상남도|"
    r"제주특별자치도)\s+[가-힣0-9\s-]{2,50}"
)
ACCOUNT_PATTERN = re.compile(
    r"(?<!\d)(?!(?:19|20)\d{2}-\d{1,2}-\d{1,2}(?!\d))"
    r"(?:\d{2,6}[-\s]){2,3}\d{2,6}(?!\d)|"
    r"(?<!\d)\d{10,16}(?!\d)"
)
WINDOWS_PATH_PATTERN = re.compile(
    r"(?i)(?:[A-Z]:\\|\\\\)[^\r\n\t\"'<>|]+"
)
API_KEY_PATTERN = re.compile(
    r"(?i)(?:AIza[0-9A-Za-z_-]{20,}|sk-[0-9A-Za-z_-]{16,}|"
    r"(?:api[_ -]?key|token)\s*[:=]\s*[0-9A-Za-z_-]{12,})"
)
SCHOOL_PATTERN = re.compile(
    r"[가-힣A-Za-z0-9·]{2,30}(?:초등학교|중학교|고등학교|대학교)"
)


class FeedbackExportError(ValueError):
    pass


def _now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def _project_dir(project_id: str) -> Path:
    return grading_service.PROJECTS_DIR / project_id


def _participant_numbers(config: ProjectConfig) -> list[int]:
    numbers = {student.number for student in config.roster_students}
    try:
        numbers.update(int(item["number"]) for item in find_materials(config))
    except (OSError, ValueError):
        pass
    for summary in list_round_summaries(config):
        numbers.update(load_completed(config.id, summary["id"]).keys())
    return sorted(number for number in numbers if number > 0)


def _privacy_sources(config: ProjectConfig, numbers: list[int]) -> dict:
    roster = {student.number: student for student in config.roster_students}
    names_by_number: dict[int, set[str]] = {number: set() for number in numbers}
    identifiers: set[str] = set()
    for number, student in roster.items():
        if student.name.strip():
            names_by_number.setdefault(number, set()).add(student.name.strip())
            identifiers.add(student.name.strip())
        for value in (student.grade, student.class_name, student.student_id):
            if str(value).strip():
                identifiers.add(str(value).strip())
    try:
        participants = find_materials(config)
    except (OSError, ValueError):
        participants = []
    material_files = []
    for participant in participants:
        number = int(participant["number"])
        name = str(participant.get("name", "")).strip()
        if name:
            names_by_number.setdefault(number, set()).add(name)
            identifiers.add(name)
        for file_info in participant.get("files", []):
            path = Path(file_info["path"])
            material_files.append({
                "number": number,
                "path": path,
                "type": file_info.get("type", ""),
                "ext": file_info.get("ext", path.suffix.lower()),
                "size_mb": file_info.get("size_mb", 0),
            })
            identifiers.update({path.name, path.stem, str(path)})
    for summary in list_round_summaries(config):
        for number, result in load_completed(config.id, summary["id"]).items():
            name = str(result.get("team_name", "")).strip()
            if name:
                names_by_number.setdefault(int(number), set()).add(name)
                identifiers.add(name)
    identifiers.update({
        config.name,
        config.description,
        config.setup.target,
        config.setup.assessment_name,
        config.materials.folder_path,
        config.exam.question_source_path,
        config.exam.rubric_source_path,
        config.exam.scan_split.source_path,
    })
    identifiers = {value.strip() for value in identifiers if str(value).strip()}
    return {
        "roster": roster,
        "names_by_number": names_by_number,
        "identifiers": identifiers,
        "material_files": material_files,
    }


def _build_aliases(numbers: list[int], export_id: str) -> dict[int, dict]:
    shuffled = list(numbers)
    secrets.SystemRandom().shuffle(shuffled)
    student_index_by_number = {
        number: index
        for index, number in enumerate(shuffled, 1)
    }
    shuffled_seats = list(range(1, len(numbers) + 1))
    secrets.SystemRandom().shuffle(shuffled_seats)
    seat_by_number = {
        number: shuffled_seats[index - 1]
        for index, number in enumerate(sorted(numbers), 1)
    }
    grade = f"가상학년-{secrets.randbelow(90) + 10}"
    class_name = f"가상반-{secrets.randbelow(90) + 10}"
    return {
        number: {
            "student": f"S-{student_index_by_number[number]:03d}",
            "grade": grade,
            "class": class_name,
            "seat": f"가상번호-{seat_by_number[number]:03d}",
            "export_id": export_id,
        }
        for number in sorted(numbers)
    }


def _scrubber(
    aliases: dict[int, dict],
    privacy: dict,
):
    replacements = []
    grade_aliases = {}
    class_aliases = {}
    for number, names in privacy["names_by_number"].items():
        alias = aliases.get(number, {}).get("student", "학생")
        replacements.extend((name, alias) for name in names if name)
    for number, student in privacy["roster"].items():
        alias = aliases.get(number, {}).get("student", "학생")
        if student.student_id:
            replacements.append((student.student_id, alias))
        if str(student.grade).strip():
            grade_aliases[str(student.grade).strip()] = aliases[number]["grade"]
        if str(student.class_name).strip():
            class_aliases[str(student.class_name).strip()] = aliases[number]["class"]
    for file_info in privacy["material_files"]:
        replacements.extend([
            (file_info["path"].name, "원본파일"),
            (file_info["path"].stem, "원본파일"),
            (str(file_info["path"]), "원본경로"),
        ])
    replacements.sort(key=lambda item: len(item[0]), reverse=True)

    def scrub(value) -> str:
        text = str(value or "")
        for source, replacement in replacements:
            text = text.replace(source, replacement)
        for number, alias in aliases.items():
            text = re.sub(
                rf"(?<!\d){re.escape(str(number))}\s*번",
                alias["student"],
                text,
            )
            text = re.sub(
                rf"(?:번호|학생)\s*[:#-]?\s*{re.escape(str(number))}(?!\d)",
                alias["student"],
                text,
            )
        for grade, alias in grade_aliases.items():
            text = re.sub(
                rf"(?<!\d){re.escape(grade)}\s*학년",
                alias,
                text,
            )
        for class_name, alias in class_aliases.items():
            text = re.sub(
                rf"(?<!\d){re.escape(class_name)}\s*반",
                alias,
                text,
            )
        text = EMAIL_PATTERN.sub("[이메일 삭제]", text)
        text = PHONE_PATTERN.sub("[전화번호 삭제]", text)
        text = ADDRESS_PATTERN.sub("[주소 삭제]", text)
        text = ACCOUNT_PATTERN.sub("[계정번호 삭제]", text)
        text = WINDOWS_PATH_PATTERN.sub("[경로 삭제]", text)
        text = API_KEY_PATTERN.sub("[인증정보 삭제]", text)
        text = SCHOOL_PATTERN.sub("[학교명 삭제]", text)
        return text

    return scrub


def _safe_score(value):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return round(number, 4) if number == number else None


def _round_item_value(result: dict, item_id: str, *, raw: bool):
    source = result.get("ai_original") if raw else result
    if raw and not isinstance(source, dict):
        source = result
    return _safe_score((source or {}).get(item_id))


def _round_total_value(result: dict, *, raw: bool):
    source = result.get("ai_original") if raw else result
    if raw and not isinstance(source, dict):
        source = result
    return _safe_score((source or {}).get("total_score"))


def _collect_export_data(
    config: ProjectConfig,
    *,
    export_id: str,
    scan_pdf_pages: bool = True,
) -> dict:
    numbers = _participant_numbers(config)
    privacy = _privacy_sources(config, numbers)
    aliases = _build_aliases(numbers, export_id)
    scrub = _scrubber(aliases, privacy)
    dashboard = build_review_dashboard(
        config,
        participant_numbers=numbers,
    )
    students = {
        student["team_number"]: student
        for student in dashboard["students"]
    }
    items = scoring_items(config)
    generic_items = {
        item["id"]: {
            "id": f"I-{index:02d}",
            "label": f"항목 {index:02d}",
            "max_score": item["max_score"],
        }
        for index, item in enumerate(items, 1)
    }

    records = []
    text_field_count = 0
    for summary in dashboard["rounds"]:
        round_id = summary["id"]
        metadata = load_round_metadata(config.id, round_id)
        context = metadata.get("execution_context", {})
        for number, result in load_completed(config.id, round_id).items():
            if int(number) not in aliases:
                continue
            student = students.get(int(number), {})
            decision = student.get("decision") or {}
            for item in items or [{
                "id": "total_score", "label": "총점", "max_score": dashboard["max_score"]
            }]:
                item_id = item["id"]
                generic = generic_items.get(item_id, {
                    "id": "TOTAL", "label": "총점", "max_score": item["max_score"]
                })
                raw_score = (
                    _round_total_value(result, raw=True)
                    if item_id == "total_score"
                    else _round_item_value(result, item_id, raw=True)
                )
                adjusted_score = (
                    _round_total_value(result, raw=False)
                    if item_id == "total_score"
                    else _round_item_value(result, item_id, raw=False)
                )
                summary_text = scrub(result.get(f"{item_id}_answer_summary", ""))
                reason_text = scrub(result.get(f"{item_id}_reason", ""))
                text_field_count += int(bool(summary_text)) + int(bool(reason_text))
                manual_item = next(
                    (
                        row.get("manual_score")
                        for row in student.get("items", [])
                        if row["id"] == item_id
                    ),
                    None,
                )
                final_item = (decision.get("item_scores") or {}).get(item_id)
                record = {
                    "schema_version": FEEDBACK_SCHEMA_VERSION,
                    "dataset_id": export_id,
                    "virtual_student_id": aliases[int(number)]["student"],
                    "virtual_grade": aliases[int(number)]["grade"],
                    "virtual_class": aliases[int(number)]["class"],
                    "virtual_seat": aliases[int(number)]["seat"],
                    "workflow_type": config.workflow_type,
                    "project_type": config.project_type,
                    "item_id": generic["id"],
                    "item_label": generic["label"],
                    "max_score": generic["max_score"],
                    "round_id": round_id,
                    "provider": context.get("provider") or summary.get("provider", ""),
                    "model": context.get("model") or summary.get("model", ""),
                    "criteria_version": (
                        context.get("criteria_version")
                        or summary.get("criteria_version", 0)
                    ),
                    "raw_ai_score": raw_score,
                    "adjusted_score": adjusted_score,
                    "score_adjusted": (
                        raw_score is not None
                        and adjusted_score is not None
                        and abs(raw_score - adjusted_score) > 0.001
                    ),
                    "confidence": _safe_score(result.get(f"{item_id}_confidence")),
                    "review_required": bool(
                        result.get(f"{item_id}_review_required", False)
                    ),
                    "answer_summary": summary_text,
                    "reason": reason_text,
                    "manual_score": manual_item,
                    "final_score": final_item,
                    "final_status": decision.get("status", "pending"),
                    "ai_manual_difference": (
                        round(raw_score - float(manual_item), 4)
                        if raw_score is not None and manual_item is not None
                        else None
                    ),
                    "ai_final_difference": (
                        round(raw_score - float(final_item), 4)
                        if raw_score is not None and final_item is not None
                        else None
                    ),
                    "audit_event_count": len(student.get("audit_log", [])),
                }
                records.append(record)

    diagnostics = {
        "schema_version": FEEDBACK_SCHEMA_VERSION,
        "dataset_id": export_id,
        "created_at": _now_iso(),
        "workflow_type": config.workflow_type,
        "project_type": config.project_type,
        "rounds": [],
    }
    for summary in list_round_summaries(config):
        metadata = load_round_metadata(config.id, summary["id"])
        failures = []
        for failure in metadata.get("failures", []):
            number = int(failure.get("team_number", 0) or 0)
            failures.append({
                "virtual_student_id": aliases.get(number, {}).get("student", "미연결"),
                "category": failure.get("category", "unknown"),
                "retryable": bool(failure.get("retryable", False)),
                "error": scrub(failure.get("error", "")),
                "action": scrub(failure.get("action", "")),
            })
        attempts = []
        for attempt in metadata.get("attempts", []):
            attempts.append({
                "status": attempt.get("status", ""),
                "started_at": attempt.get("started_at", ""),
                "finished_at": attempt.get("finished_at", ""),
                "requested_students": [
                    aliases.get(int(number), {}).get("student", "미연결")
                    for number in attempt.get("requested_team_numbers", [])
                ],
                "success_count": attempt.get("success_count", 0),
                "failure_count": attempt.get("failure_count", 0),
            })
        request_plan = metadata.get("request_plan", {})
        diagnostics["rounds"].append({
            "round_id": summary["id"],
            "status": summary.get("status", ""),
            "provider": summary.get("provider", ""),
            "model": summary.get("model", ""),
            "criteria_version": summary.get("criteria_version", 0),
            "criteria_status": summary.get("criteria_status", ""),
            "started_at": summary.get("started_at", ""),
            "finished_at": summary.get("finished_at", ""),
            "target_count": summary.get("target_count", 0),
            "completed_count": summary.get("completed_count", 0),
            "failure_count": summary.get("failure_count", 0),
            "remaining_count": summary.get("remaining_count", 0),
            "attempt_count": summary.get("attempt_count", 0),
            "expected_requests": request_plan.get("expected_requests"),
            "estimated_minutes_range": request_plan.get("estimated_minutes_range"),
            "estimated_cost_range_krw": request_plan.get(
                "estimated_cost_range_krw"
            ),
            "attempts": attempts,
            "failures": failures,
        })

    materials = []
    for file_info in privacy["material_files"]:
        number = file_info["number"]
        if number not in aliases:
            continue
        pages = None
        if scan_pdf_pages and file_info["path"].suffix.lower() == ".pdf":
            try:
                from pypdf import PdfReader
                pages = len(PdfReader(str(file_info["path"])).pages)
            except Exception:
                pages = None
        materials.append({
            "virtual_student_id": aliases[number]["student"],
            "file_type": file_info["type"] or "unknown",
            "extension": file_info["ext"],
            "size_mb": file_info["size_mb"],
            "page_count": pages,
        })

    analysis = build_analysis_summary(dashboard, include_manual=True)
    return {
        "numbers": numbers,
        "privacy": privacy,
        "aliases": aliases,
        "dashboard": dashboard,
        "analysis": analysis,
        "records": records,
        "diagnostics": diagnostics,
        "materials": materials,
        "text_field_count": text_field_count,
    }


def preview_feedback_export(config: ProjectConfig) -> dict:
    data = _collect_export_data(
        config,
        export_id=f"PREVIEW-{uuid.uuid4().hex[:8]}",
        scan_pdf_pages=False,
    )
    analysis = data["analysis"]
    return {
        "student_count": len(data["numbers"]),
        "round_count": len(data["diagnostics"]["rounds"]),
        "record_count": len(data["records"]),
        "manual_count": analysis["manual_count"],
        "approved_count": analysis["approved_count"],
        "material_count": len(data["materials"]),
        "excluded_original_count": len(data["materials"]),
        "text_field_count": data["text_field_count"],
        "manual_measurement": analysis["manual_measurement"],
        "final_measurement": analysis["final_measurement"],
        "mapping_location": "프로젝트 내부 private/feedback_mappings",
        "originals_included": False,
    }


def _style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="245A7A")
    font = Font(name="맑은 고딕", bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    ws.row_dimensions[1].height = 32
    for index in range(1, ws.max_column + 1):
        values = [
            str(ws.cell(row=row, column=index).value or "")
            for row in range(1, min(ws.max_row, 100) + 1)
        ]
        width = min(max(max((
            sum(
                2 if unicodedata.east_asian_width(char) in {"W", "F", "A"} else 1
                for char in value
            )
            for value in values
        ), default=6) + 2, 10), 36)
        ws.column_dimensions[get_column_letter(index)].width = width


def _performance_workbook(data: dict) -> bytes:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "성능 요약"
    summary_ws.append(["지표", "값", "해석"])
    analysis = data["analysis"]
    rows = [
        ("가상 학생 수", len(data["numbers"]), "실제 학생 정보는 로컬 대응표에만 저장"),
        ("채점 회차", len(data["diagnostics"]["rounds"]), "결과가 저장된 전체 회차"),
        ("채점 레코드", len(data["records"]), "학생×회차×항목"),
        ("수동 점수 보유", analysis["manual_count"], analysis["manual_measurement"]),
        (
            "AI-수동 평균 절대차",
            analysis["manual_mae"] if analysis["manual_mae"] is not None else "측정 불가",
            "수동 점수가 있는 학생만 계산",
        ),
        ("최종 확정 학생", analysis["approved_count"], analysis["final_measurement"]),
        (
            "AI-확정 평균 절대차",
            analysis["final_mae"] if analysis["final_mae"] is not None else "측정 불가",
            "교사 최종 확정 학생만 계산",
        ),
        ("원본 파일 포함", "아니오", "PDF·이미지·동영상·음성은 모두 제외"),
    ]
    for row in rows:
        summary_ws.append(row)
    _style_sheet(summary_ws)

    student_ws = wb.create_sheet("학생별")
    student_ws.append([
        "가상 학생", "가상 학년", "가상 반", "가상 번호",
        "AI 평균", "표준편차", "수동 점수", "AI-수동 차이",
        "AI 제안", "최종 확정", "확정 상태", "검토 우선순위",
    ])
    for student in data["dashboard"]["students"]:
        alias = data["aliases"].get(student["team_number"])
        if not alias:
            continue
        decision = student.get("decision") or {}
        student_ws.append([
            alias["student"], alias["grade"], alias["class"], alias["seat"],
            student.get("ai_average"), student.get("std_dev"),
            student.get("manual_score"), student.get("ai_manual_difference"),
            student.get("ai_suggested_score"),
            decision.get("total_score") if decision.get("status") == "approved" else None,
            decision.get("status", "pending"), student.get("priority", "normal"),
        ])
    _style_sheet(student_ws)
    if student_ws.max_row > 1:
        student_ws.conditional_formatting.add(
            f"F2:F{student_ws.max_row}",
            ColorScaleRule(
                start_type="min", start_color="E2F0D9",
                mid_type="percentile", mid_value=50, mid_color="FFF2CC",
                end_type="max", end_color="F4CCCC",
            ),
        )

    item_ws = wb.create_sheet("항목별")
    item_ws.append([
        "가상 항목", "배점", "AI 채점 학생", "AI 평균", "평균 표준편차",
        "수동 비교 학생", "AI-수동 평균 절대차",
        "확정 학생", "AI-확정 평균 절대차",
    ])
    for index, item in enumerate(analysis["items"], 1):
        item_ws.append([
            f"항목 {index:02d}", item["max_score"], item["student_count"],
            item["ai_average"], item["average_std_dev"], item["manual_count"],
            item["manual_mae"] if item["manual_mae"] is not None else "측정 불가",
            item["final_count"],
            item["final_mae"] if item["final_mae"] is not None else "측정 불가",
        ])
    _style_sheet(item_ws)

    round_ws = wb.create_sheet("회차별")
    round_ws.append([
        "회차", "상태", "공급자", "모델", "기준 버전",
        "대상", "완료", "실패", "남음", "시도 횟수", "예상 요청",
    ])
    for item in data["diagnostics"]["rounds"]:
        round_ws.append([
            item["round_id"], item["status"], item["provider"], item["model"],
            item["criteria_version"], item["target_count"], item["completed_count"],
            item["failure_count"], item["remaining_count"], item["attempt_count"],
            item["expected_requests"],
        ])
    _style_sheet(round_ws)
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _csv_bytes(rows: list[dict]) -> bytes:
    stream = io.StringIO(newline="")
    fields = [
        "virtual_student_id", "file_type", "extension", "size_mb", "page_count"
    ]
    writer = csv.DictWriter(stream, fieldnames=fields)
    writer.writeheader()
    writer.writerows(rows)
    return stream.getvalue().encode("utf-8-sig")


def _guide_text(data: dict) -> str:
    analysis = data["analysis"]
    return "\n".join([
        "AI 채점기 개발 피드백 묶음",
        "",
        "이 파일은 사용자가 채점에 이미 입력한 정보만으로 자동 생성되었습니다.",
        "학생·학교·교사·컴퓨터 경로·원본 파일명은 포함하지 않습니다.",
        "가상 학생 ID는 이 묶음 안에서만 일관되며 다음 내보내기에서는 바뀝니다.",
        "실제 학생과의 대응표는 ZIP 밖의 프로젝트 private 폴더에 따로 저장됩니다.",
        "",
        "포함:",
        "- 성능요약.xlsx: 학생·항목·회차 단위의 점수 신뢰도 요약",
        "- 채점데이터.jsonl: 익명 점수·근거·확신도 레코드",
        "- 실행진단.json: 모델·회차·오류 유형·시도 횟수",
        "- 자료목록.csv: 익명화된 파일 형식·크기·PDF 쪽 수",
        "",
        "제외:",
        "- 학생 답안 PDF, 이미지, 동영상, 음성 원본",
        "- 문제지·정답지·루브릭 원본",
        "- 실제 이름·학년·반·번호·학번과 로컬 경로",
        "- API 키와 로그인 정보",
        "",
        f"수동 채점 비교: {analysis['manual_measurement']}",
        f"최종 확정 비교: {analysis['final_measurement']}",
        "",
        "주의: 자동 비식별화는 알려진 값과 일반 개인정보 패턴을 검사하지만,",
        "개발자에게 보내기 전 안내 파일과 표의 내용을 한 번 확인하는 것을 권장합니다.",
    ])


def _forbidden_tokens(data: dict) -> list[str]:
    tokens = []
    generic_labels = {
        "학생", "참가자", "팀", "보고서", "시험", "평가", "프로젝트",
    }
    for value in data["privacy"]["identifiers"]:
        stripped = str(value).strip()
        # 한 글자 값과 단순 숫자는 일반 문장·점수와 충돌하므로 패턴 검사에서 제외한다.
        if (
            len(stripped) >= 2
            and not stripped.isdigit()
            and stripped not in generic_labels
        ):
            tokens.append(stripped)
    return sorted(set(tokens), key=len, reverse=True)


def _validate_text(text: str, *, forbidden: list[str], location: str) -> list[str]:
    errors = []
    for token in forbidden:
        if token and token.casefold() in text.casefold():
            errors.append(f"{location}: 알려진 개인정보 또는 원본 식별자 발견")
            break
    for label, pattern in (
        ("이메일", EMAIL_PATTERN),
        ("전화번호", PHONE_PATTERN),
        ("주소", ADDRESS_PATTERN),
        ("계정번호", ACCOUNT_PATTERN),
        ("절대경로", WINDOWS_PATH_PATTERN),
        ("인증정보", API_KEY_PATTERN),
        ("학교명", SCHOOL_PATTERN),
    ):
        if pattern.search(text):
            errors.append(f"{location}: {label} 패턴 발견")
    return errors


def validate_feedback_bundle(path: Path, data: dict) -> None:
    forbidden = _forbidden_tokens(data)
    errors = []
    with zipfile.ZipFile(path) as bundle:
        names = set(bundle.namelist())
        if names != BUNDLE_FILENAMES:
            errors.append("ZIP 구성 파일이 허용목록과 다릅니다.")
        for name in sorted(names & TEXT_ENTRY_NAMES):
            text = bundle.read(name).decode("utf-8-sig", errors="replace")
            errors.extend(_validate_text(text, forbidden=forbidden, location=name))
        if "성능요약.xlsx" in names:
            with zipfile.ZipFile(io.BytesIO(bundle.read("성능요약.xlsx"))) as workbook:
                for member in workbook.namelist():
                    if not member.endswith((".xml", ".rels")):
                        continue
                    text = workbook.read(member).decode("utf-8", errors="replace")
                    errors.extend(_validate_text(
                        text,
                        forbidden=forbidden,
                        location=f"성능요약.xlsx/{member}",
                    ))
    if errors:
        raise FeedbackExportError(
            "개인정보 자동 검사에 통과하지 못해 묶음을 저장하지 않았습니다. "
            + " / ".join(sorted(set(errors))[:5])
        )


def create_feedback_bundle(config: ProjectConfig) -> tuple[Path, dict]:
    export_id = f"FD-{uuid.uuid4().hex[:12].upper()}"
    data = _collect_export_data(config, export_id=export_id)
    export_dir = _project_dir(config.id) / "exports" / "feedback"
    mapping_dir = _project_dir(config.id) / "private" / "feedback_mappings"
    export_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = export_dir / f"개발용_성능자료_{timestamp}_{export_id[-4:]}.zip"
    temporary = path.with_suffix(".zip.tmp")
    try:
        with zipfile.ZipFile(
            temporary, "w", compression=zipfile.ZIP_DEFLATED
        ) as bundle:
            bundle.writestr("성능요약.xlsx", _performance_workbook(data))
            bundle.writestr(
                "채점데이터.jsonl",
                "\n".join(
                    json.dumps(record, ensure_ascii=False, sort_keys=True)
                    for record in data["records"]
                ).encode("utf-8"),
            )
            bundle.writestr(
                "실행진단.json",
                json.dumps(
                    data["diagnostics"], ensure_ascii=False, indent=2
                ).encode("utf-8"),
            )
            bundle.writestr("자료목록.csv", _csv_bytes(data["materials"]))
            bundle.writestr("안내.txt", _guide_text(data).encode("utf-8"))
        validate_feedback_bundle(temporary, data)
        temporary.replace(path)
    except Exception:
        temporary.unlink(missing_ok=True)
        path.unlink(missing_ok=True)
        raise

    mapping_dir.mkdir(parents=True, exist_ok=True)
    mapping_path = mapping_dir / f"{export_id}.json"
    roster = data["privacy"]["roster"]
    mapping_payload = {
        "schema_version": FEEDBACK_SCHEMA_VERSION,
        "export_id": export_id,
        "created_at": _now_iso(),
        "bundle_filename": path.name,
        "students": [
            {
                "actual_number": number,
                "actual_name": (
                    roster[number].name
                    if number in roster
                    else next(
                        iter(data["privacy"]["names_by_number"].get(number, [])),
                        "",
                    )
                ),
                "actual_grade": roster[number].grade if number in roster else "",
                "actual_class": (
                    roster[number].class_name if number in roster else ""
                ),
                "actual_student_id": (
                    roster[number].student_id if number in roster else ""
                ),
                **data["aliases"][number],
            }
            for number in data["numbers"]
        ],
    }
    mapping_temporary = mapping_path.with_suffix(".json.tmp")
    try:
        mapping_temporary.write_text(
            json.dumps(mapping_payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        mapping_temporary.replace(mapping_path)
    except Exception:
        mapping_temporary.unlink(missing_ok=True)
        path.unlink(missing_ok=True)
        raise
    return path, {
        "export_id": export_id,
        "mapping_path": mapping_path,
        "student_count": len(data["numbers"]),
        "record_count": len(data["records"]),
    }
