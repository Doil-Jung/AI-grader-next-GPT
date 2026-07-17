"""대회·탐구 심사의 원 평가점수와 최종 순위점수를 분리해 관리한다."""

from __future__ import annotations

import hashlib
import json
import math
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models.project import ProjectConfig
from services import grading as grading_service
from services.grading import load_completed


COMPETITION_SCHEMA_VERSION = 1


def _now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def _state_path(project_id: str) -> Path:
    return grading_service.PROJECTS_DIR / project_id / "competition_ranking.json"


def _atomic_write(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporary.replace(path)


def valid_total_scores(config: ProjectConfig) -> list[float]:
    """루브릭 척도 조합으로 만들 수 있는 총점을 큰 순서로 반환한다."""
    possible = {0.0}
    for criterion in config.all_criteria:
        possible = {
            round(previous + float(score), 4)
            for previous in possible
            for score in criterion.scale
        }
    return sorted(possible, reverse=True)


def _number(value, *, allow_none: bool = False) -> float | None:
    if value in (None, "") and allow_none:
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        if allow_none:
            return None
        raise ValueError("점수는 숫자로 입력하세요.")
    if not math.isfinite(result):
        raise ValueError("점수는 유한한 숫자여야 합니다.")
    return round(result, 4)


def _normalize_allowed_scores(
    config: ProjectConfig,
    values: list | None,
) -> list[float]:
    defaults = valid_total_scores(config)
    if not defaults:
        raise ValueError("대회 최종 점수를 배정하려면 평가기준과 점수 척도가 필요합니다.")
    maximum = max(defaults)
    if values:
        normalized = sorted({
            _number(value)
            for value in values
        }, reverse=True)
        if any(value < 0 or value > maximum for value in normalized):
            raise ValueError(f"허용 점수는 0~{maximum:g}점 범위여야 합니다.")
        if not normalized:
            raise ValueError("허용 점수를 하나 이상 입력하세요.")
        return normalized
    return defaults


def _distribute(group_count: int, allowed_scores: list[float]) -> list[float]:
    """순위 그룹 수에 맞춰 허용 점수 전체 범위를 결정적으로 배분한다."""
    if group_count <= 0:
        return []
    if group_count == 1:
        return [allowed_scores[0]]
    if len(allowed_scores) == 1:
        return [allowed_scores[0]] * group_count
    if group_count <= len(allowed_scores):
        indexes = [
            round(index * (len(allowed_scores) - 1) / (group_count - 1))
            for index in range(group_count)
        ]
        return [allowed_scores[index] for index in indexes]

    # 허용 점수보다 순위 그룹이 많으면 전체 범위를 고르게 사용하고 같은 점수를 허용한다.
    return [
        allowed_scores[
            round(index * (len(allowed_scores) - 1) / (group_count - 1))
        ]
        for index in range(group_count)
    ]


def _source_fingerprint(config: ProjectConfig, round_id: int, completed: dict) -> str:
    payload = []
    for number, result in sorted(completed.items()):
        payload.append({
            "team_number": int(number),
            "team_name": result.get("team_name", ""),
            "total_score": result.get("total_score"),
            "items": {
                criterion.id: result.get(criterion.id)
                for criterion in config.all_criteria
            },
        })
    encoded = json.dumps(
        {"round": round_id, "results": payload},
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def build_competition_plan(
    config: ProjectConfig,
    *,
    source_round: int,
    teams: list[dict],
    allowed_scores: list | None = None,
) -> dict:
    if config.workflow_type != "competition":
        raise ValueError("대회·탐구 심사 프로젝트에서만 순위를 조정할 수 있습니다.")
    completed = load_completed(config.id, int(source_round))
    if not completed:
        raise ValueError(f"{source_round}회차의 평가 결과가 없습니다.")
    if not teams:
        raise ValueError("순위를 조정할 팀 목록이 없습니다.")

    source_numbers = {int(number) for number in completed}
    ordered_numbers = []
    normalized_teams = []
    for index, item in enumerate(teams):
        try:
            number = int(item.get("team_number"))
        except (TypeError, ValueError):
            raise ValueError(f"{index + 1}번째 팀 번호가 올바르지 않습니다.")
        if number in ordered_numbers:
            raise ValueError(f"{number}번 팀이 두 번 포함되었습니다.")
        if number not in source_numbers:
            raise ValueError(f"{number}번 팀은 {source_round}회차 결과에 없습니다.")
        ordered_numbers.append(number)
        normalized_teams.append({
            "team_number": number,
            "tie_with_previous": bool(item.get("tie_with_previous")) and index > 0,
            "override_score": _number(item.get("override_score"), allow_none=True),
            "exception_reason": str(item.get("exception_reason", "")).strip(),
        })
    if set(ordered_numbers) != source_numbers:
        missing = sorted(source_numbers - set(ordered_numbers))
        raise ValueError(
            "평가 결과가 있는 모든 팀을 포함해야 합니다. 누락: "
            + ", ".join(f"{number}번" for number in missing)
        )

    normalized_allowed = _normalize_allowed_scores(config, allowed_scores)
    maximum = max(normalized_allowed)
    groups: list[list[dict]] = []
    for item in normalized_teams:
        if item["tie_with_previous"] and groups:
            groups[-1].append(item)
        else:
            groups.append([item])
    auto_scores = _distribute(len(groups), normalized_allowed)

    entries = []
    previous_final = None
    preceding_count = 0
    for group_index, members in enumerate(groups):
        rank = preceding_count + 1
        preceding_count += len(members)
        overrides = {
            member["override_score"]
            for member in members
            if member["override_score"] is not None
        }
        if len(overrides) > 1:
            raise ValueError(f"공동 {rank}위 팀의 예외 점수는 같아야 합니다.")
        auto_score = auto_scores[group_index]
        final_score = next(iter(overrides), auto_score)
        if final_score < 0 or final_score > maximum:
            raise ValueError(f"최종 점수는 0~{maximum:g}점 범위여야 합니다.")
        reasons = [
            member["exception_reason"]
            for member in members
            if member["exception_reason"]
        ]
        manual_exception = bool(overrides and abs(final_score - auto_score) > 0.001)
        if manual_exception and not reasons:
            raise ValueError(
                f"{rank}위의 자동 배정 점수를 바꾸려면 예외 사유를 입력하세요."
            )
        if previous_final is not None and final_score > previous_final:
            raise ValueError(
                f"{rank}위 최종 점수는 앞 순위의 {previous_final:g}점보다 "
                "높을 수 없습니다."
            )
        previous_final = final_score
        reason = " / ".join(dict.fromkeys(reasons))
        for member in members:
            source = completed[member["team_number"]]
            evaluation_score = _number(source.get("total_score"), allow_none=True)
            entries.append({
                "rank": rank,
                "tie": len(members) > 1,
                "team_number": member["team_number"],
                "team_name": source.get(
                    "team_name", f"참가자 {member['team_number']}"
                ),
                "evaluation_score": evaluation_score,
                "evaluation_items": {
                    criterion.id: _number(
                        source.get(criterion.id), allow_none=True
                    )
                    for criterion in config.all_criteria
                },
                "auto_final_score": auto_score,
                "final_score": final_score,
                "score_change": (
                    round(final_score - evaluation_score, 4)
                    if evaluation_score is not None
                    else None
                ),
                "manual_exception": manual_exception,
                "exception_reason": reason,
            })

    return {
        "schema_version": COMPETITION_SCHEMA_VERSION,
        "source_round": int(source_round),
        "source_fingerprint": _source_fingerprint(
            config, int(source_round), completed
        ),
        "allowed_scores": normalized_allowed,
        "entries": entries,
        "summary": {
            "team_count": len(entries),
            "rank_group_count": len(groups),
            "tie_team_count": sum(1 for entry in entries if entry["tie"]),
            "manual_exception_count": sum(
                1 for entry in entries if entry["manual_exception"]
            ),
            "changed_score_count": sum(
                1
                for entry in entries
                if entry["score_change"] is not None
                and abs(entry["score_change"]) > 0.001
            ),
        },
    }


def load_competition_state(project_id: str) -> dict:
    path = _state_path(project_id)
    if not path.exists():
        return {
            "schema_version": COMPETITION_SCHEMA_VERSION,
            "updated_at": "",
            "current": None,
            "history": [],
        }
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        value = {}
    if not isinstance(value, dict):
        value = {}
    value.setdefault("schema_version", COMPETITION_SCHEMA_VERSION)
    value.setdefault("updated_at", "")
    value.setdefault("current", None)
    value.setdefault("history", [])
    return value


def competition_state_view(config: ProjectConfig) -> dict:
    state = load_competition_state(config.id)
    current = state.get("current")
    stale = False
    if current:
        completed = load_completed(config.id, int(current["source_round"]))
        stale = (
            not completed
            or current.get("source_fingerprint")
            != _source_fingerprint(
                config, int(current["source_round"]), completed
            )
        )
    return {
        **state,
        "current_stale": stale,
        "default_allowed_scores": valid_total_scores(config),
    }


def approve_competition_plan(
    config: ProjectConfig,
    *,
    source_round: int,
    teams: list[dict],
    allowed_scores: list | None,
    approval_note: str = "",
) -> dict:
    plan = build_competition_plan(
        config,
        source_round=source_round,
        teams=teams,
        allowed_scores=allowed_scores,
    )
    state = load_competition_state(config.id)
    version = max(
        [int(item.get("version", 0)) for item in state["history"]] or [0]
    ) + 1
    approved_at = _now_iso()
    current = {
        **plan,
        "status": "approved",
        "version": version,
        "approved_at": approved_at,
        "approval_note": str(approval_note or "").strip(),
    }
    state["schema_version"] = COMPETITION_SCHEMA_VERSION
    state["updated_at"] = approved_at
    state["current"] = current
    state["history"].append({
        "version": version,
        "approved_at": approved_at,
        "source_round": plan["source_round"],
        "source_fingerprint": plan["source_fingerprint"],
        "allowed_scores": plan["allowed_scores"],
        "approval_note": current["approval_note"],
        "summary": plan["summary"],
        "entries": [
            {
                "rank": entry["rank"],
                "tie": entry["tie"],
                "team_number": entry["team_number"],
                "team_name": entry["team_name"],
                "evaluation_score": entry["evaluation_score"],
                "final_score": entry["final_score"],
                "score_change": entry["score_change"],
                "manual_exception": entry["manual_exception"],
                "exception_reason": entry["exception_reason"],
            }
            for entry in plan["entries"]
        ],
    })
    _atomic_write(_state_path(config.id), state)
    return current


def _style_sheet(ws, *, text_columns: set[int] | None = None) -> None:
    text_columns = text_columns or set()
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill("solid", fgColor="5A3D7A")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(name="맑은 고딕", bold=True, color="FFFFFF")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    ws.row_dimensions[1].height = 32
    for index in range(1, ws.max_column + 1):
        values = [
            str(ws.cell(row=row, column=index).value or "")
            for row in range(1, min(ws.max_row, 100) + 1)
        ]
        width = min(max(max((len(value) for value in values), default=6) + 3, 10), 36)
        ws.column_dimensions[get_column_letter(index)].width = width
        if index in text_columns:
            for cell in ws[get_column_letter(index)]:
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )


def build_competition_workbook(
    config: ProjectConfig,
    current: dict,
    history: list[dict],
) -> Workbook:
    if not current or current.get("status") != "approved":
        raise ValueError("교사가 확정한 순위 조정안이 없습니다.")
    wb = Workbook()
    final_ws = wb.active
    final_ws.title = "최종 순위표"
    final_ws.append([
        "최종 순위", "공동 순위", "번호", "이름",
        "원 평가점수", "최종 배정점수", "점수 변화",
        "자동 배정점수", "수동 예외", "예외 사유",
        "원 평가 회차", "확정 버전", "확정 시각",
    ])
    for entry in current["entries"]:
        final_ws.append([
            entry["rank"], "공동" if entry["tie"] else "",
            entry["team_number"], entry["team_name"],
            entry["evaluation_score"], entry["final_score"],
            entry["score_change"], entry["auto_final_score"],
            "예외" if entry["manual_exception"] else "",
            entry["exception_reason"], current["source_round"],
            current["version"], current["approved_at"],
        ])
    _style_sheet(final_ws, text_columns={10})

    source_ws = wb.create_sheet("원 평가점수")
    source_ws.append([
        "번호", "이름",
        *[criterion.name for criterion in config.all_criteria],
        "원 평가 총점", "원 평가 회차",
    ])
    for entry in sorted(current["entries"], key=lambda item: item["team_number"]):
        source_ws.append([
            entry["team_number"], entry["team_name"],
            *[
                entry["evaluation_items"].get(criterion.id)
                for criterion in config.all_criteria
            ],
            entry["evaluation_score"], current["source_round"],
        ])
    _style_sheet(source_ws)

    history_ws = wb.create_sheet("변경 이력")
    history_ws.append([
        "버전", "확정 시각", "원 평가 회차", "팀 수",
        "순위 그룹", "공동순위 팀", "수동 예외", "점수 변경",
        "확정 메모",
    ])
    for item in history:
        summary = item.get("summary") or {}
        history_ws.append([
            item.get("version"), item.get("approved_at"),
            item.get("source_round"), summary.get("team_count"),
            summary.get("rank_group_count"), summary.get("tie_team_count"),
            summary.get("manual_exception_count"),
            summary.get("changed_score_count"),
            item.get("approval_note", ""),
        ])
    _style_sheet(history_ws, text_columns={9})
    return wb
