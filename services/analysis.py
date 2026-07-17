"""채점 결과의 신뢰도 요약과 최종 확정 점수 중심 Excel을 만든다."""

from __future__ import annotations

import statistics
import unicodedata
from collections.abc import Callable

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models.project import ProjectConfig


HEADER_FILL = PatternFill("solid", fgColor="245A7A")
SUBHEADER_FILL = PatternFill("solid", fgColor="DCEAF2")
TITLE_FILL = PatternFill("solid", fgColor="EAF3F8")
APPROVED_FILL = PatternFill("solid", fgColor="E2F0D9")
PENDING_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="맑은 고딕", size=10)
SECTION_FONT = Font(name="맑은 고딕", bold=True, color="163B50", size=12)
THIN_GRAY = Side(style="thin", color="D9E1E5")
BOTTOM_BORDER = Border(bottom=THIN_GRAY)
CENTER = Alignment(horizontal="center", vertical="center")
HEADER_CENTER = Alignment(
    horizontal="center", vertical="center", wrap_text=True
)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _mean(values: list[float]) -> float | None:
    return round(statistics.fmean(values), 2) if values else None


def _mae(values: list[float]) -> float | None:
    return round(statistics.fmean(abs(value) for value in values), 2) if values else None


def build_analysis_summary(dashboard: dict, *, include_manual: bool) -> dict:
    """화면과 파일이 같은 기준으로 읽는 총점·문항별 분석을 반환한다."""
    students = dashboard.get("students", [])
    manual_differences = [
        student["ai_manual_difference"]
        for student in students
        if include_manual and student.get("ai_manual_difference") is not None
    ]
    final_differences = []
    for student in students:
        decision = student.get("decision") or {}
        ai_average = student.get("ai_average")
        if decision.get("status") == "approved" and ai_average is not None:
            final_differences.append(ai_average - float(decision["total_score"]))

    item_rows = []
    item_definitions: dict[str, dict] = {}
    for student in students:
        for item in student.get("items", []):
            aggregate = item_definitions.setdefault(item["id"], {
                "id": item["id"],
                "label": item["label"],
                "max_score": item["max_score"],
                "ai_averages": [],
                "std_devs": [],
                "manual_differences": [],
                "final_differences": [],
                "manual_count": 0,
                "final_count": 0,
            })
            if item.get("ai_average") is not None:
                aggregate["ai_averages"].append(float(item["ai_average"]))
                aggregate["std_devs"].append(float(item.get("std_dev") or 0))
            if include_manual and item.get("manual_score") is not None:
                aggregate["manual_count"] += 1
                if item.get("ai_manual_difference") is not None:
                    aggregate["manual_differences"].append(
                        float(item["ai_manual_difference"])
                    )
            if item.get("final_score") is not None:
                aggregate["final_count"] += 1
                if item.get("ai_average") is not None:
                    aggregate["final_differences"].append(
                        float(item["ai_average"]) - float(item["final_score"])
                    )

    for aggregate in item_definitions.values():
        item_rows.append({
            "id": aggregate["id"],
            "label": aggregate["label"],
            "max_score": aggregate["max_score"],
            "student_count": len(aggregate["ai_averages"]),
            "ai_average": _mean(aggregate["ai_averages"]),
            "average_std_dev": _mean(aggregate["std_devs"]),
            "manual_count": aggregate["manual_count"],
            "manual_mae": _mae(aggregate["manual_differences"]),
            "final_count": aggregate["final_count"],
            "final_mae": _mae(aggregate["final_differences"]),
        })

    summary = dashboard.get("summary", {})
    return {
        "student_count": len(students),
        "round_count": len(dashboard.get("selected_round_ids", [])),
        "manual_count": summary.get("manual_count", 0) if include_manual else 0,
        "approved_count": summary.get(
            "effective_approved_count", summary.get("approved_count", 0)
        ),
        "attention_count": summary.get("attention_count", 0),
        "manual_mae": _mae(manual_differences),
        "final_mae": _mae(final_differences),
        "manual_measurement": (
            f"{len(manual_differences)}명 비교"
            if manual_differences
            else "측정 불가"
        ),
        "final_measurement": (
            f"{len(final_differences)}명 비교"
            if final_differences
            else "측정 불가"
        ),
        "items": item_rows,
    }


def build_analysis_rows(
    dashboard: dict,
    *,
    include_manual: bool,
) -> list[dict]:
    results = []
    for student in dashboard.get("students", []):
        scores = {
            f"{round_id}회": score
            for round_id, score in student.get("scores_by_round", {}).items()
        }
        calc_scores = list(scores.values())
        manual_score = student.get("manual_score") if include_manual else None
        if not calc_scores and manual_score is None:
            continue
        trimmed_average = student.get("ai_average")
        is_trimmed = False
        if len(calc_scores) >= 5:
            trimmed_list = sorted(calc_scores)[1:-1]
            trimmed_average = round(sum(trimmed_list) / len(trimmed_list), 2)
            is_trimmed = True
        decision = student.get("decision") or {}
        results.append({
            "team_number": student["team_number"],
            "team_name": student["team_name"],
            "manual_score": manual_score,
            "scores_by_round": scores,
            "score_count": len(calc_scores),
            "average": student.get("ai_average"),
            "median": student.get("ai_median"),
            "trimmed_average": trimmed_average,
            "is_trimmed": is_trimmed,
            "std_dev": student.get("std_dev", 0),
            "ai_manual_difference": (
                student.get("ai_manual_difference") if include_manual else None
            ),
            "ai_suggested_score": student.get("ai_suggested_score"),
            "final_score": (
                decision.get("total_score")
                if decision.get("status") == "approved"
                else None
            ),
            "final_status": decision.get("status", "pending"),
            "final_source": decision.get("decision_source", ""),
            "decision_stale": bool(student.get("decision_stale")),
            "priority": student.get("priority", "normal"),
            "review_reasons": student.get("review_reasons", []),
        })
    results.sort(key=lambda value: (
        value["average"] is None,
        -(value["average"] or 0),
        value["team_number"],
    ))
    return results


def _style_table(ws, *, header_row: int, max_row: int, max_column: int) -> None:
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(max_column)}{max(header_row, max_row)}"
    )
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_CENTER
    ws.row_dimensions[header_row].height = 34
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=max_row,
        min_col=1,
        max_col=max_column,
    ):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BOTTOM_BORDER
            cell.alignment = CENTER


def _fit_columns(ws, *, text_columns: set[int] | None = None) -> None:
    text_columns = text_columns or set()

    def display_width(value: str) -> int:
        return sum(
            2 if unicodedata.east_asian_width(char) in {"W", "F", "A"} else 1
            for char in value
        )

    for index in range(1, ws.max_column + 1):
        values = [
            str(ws.cell(row=row, column=index).value or "")
            for row in range(1, min(ws.max_row, 80) + 1)
        ]
        width = min(
            max(max((display_width(value) for value in values), default=4) + 2, 10),
            38,
        )
        ws.column_dimensions[get_column_letter(index)].width = width
        if index in text_columns:
            for cell in ws[get_column_letter(index)]:
                cell.alignment = LEFT_WRAP


def _final_item_columns(config: ProjectConfig) -> list[dict]:
    if config.project_type != "exam":
        return [
            {
                "kind": "item",
                "id": criterion.id,
                "label": f"확정 {criterion.name}",
            }
            for category in config.categories
            for criterion in category.criteria
        ]
    columns = []
    for question in config.exam.top_level_questions():
        children = config.exam.children_of(question.id)
        if children:
            columns.append({
                "kind": "group",
                "ids": [child.id for child in children],
                "label": f"확정 {question.number}번 합계",
            })
            columns.extend({
                "kind": "item",
                "id": child.id,
                "label": f"확정 {child.number}번",
            } for child in children)
        else:
            columns.append({
                "kind": "item",
                "id": question.id,
                "label": f"확정 {question.number}번",
            })
    return columns


def _final_item_value(decision: dict, column: dict):
    item_scores = decision.get("item_scores") or {}
    if decision.get("status") != "approved":
        return ""
    if column["kind"] == "item":
        return item_scores.get(column["id"], "")
    values = [item_scores.get(item_id) for item_id in column["ids"]]
    return round(sum(values), 4) if values and all(value is not None for value in values) else ""


def build_analysis_workbook(
    config: ProjectConfig,
    dashboard: dict,
    *,
    include_manual: bool,
    load_completed: Callable[[str, int], dict],
) -> Workbook:
    """확정 점수표, 총점 분석, 항목별 분석/비교를 한 파일에 보존한다."""
    rows = []
    for student in dashboard.get("students", []):
        latest_result = {}
        for round_id in reversed(dashboard.get("selected_round_ids", [])):
            candidate = load_completed(config.id, round_id).get(
                student["team_number"], {}
            )
            if candidate:
                latest_result = candidate
                break
        rows.append({
            **student,
            "overall_comment": latest_result.get("overall_comment", ""),
            "seteuk": latest_result.get("seteuk", ""),
        })
    rows.sort(key=lambda value: value["team_number"])
    round_keys = [f"{round_id}회" for round_id in dashboard.get("selected_round_ids", [])]
    summary = build_analysis_summary(dashboard, include_manual=include_manual)

    wb = Workbook()
    final_ws = wb.active
    final_ws.title = "확정 점수표"
    item_columns = _final_item_columns(config)
    final_headers = [
        "번호", "이름", "확정 상태", "최종 확정 총점", "확정 근거",
        "AI 제안점수", "수동 점수", "차이(AI-수동)", "AI 표준편차",
    ] + [column["label"] for column in item_columns] + [
        "검토 사유", "교사 확정 메모", "종합 의견", "세특 초안",
    ]
    final_ws.append(final_headers)
    for student in rows:
        decision = student.get("decision") or {}
        approved = decision.get("status") == "approved"
        values = [
            student["team_number"],
            student["team_name"],
            "최종 확정" if approved else "확정 대기",
            decision.get("total_score", "") if approved else "",
            decision.get("decision_source", "") if approved else "",
            student.get("ai_suggested_score", ""),
            student.get("manual_score", "") if include_manual else "",
            student.get("ai_manual_difference", "") if include_manual else "",
            student.get("std_dev", 0),
        ]
        values.extend(_final_item_value(decision, column) for column in item_columns)
        values.extend([
            " · ".join(reason["label"] for reason in student.get("review_reasons", [])),
            decision.get("teacher_note", "") if approved else "",
            student.get("overall_comment", ""),
            student.get("seteuk", ""),
        ])
        final_ws.append(values)
        final_ws.cell(
            row=final_ws.max_row, column=3
        ).fill = APPROVED_FILL if approved else PENDING_FILL
    _style_table(
        final_ws, header_row=1, max_row=final_ws.max_row, max_column=final_ws.max_column
    )
    _fit_columns(final_ws, text_columns={
        final_ws.max_column - 3,
        final_ws.max_column - 2,
        final_ws.max_column - 1,
        final_ws.max_column,
    })

    total_ws = wb.create_sheet("종합 분석")
    total_headers = [
        "순위", "번호", "이름", *round_keys,
        "AI 평균", "AI 중앙값", "절사평균", "AI 표준편차",
        "수동채점", "차이(AI-수동)", "AI 제안점수",
        "교사 확정점수", "확정 상태", "확정 근거", "검토 사유",
        "종합의견", "세특",
    ]
    total_ws.append(total_headers)
    ranked = sorted(rows, key=lambda value: (
        value.get("ai_average") is None,
        -(value.get("ai_average") or 0),
        value["team_number"],
    ))
    for rank, student in enumerate(ranked, 1):
        scores = {
            f"{round_id}회": score
            for round_id, score in student.get("scores_by_round", {}).items()
        }
        score_values = list(scores.values())
        trimmed = student.get("ai_average")
        if len(score_values) >= 5:
            trimmed_values = sorted(score_values)[1:-1]
            trimmed = round(sum(trimmed_values) / len(trimmed_values), 2)
        decision = student.get("decision") or {}
        approved = decision.get("status") == "approved"
        total_ws.append([
            rank, student["team_number"], student["team_name"],
            *[scores.get(key, "") for key in round_keys],
            student.get("ai_average", ""),
            student.get("ai_median", ""),
            trimmed if trimmed is not None else "",
            student.get("std_dev", 0),
            student.get("manual_score", "") if include_manual else "",
            student.get("ai_manual_difference", "") if include_manual else "",
            student.get("ai_suggested_score", ""),
            decision.get("total_score", "") if approved else "",
            "최종 확정" if approved else "확정 대기",
            decision.get("decision_source", "") if approved else "",
            " · ".join(reason["label"] for reason in student.get("review_reasons", [])),
            student.get("overall_comment", ""),
            student.get("seteuk", ""),
        ])
    _style_table(
        total_ws, header_row=1, max_row=total_ws.max_row, max_column=total_ws.max_column
    )
    _fit_columns(total_ws, text_columns={
        total_ws.max_column - 2,
        total_ws.max_column - 1,
        total_ws.max_column,
    })

    detail_ws = wb.create_sheet("항목별 비교")
    detail_headers = [
        "번호", "이름", "항목·문항", "배점", *round_keys,
        "AI 평균", "AI 중앙값", "표준편차", "수동 점수",
        "차이(AI-수동)", "최종 확정",
    ]
    detail_ws.append(detail_headers)
    for student in rows:
        for item in student.get("items", []):
            detail_ws.append([
                student["team_number"], student["team_name"], item["label"],
                item["max_score"],
                *[
                    item.get("scores_by_round", {}).get(int(key[:-1]), "")
                    for key in round_keys
                ],
                item.get("ai_average", ""),
                item.get("ai_median", ""),
                item.get("std_dev", 0),
                item.get("manual_score", "") if include_manual else "",
                item.get("ai_manual_difference", "") if include_manual else "",
                item.get("final_score", ""),
            ])
    _style_table(
        detail_ws, header_row=1, max_row=detail_ws.max_row, max_column=detail_ws.max_column
    )
    _fit_columns(detail_ws, text_columns={3})

    item_ws = wb.create_sheet("문항별 분석")
    item_ws.append([
        "항목·문항", "배점", "AI 채점 학생", "AI 평균",
        "평균 표준편차", "수동 비교 학생", "AI-수동 평균 절대차",
        "확정 학생", "AI-확정 평균 절대차",
    ])
    for item in summary["items"]:
        item_ws.append([
            item["label"], item["max_score"], item["student_count"],
            item["ai_average"], item["average_std_dev"], item["manual_count"],
            item["manual_mae"] if item["manual_mae"] is not None else "측정 불가",
            item["final_count"],
            item["final_mae"] if item["final_mae"] is not None else "측정 불가",
        ])
    _style_table(
        item_ws, header_row=1, max_row=item_ws.max_row, max_column=item_ws.max_column
    )
    if item_ws.max_row > 1:
        item_ws.conditional_formatting.add(
            f"E2:E{item_ws.max_row}",
            ColorScaleRule(
                start_type="min", start_color="E2F0D9",
                mid_type="percentile", mid_value=50, mid_color="FFF2CC",
                end_type="max", end_color="F4CCCC",
            ),
        )
    _fit_columns(item_ws, text_columns={1})
    return wb
