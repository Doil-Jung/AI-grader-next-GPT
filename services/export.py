"""
Excel 내보내기 서비스
- 동적 rubric에 맞는 채점 결과 Excel 생성
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from models.project import ProjectConfig


def generate_excel(config: ProjectConfig, results: list[dict], output_path: Path):
    """채점 결과 Excel 생성 (동적 rubric 기반)"""
    if config.project_type == "exam":
        return _generate_exam_excel(config, results, output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "채점 결과"

    # ─── 스타일 정의 ───
    header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    sub_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    data_font = Font(name="맑은 고딕", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    silver_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    bronze_fill = PatternFill(start_color="FFE4C4", end_color="FFE4C4", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    criteria = config.all_criteria
    categories = config.categories

    # ─── 타이틀 행 ───
    total_cols = 3 + len(criteria) + len(categories) + 3  # 순위, 번호, 이름, 항목들, 소계들, 합계, 종합의견, 세특
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{config.name} 채점 결과"
    title_cell.font = Font(name="맑은 고딕", bold=True, size=14, color="2B579A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ─── 헤더 구성 ───
    # Row 2: 카테고리 병합 헤더
    # Row 3: 세부 항목 헤더

    # 순위, 번호, 이름 (2행~3행 병합)
    fixed_headers = ["순위", "번호", "이름"]
    for i, h in enumerate(fixed_headers, 1):
        ws.merge_cells(start_row=2, start_column=i, end_row=3, end_column=i)
        ws.cell(row=2, column=i, value=h)

    col = 4  # 시작 열
    cat_start_cols = {}
    
    for cat in categories:
        start_col = col
        cat_start_cols[cat.name] = {"start": start_col, "criteria": []}
        
        for c in cat.criteria:
            ws.cell(row=3, column=col, value=f"{c.name}\n({c.max_score})")
            cat_start_cols[cat.name]["criteria"].append(col)
            col += 1
        
        end_col = col - 1
        if start_col <= end_col:
            if start_col < end_col:
                ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
            ws.cell(row=2, column=start_col, value=f"{cat.name} ({cat.max_score})")

    # 카테고리 소계 열
    subtotal_cols = {}
    for cat in categories:
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        ws.cell(row=2, column=col, value=f"{cat.name}\n소계({cat.max_score})")
        subtotal_cols[cat.name] = col
        col += 1

    # 합계 열
    total_col = col
    ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
    ws.cell(row=2, column=col, value=f"합계\n({config.total_max_score})")
    col += 1

    # 종합의견 열
    comment_col = col
    ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
    ws.cell(row=2, column=col, value="종합의견")
    col += 1

    # 세특 열
    seteuk_col = col
    ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
    ws.cell(row=2, column=col, value="세특")
    col += 1

    last_col = col - 1

    # ─── 헤더 스타일 ───
    for row in [2, 3]:
        for c in range(1, last_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill if row == 2 else sub_header_fill
            cell.alignment = center_align
            cell.border = thin_border

    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 40

    # ─── 데이터 행 ───
    for rank, result in enumerate(results, 1):
        row = rank + 3
        
        # 순위, 번호, 이름
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=result.get("team_number", ""))
        ws.cell(row=row, column=3, value=result.get("team_name", ""))
        
        # 세부항목 점수
        c_col = 4
        for cat in categories:
            for criterion in cat.criteria:
                ws.cell(row=row, column=c_col, value=result.get(criterion.id, "-"))
                c_col += 1
        
        # 카테고리 소계
        for cat in categories:
            cat_key = cat.name.replace(" ", "_") + "_total"
            ws.cell(row=row, column=subtotal_cols[cat.name], value=result.get(cat_key, "-"))
        
        # 합계
        ws.cell(row=row, column=total_col, value=result.get("total_score", "-"))
        
        # 종합의견
        ws.cell(row=row, column=comment_col, value=result.get("overall_comment", ""))
        
        # 세특
        ws.cell(row=row, column=seteuk_col, value=result.get("seteuk", ""))
        
        # 스타일 적용
        for c in range(1, last_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = data_font
            cell.border = thin_border
            if c <= 2 or (4 <= c <= last_col - 1):
                cell.alignment = center_align
            else:
                cell.alignment = left_align
        
        # 순위별 색상
        fill = None
        if rank == 1: fill = gold_fill
        elif rank <= 3: fill = silver_fill
        elif rank <= 10: fill = bronze_fill
        elif rank <= 30: fill = green_fill
        if fill:
            for c in range(1, last_col + 1):
                ws.cell(row=row, column=c).fill = fill

    # ─── 열 너비 ───
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 18
    for c in range(4, last_col + 1):
        letter = get_column_letter(c)
        if c == comment_col:
            ws.column_dimensions[letter].width = 45
        elif c == seteuk_col:
            ws.column_dimensions[letter].width = 50
        else:
            ws.column_dimensions[letter].width = 12

    # ─── 채점 근거 시트 ───
    ws2 = wb.create_sheet("채점 근거")
    detail_headers = ["번호", "이름"]
    for c in criteria:
        detail_headers.extend([f"{c.name} 점수", f"{c.name} 근거"])
    detail_headers.append("종합의견")
    detail_headers.append("세특")
    
    for c_idx, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for i, r in enumerate(results):
        row = i + 2
        vals = [r.get("team_number", ""), r.get("team_name", "")]
        for c in criteria:
            vals.append(r.get(c.id, "-"))
            vals.append(r.get(f"{c.id}_reason", ""))
        vals.append(r.get("overall_comment", ""))
        vals.append(r.get("seteuk", ""))
        
        for c_idx, val in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=c_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = left_align if c_idx % 2 == 0 and c_idx >= 4 else center_align

    for c_idx in range(1, len(detail_headers) + 1):
        letter = get_column_letter(c_idx)
        if c_idx <= 2:
            ws2.column_dimensions[letter].width = 15
        elif c_idx % 2 == 1:
            ws2.column_dimensions[letter].width = 12
        else:
            ws2.column_dimensions[letter].width = 40

    wb.save(output_path)


def _generate_exam_excel(config: ProjectConfig, results: list[dict], output_path: Path):
    """교사 승인 상태와 문항별 근거를 포함한 정기고사 채점표."""
    wb = Workbook()
    ws = wb.active
    ws.title = "서술형 채점 결과"

    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    pending_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    approved_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    questions = config.exam.scored_questions()
    headers = ["번호", "이름"] + [f"{q.number}번 ({q.max_score})" for q in questions]
    headers += [f"총점 ({config.total_max_score})", "검토 필요", "교사 승인", "승인 시각", "교사 메모"]
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    for row_index, result in enumerate(sorted(results, key=lambda r: r.get("team_number", 0)), 2):
        values = [result.get("team_number", ""), result.get("team_name", "")]
        values.extend(result.get(q.id, "") for q in questions)
        values.extend([
            result.get("total_score", ""),
            result.get("review_required_count", 0),
            "승인" if result.get("teacher_status") == "approved" else "검토 대기",
            result.get("teacher_approved_at", ""),
            result.get("teacher_note", ""),
        ])
        fill = approved_fill if result.get("teacher_status") == "approved" else pending_fill
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_index, column=col, value=value)
            cell.border = thin_border
            cell.alignment = left if col in (2, len(headers)) else center
            cell.fill = fill

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 16
    for col in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14 if col < len(headers) else 28

    detail = wb.create_sheet("문항별 근거")
    detail_headers = ["번호", "이름", "문항", "점수", "답안 요약", "채점 근거", "확신도", "교사 확인 필요", "승인 상태"]
    for col, title in enumerate(detail_headers, 1):
        cell = detail.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center

    row = 2
    for result in sorted(results, key=lambda r: r.get("team_number", 0)):
        for question in questions:
            values = [
                result.get("team_number", ""), result.get("team_name", ""), question.number,
                result.get(question.id, ""), result.get(f"{question.id}_answer_summary", ""),
                result.get(f"{question.id}_reason", ""), result.get(f"{question.id}_confidence", ""),
                "예" if result.get(f"{question.id}_review_required") else "아니오",
                "승인" if result.get("teacher_status") == "approved" else "검토 대기",
            ]
            for col, value in enumerate(values, 1):
                cell = detail.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = left if col in (5, 6) else center
            row += 1

    widths = [8, 16, 8, 10, 35, 45, 10, 14, 12]
    for col, width in enumerate(widths, 1):
        detail.column_dimensions[get_column_letter(col)].width = width
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    wb.save(output_path)
